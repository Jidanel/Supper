# ===================================================================
# Fichier : accounts/views_import.py - VERSION MISE À JOUR
# Vues d'import Excel pour postes et utilisateurs
# Avec habilitations granulaires et journalisation détaillée
# ===================================================================

import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ValidationError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import logging

from accounts.models import Poste, UtilisateurSUPPER
from common.utils import (
    log_user_action,
    log_erreur_action,
    require_permission,
    require_any_permission,
    require_habilitation,
    get_user_description,
    get_user_short_description,
    get_habilitation_normalisee,
    get_habilitation_label,
    is_admin_user,
    is_service_central,
    format_montant_fcfa,
    HABILITATIONS_ADMIN,
    HABILITATIONS_SERVICES_CENTRAUX,
    HABILITATIONS_LABELS,
)

logger = logging.getLogger('supper')


# ===================================================================
# CONSTANTES ET CONFIGURATIONS
# ===================================================================

# Régions du Cameroun (pour validation)
REGIONS_CAMEROUN = [
    'ADAMAOUA', 'CENTRE', 'EST', 'EXTREME-NORD', 'EXTRÊME-NORD',
    'LITTORAL', 'NORD', 'NORD-OUEST', 'OUEST', 'SUD', 'SUD-OUEST'
]

# Mapping des régions (normalisation)
REGIONS_MAPPING = {
    'ADAMAOUA': 'Adamaoua',
    'CENTRE': 'Centre',
    'EST': 'Est',
    'EXTREME-NORD': 'Extrême-Nord',
    'EXTRÊME-NORD': 'Extrême-Nord',
    'EXTREME NORD': 'Extrême-Nord',
    'LITTORAL': 'Littoral',
    'NORD': 'Nord',
    'NORD-OUEST': 'Nord-Ouest',
    'NORD OUEST': 'Nord-Ouest',
    'OUEST': 'Ouest',
    'SUD': 'Sud',
    'SUD-OUEST': 'Sud-Ouest',
    'SUD OUEST': 'Sud-Ouest',
}

# Types de postes valides
TYPES_POSTES_MAPPING = {
    'PEAGE': 'peage',
    'PÉAGE': 'peage',
    'peage': 'peage',
    'péage': 'peage',
    'PG': 'peage',
    'P': 'peage',
    'PESAGE': 'pesage',
    'pesage': 'pesage',
    'PS': 'pesage',
    'S': 'pesage',
}

# Habilitations valides pour import
HABILITATIONS_IMPORT_MAPPING = {
    # Administrateurs
    'ADMIN_PRINCIPAL': 'admin_principal',
    'ADMIN': 'admin_principal',
    'ADMINISTRATEUR': 'admin_principal',
    'COORD_PSRR': 'coord_psrr',
    'COORDONNATEUR': 'coord_psrr',
    'COORDONNATEUR PSRR': 'coord_psrr',
    'SERV_INFO': 'serv_info',
    'SERVICE INFO': 'serv_info',
    'SERVICE INFORMATIQUE': 'serv_info',
    'INFORMATIQUE': 'serv_info',
    
    # Services centraux
    'SERV_EMISSION': 'serv_emission',
    'SERVICE EMISSION': 'serv_emission',
    'EMISSION': 'serv_emission',
    'CHEF_AG': 'chef_ag',
    'CHEF AG': 'chef_ag',
    'AFFAIRES GENERALES': 'chef_ag',
    'SERV_CONTROLE': 'serv_controle',
    'SERVICE CONTROLE': 'serv_controle',
    'CONTROLE': 'serv_controle',
    'SERV_ORDRE': 'serv_ordre',
    'SERVICE ORDRE': 'serv_ordre',
    'ORDRE': 'serv_ordre',
    
    # CISOP
    'CISOP_PEAGE': 'cisop_peage',
    'CISOP PEAGE': 'cisop_peage',
    'CISOP_PESAGE': 'cisop_pesage',
    'CISOP PESAGE': 'cisop_pesage',
    
    # Chefs de poste
    'CHEF_PEAGE': 'chef_peage',
    'CHEF PEAGE': 'chef_peage',
    'CHEF DE POSTE PEAGE': 'chef_peage',
    'CHEF_STATION_PESAGE': 'chef_station_pesage',
    'CHEF STATION PESAGE': 'chef_station_pesage',
    'CHEF DE STATION PESAGE': 'chef_station_pesage',
    'CHEF PESAGE': 'chef_station_pesage',
    
    # Opérationnels pesage
    'REGISSEUR_PESAGE': 'regisseur_pesage',
    'REGISSEUR PESAGE': 'regisseur_pesage',
    'CHEF_EQUIPE_PESAGE': 'chef_equipe_pesage',
    'CHEF EQUIPE PESAGE': 'chef_equipe_pesage',
    
    # Autres
    'AGENT_INVENTAIRE': 'agent_inventaire',
    'AGENT INVENTAIRE': 'agent_inventaire',
    'INVENTAIRE': 'agent_inventaire',
    'CAISSIER': 'caissier',
    'FOCAL_REGIONAL': 'focal_regional',
    'FOCAL REGIONAL': 'focal_regional',
    'POINT FOCAL': 'focal_regional',
    'CHEF_SERVICE': 'chef_service',
    'CHEF SERVICE': 'chef_service',
    'REGISSEUR': 'regisseur',
    'COMPTABLE_MAT': 'comptable_mat',
    'COMPTABLE MATIERES': 'comptable_mat',
    'IMPRIMERIE': 'imprimerie',
}


# ===================================================================
# VUES D'IMPORT DES POSTES
# ===================================================================

@login_required
@require_any_permission('peut_creer_poste_masse', 'peut_gerer_postes', 'peut_ajouter_poste')
def import_postes_excel(request):
    """
    Vue pour importer des postes depuis un fichier Excel
    
    Permissions requises:
    - peut_creer_poste_masse : Permission principale pour import en masse
    - peut_gerer_postes : Permission générale de gestion des postes
    - peut_ajouter_poste : Permission d'ajout de poste
    """
    
    user_desc = get_user_description(request.user)
    
    if request.method == 'POST':
        fichier = request.FILES.get('fichier_excel')
        action_doublon = request.POST.get('action_doublon', 'sauter')
        
        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier Excel")
            log_user_action(
                request.user,
                "IMPORT_POSTES_ECHEC",
                "Tentative d'import sans fichier",
                request
            )
            return redirect('accounts:import_postes')
        
        if not fichier.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Le fichier doit être au format Excel (.xlsx ou .xls)")
            log_user_action(
                request.user,
                "IMPORT_POSTES_ECHEC",
                f"Format de fichier invalide: {fichier.name}",
                request
            )
            return redirect('accounts:import_postes')
        
        try:
            # Journaliser le début de l'import
            log_user_action(
                request.user,
                "IMPORT_POSTES_DEBUT",
                f"Début d'import de postes depuis le fichier {fichier.name}",
                request,
                fichier=fichier.name,
                taille=f"{fichier.size / 1024:.1f} Ko",
                action_doublon=action_doublon
            )
            
            # Lire le fichier Excel
            df = pd.read_excel(fichier)
            
            # Traiter l'import
            resultats = traiter_import_postes(df, action_doublon, request.user)
            
            # Journaliser les résultats
            if resultats['success']:
                messages.success(
                    request,
                    f"Import terminé ! {resultats['nb_crees']} postes créés, "
                    f"{resultats['nb_modifies']} modifiés, {resultats['nb_sautes']} sautés, "
                    f"{resultats['nb_erreurs']} erreurs"
                )
                
                log_user_action(
                    request.user,
                    "IMPORT_POSTES_SUCCES",
                    f"Import de postes terminé avec succès",
                    request,
                    fichier=fichier.name,
                    nb_crees=resultats['nb_crees'],
                    nb_modifies=resultats['nb_modifies'],
                    nb_sautes=resultats['nb_sautes'],
                    nb_erreurs=resultats['nb_erreurs']
                )
                
                logger.info(
                    f"📥 IMPORT POSTES RÉUSSI | {get_user_short_description(request.user)} | "
                    f"Créés: {resultats['nb_crees']}, Modifiés: {resultats['nb_modifies']}, "
                    f"Sautés: {resultats['nb_sautes']}, Erreurs: {resultats['nb_erreurs']}"
                )
            else:
                messages.error(request, f"Erreur lors de l'import: {resultats['erreur']}")
                
                log_erreur_action(
                    request.user,
                    "IMPORT_POSTES",
                    resultats['erreur'],
                    request
                )
            
            context = {
                'resultats': resultats,
                'fichier_nom': fichier.name,
                'user_desc': user_desc,
            }
            return render(request, 'accounts/import_postes_resultats.html', context)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier: {str(e)}")
            log_erreur_action(request.user, "IMPORT_POSTES", str(e), request)
            return redirect('accounts:import_postes')
    
    # GET request - afficher le formulaire
    context = {
        'user_desc': user_desc,
        'regions_disponibles': list(REGIONS_MAPPING.values()),
        'types_postes': ['peage', 'pesage'],
    }
    return render(request, 'accounts/import_postes_form.html', context)


def traiter_import_postes(df, action_doublon, user):
    """
    Traite l'import des postes depuis Excel avec validation complète
    
    Args:
        df: DataFrame pandas avec les données
        action_doublon: 'sauter' ou 'ecraser'
        user: Utilisateur effectuant l'import
    
    Returns:
        dict: Résultats de l'import
    """
    
    # Normaliser les noms de colonnes
    df.columns = df.columns.str.strip().str.lower()
    
    nb_crees = 0
    nb_modifies = 0
    nb_sautes = 0
    nb_erreurs = 0
    erreurs_detail = []
    postes_crees = []
    postes_modifies = []
    regions_non_trouvees = set()
    types_invalides = set()
    
    # Vérifier les colonnes requises
    colonnes_requises = ['nom', 'region', 'code', 'type']
    colonnes_manquantes = [col for col in colonnes_requises if col not in df.columns]
    
    if colonnes_manquantes:
        return {
            'success': False,
            'erreur': f"Colonnes manquantes dans le fichier: {', '.join(colonnes_manquantes)}. "
                      f"Colonnes attendues: {', '.join(colonnes_requises)}",
            'nb_crees': 0,
            'nb_modifies': 0,
            'nb_sautes': 0,
            'nb_erreurs': 0
        }
    
    # Colonnes optionnelles
    colonnes_optionnelles = ['departement', 'axe_routier', 'latitude', 'longitude', 'description']
    
    try:
        with transaction.atomic():
            for index, row in df.iterrows():
                ligne_num = index + 2  # +2 car Excel commence à 1 et il y a l'en-tête
                
                try:
                    # Extraire et nettoyer les données obligatoires
                    nom_poste = str(row['nom']).strip() if pd.notna(row['nom']) else ''
                    code_poste = str(row['code']).strip().upper() if pd.notna(row['code']) else ''
                    nom_region = str(row['region']).strip().upper() if pd.notna(row['region']) else ''
                    type_poste_raw = str(row['type']).strip() if pd.notna(row['type']) else ''
                    
                    # Vérifier données non vides
                    if not nom_poste or nom_poste.lower() == 'nan':
                        continue  # Ligne vide, ignorer silencieusement
                    
                    if not code_poste or code_poste.lower() == 'nan':
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Code manquant pour le poste '{nom_poste}'",
                            'type': 'code_manquant'
                        })
                        nb_erreurs += 1
                        continue
                    
                    if not nom_region or nom_region.lower() == 'nan':
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Région manquante pour le poste '{nom_poste}'",
                            'type': 'region_manquante'
                        })
                        nb_erreurs += 1
                        continue
                    
                    if not type_poste_raw or type_poste_raw.lower() == 'nan':
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Type manquant pour le poste '{nom_poste}'",
                            'type': 'type_manquant'
                        })
                        nb_erreurs += 1
                        continue
                    
                    # Normaliser le type de poste
                    type_poste = TYPES_POSTES_MAPPING.get(type_poste_raw.upper())
                    
                    if not type_poste:
                        types_invalides.add(type_poste_raw)
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Type '{type_poste_raw}' invalide pour '{nom_poste}'. "
                                      f"Types acceptés: peage, pesage",
                            'type': 'type_invalide'
                        })
                        nb_erreurs += 1
                        continue
                    
                    # Normaliser la région
                    region_normalisee = REGIONS_MAPPING.get(nom_region)
                    
                    if not region_normalisee:
                        # Essayer une correspondance partielle
                        region_normalisee = _trouver_region_proche(nom_region)
                    
                    if not region_normalisee:
                        regions_non_trouvees.add(nom_region)
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Région '{nom_region}' introuvable pour '{nom_poste}'",
                            'type': 'region_invalide'
                        })
                        nb_erreurs += 1
                        continue
                    
                    # Extraire les données optionnelles
                    departement = ''
                    axe_routier = ''
                    latitude = None
                    longitude = None
                    description = ''
                    
                    if 'departement' in df.columns and pd.notna(row.get('departement')):
                        departement = str(row['departement']).strip()
                    
                    if 'axe_routier' in df.columns and pd.notna(row.get('axe_routier')):
                        axe_routier = str(row['axe_routier']).strip()
                    
                    if 'latitude' in df.columns and pd.notna(row.get('latitude')):
                        try:
                            latitude = float(row['latitude'])
                        except (ValueError, TypeError):
                            pass
                    
                    if 'longitude' in df.columns and pd.notna(row.get('longitude')):
                        try:
                            longitude = float(row['longitude'])
                        except (ValueError, TypeError):
                            pass
                    
                    if 'description' in df.columns and pd.notna(row.get('description')):
                        description = str(row['description']).strip()
                    
                    # Vérifier si le poste existe déjà
                    poste_existant = Poste.objects.filter(code=code_poste).first()
                    
                    if poste_existant:
                        if action_doublon == 'ecraser':
                            # Mise à jour du poste existant
                            poste_existant.nom = nom_poste
                            poste_existant.region = region_normalisee
                            poste_existant.type = type_poste
                            
                            if departement:
                                poste_existant.departement = departement
                            if axe_routier:
                                poste_existant.axe_routier = axe_routier
                            if latitude is not None:
                                poste_existant.latitude = latitude
                            if longitude is not None:
                                poste_existant.longitude = longitude
                            if description:
                                poste_existant.description = description
                            
                            poste_existant.save()
                            postes_modifies.append({
                                'code': code_poste,
                                'nom': nom_poste,
                                'type': type_poste
                            })
                            nb_modifies += 1
                        else:
                            # Sauter
                            nb_sautes += 1
                    else:
                        # Créer nouveau poste
                        nouveau_poste = Poste.objects.create(
                            code=code_poste,
                            nom=nom_poste,
                            region=region_normalisee,
                            type=type_poste,
                            departement=departement,
                            axe_routier=axe_routier,
                            latitude=latitude,
                            longitude=longitude,
                            description=description,
                            is_active=True,
                            nouveau=True  # Marquer comme nouveau pour suivi
                        )
                        postes_crees.append({
                            'code': code_poste,
                            'nom': nom_poste,
                            'type': type_poste,
                            'region': region_normalisee
                        })
                        nb_crees += 1
                        
                except Exception as e:
                    erreurs_detail.append({
                        'ligne': ligne_num,
                        'erreur': str(e),
                        'type': 'exception'
                    })
                    nb_erreurs += 1
        
        return {
            'success': True,
            'nb_crees': nb_crees,
            'nb_modifies': nb_modifies,
            'nb_sautes': nb_sautes,
            'nb_erreurs': nb_erreurs,
            'erreurs_detail': erreurs_detail[:50],  # Limiter à 50 erreurs
            'postes_crees': postes_crees[:20],  # Limiter à 20 pour affichage
            'postes_modifies': postes_modifies[:20],
            'regions_non_trouvees': list(regions_non_trouvees),
            'types_invalides': list(types_invalides),
            'total_lignes': len(df)
        }
        
    except Exception as e:
        return {
            'success': False,
            'erreur': str(e),
            'nb_crees': 0,
            'nb_modifies': 0,
            'nb_sautes': 0,
            'nb_erreurs': 0
        }


def _trouver_region_proche(nom_region):
    """Essaie de trouver une correspondance proche pour la région"""
    nom_region = nom_region.upper().strip()
    
    # Enlever les accents et caractères spéciaux pour comparaison
    nom_simplifie = nom_region.replace('-', ' ').replace('É', 'E').replace('Ê', 'E')
    
    for key, value in REGIONS_MAPPING.items():
        key_simplifie = key.replace('-', ' ').replace('É', 'E').replace('Ê', 'E')
        if nom_simplifie == key_simplifie:
            return value
        if nom_simplifie in key_simplifie or key_simplifie in nom_simplifie:
            return value
    
    return None


@login_required
@require_any_permission('peut_creer_poste_masse', 'peut_gerer_postes')
def telecharger_modele_postes_excel(request):
    """
    Génère et télécharge un modèle Excel pour l'import de postes
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Postes à importer"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes avec toutes les colonnes
    headers = [
        ('NOM', 40, 'Nom complet du poste'),
        ('REGION', 20, 'Région du Cameroun'),
        ('CODE', 15, 'Code unique (ex: PG001)'),
        ('TYPE', 12, 'peage ou pesage'),
        ('DEPARTEMENT', 20, 'Département (optionnel)'),
        ('AXE_ROUTIER', 30, 'Axe routier (optionnel)'),
        ('LATITUDE', 12, 'GPS (optionnel)'),
        ('LONGITUDE', 12, 'GPS (optionnel)'),
        ('DESCRIPTION', 40, 'Description (optionnel)'),
    ]
    
    for col, (header, width, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Ligne de description
    ws.row_dimensions[2].height = 30
    for col, (_, _, desc) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = desc
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Exemples
    exemples = [
        ['Péage de Yaoundé-Nord', 'Centre', 'PG001', 'peage', 'Mfoundi', 'Yaoundé-Ngaoundéré', 3.866, 11.517, 'Entrée nord de Yaoundé'],
        ['Péage de Douala-Bonabéri', 'Littoral', 'PG002', 'peage', 'Wouri', 'Douala-Bafoussam', 4.067, 9.733, 'Pont du Wouri'],
        ['Station de Pesage de Bafoussam', 'Ouest', 'PS001', 'pesage', 'Mifi', 'Bafoussam-Douala', 5.483, 10.417, 'Sortie ouest'],
        ['Station de Pesage de Ngaoundéré', 'Adamaoua', 'PS002', 'pesage', 'Vina', 'Ngaoundéré-Garoua', 7.317, 13.583, 'Gare routière'],
    ]
    
    for row_num, exemple in enumerate(exemples, 3):
        for col, value in enumerate(exemple, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.border = border
            if col == 4:  # Colonne type
                cell.font = Font(bold=True, color="0066CC" if value == 'peage' else "CC6600")
    
    # Feuille d'instructions
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 80
    
    instructions = [
        ("INSTRUCTIONS POUR L'IMPORT DE POSTES", Font(bold=True, size=14, color="4472C4")),
        ("", None),
        ("📋 FORMAT DU FICHIER :", Font(bold=True, size=12)),
        ("", None),
        ("Colonnes OBLIGATOIRES :", Font(bold=True)),
        ("  • NOM : Nom complet du poste (ex: 'Péage de Yaoundé-Nord')", None),
        ("  • REGION : Nom de la région (doit correspondre exactement)", None),
        ("  • CODE : Code unique du poste (ex: PG001 pour péage, PS001 pour pesage)", None),
        ("  • TYPE : Type de poste (peage ou pesage)", None),
        ("", None),
        ("Colonnes OPTIONNELLES :", Font(bold=True)),
        ("  • DEPARTEMENT : Département de la région", None),
        ("  • AXE_ROUTIER : Axe routier principal (ex: Yaoundé-Douala)", None),
        ("  • LATITUDE / LONGITUDE : Coordonnées GPS", None),
        ("  • DESCRIPTION : Description ou notes", None),
        ("", None),
        ("🚗 TYPES DE POSTE ACCEPTÉS :", Font(bold=True, size=12)),
        ("", None),
        ("  Pour PÉAGE : peage, péage, PEAGE, PÉAGE, PG, P", None),
        ("  Pour PESAGE : pesage, PESAGE, PS, S", None),
        ("", None),
        ("🗺️ RÉGIONS VALIDES :", Font(bold=True, size=12)),
        ("", None),
    ]
    
    for region in sorted(set(REGIONS_MAPPING.values())):
        instructions.append((f"  • {region}", None))
    
    instructions.extend([
        ("", None),
        ("⚠️ IMPORTANT :", Font(bold=True, size=12, color="CC0000")),
        ("", None),
        ("  • Les codes doivent être UNIQUES", None),
        ("  • Les noms de régions doivent correspondre EXACTEMENT", None),
        ("  • Le type doit être 'peage' ou 'pesage'", None),
        ("  • Les lignes avec des données manquantes seront ignorées", None),
        ("", None),
        ("💡 CONSEILS :", Font(bold=True, size=12)),
        ("", None),
        ("  • Utilisez PG pour les codes de péage (PG001, PG002...)", None),
        ("  • Utilisez PS pour les codes de pesage (PS001, PS002...)", None),
        ("  • Vérifiez l'orthographe des régions avant import", None),
    ])
    
    for row_num, (text, font) in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_num, column=1)
        cell.value = text
        if font:
            cell.font = font
    
    # Feuille des régions
    ws_regions = wb.create_sheet("Régions")
    ws_regions.cell(row=1, column=1).value = "RÉGIONS DU CAMEROUN"
    ws_regions.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    for row_num, region in enumerate(sorted(set(REGIONS_MAPPING.values())), 2):
        ws_regions.cell(row=row_num, column=1).value = region
    
    ws_regions.column_dimensions['A'].width = 25
    
    # Journaliser le téléchargement
    log_user_action(
        request.user,
        "TELECHARGEMENT_MODELE",
        "Téléchargement du modèle Excel pour import de postes",
        request,
        type_modele="postes"
    )
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=modele_import_postes.xlsx'
    
    return response


# ===================================================================
# VUES D'IMPORT DES UTILISATEURS
# ===================================================================

@login_required
@require_any_permission('peut_creer_utilisateur', 'peut_gerer_utilisateurs')
def import_utilisateurs_excel(request):
    """
    Vue pour importer des utilisateurs depuis un fichier Excel
    
    Permissions requises:
    - peut_creer_utilisateur : Permission de création d'utilisateurs
    - peut_gerer_utilisateurs : Permission générale de gestion des utilisateurs
    """
    
    user_desc = get_user_description(request.user)
    
    if request.method == 'POST':
        fichier = request.FILES.get('fichier_excel')
        action_doublon = request.POST.get('action_doublon', 'sauter')
        mot_de_passe_defaut = request.POST.get('mot_de_passe_defaut', 'supper2025')
        
        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier Excel")
            return redirect('accounts:import_utilisateurs')
        
        if not fichier.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Le fichier doit être au format Excel (.xlsx ou .xls)")
            return redirect('accounts:import_utilisateurs')
        
        try:
            # Journaliser le début de l'import
            log_user_action(
                request.user,
                "IMPORT_UTILISATEURS_DEBUT",
                f"Début d'import d'utilisateurs depuis le fichier {fichier.name}",
                request,
                fichier=fichier.name,
                taille=f"{fichier.size / 1024:.1f} Ko"
            )
            
            # Lire le fichier Excel
            df = pd.read_excel(fichier)
            
            # Traiter l'import
            resultats = traiter_import_utilisateurs(
                df, action_doublon, mot_de_passe_defaut, request.user
            )
            
            if resultats['success']:
                messages.success(
                    request,
                    f"Import terminé ! {resultats['nb_crees']} utilisateurs créés, "
                    f"{resultats['nb_modifies']} modifiés, {resultats['nb_sautes']} sautés, "
                    f"{resultats['nb_erreurs']} erreurs"
                )
                
                log_user_action(
                    request.user,
                    "IMPORT_UTILISATEURS_SUCCES",
                    f"Import d'utilisateurs terminé avec succès",
                    request,
                    fichier=fichier.name,
                    nb_crees=resultats['nb_crees'],
                    nb_modifies=resultats['nb_modifies'],
                    nb_sautes=resultats['nb_sautes'],
                    nb_erreurs=resultats['nb_erreurs']
                )
                
                logger.info(
                    f"👥 IMPORT UTILISATEURS RÉUSSI | {get_user_short_description(request.user)} | "
                    f"Créés: {resultats['nb_crees']}, Modifiés: {resultats['nb_modifies']}"
                )
            else:
                messages.error(request, f"Erreur lors de l'import: {resultats['erreur']}")
                log_erreur_action(request.user, "IMPORT_UTILISATEURS", resultats['erreur'], request)
            
            context = {
                'resultats': resultats,
                'fichier_nom': fichier.name,
                'user_desc': user_desc,
            }
            return render(request, 'accounts/import_utilisateurs_resultats.html', context)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier: {str(e)}")
            log_erreur_action(request.user, "IMPORT_UTILISATEURS", str(e), request)
            return redirect('accounts:import_utilisateurs')
    
    # GET request - afficher le formulaire
    context = {
        'user_desc': user_desc,
        'habilitations_disponibles': HABILITATIONS_LABELS,
    }
    return render(request, 'accounts/import_utilisateurs_form.html', context)


def traiter_import_utilisateurs(df, action_doublon, mot_de_passe_defaut, user):
    """
    Traite l'import des utilisateurs depuis Excel avec validation complète
    """
    
    # Normaliser les noms de colonnes
    df.columns = df.columns.str.strip().str.lower()
    
    nb_crees = 0
    nb_modifies = 0
    nb_sautes = 0
    nb_erreurs = 0
    erreurs_detail = []
    utilisateurs_crees = []
    utilisateurs_modifies = []
    habilitations_non_trouvees = set()
    
    # Vérifier les colonnes requises
    colonnes_requises = ['matricule', 'nom_complet', 'telephone', 'habilitation']
    colonnes_manquantes = [col for col in colonnes_requises if col not in df.columns]
    
    if colonnes_manquantes:
        return {
            'success': False,
            'erreur': f"Colonnes manquantes: {', '.join(colonnes_manquantes)}. "
                      f"Colonnes attendues: {', '.join(colonnes_requises)}",
            'nb_crees': 0,
            'nb_modifies': 0,
            'nb_sautes': 0,
            'nb_erreurs': 0
        }
    
    try:
        with transaction.atomic():
            for index, row in df.iterrows():
                ligne_num = index + 2
                
                try:
                    # Extraire et nettoyer les données
                    matricule = str(row['matricule']).strip().upper() if pd.notna(row['matricule']) else ''
                    nom_complet = str(row['nom_complet']).strip() if pd.notna(row['nom_complet']) else ''
                    telephone = str(row['telephone']).strip() if pd.notna(row['telephone']) else ''
                    habilitation_raw = str(row['habilitation']).strip().upper() if pd.notna(row['habilitation']) else ''
                    
                    # Données optionnelles
                    email = str(row.get('email', '')).strip() if pd.notna(row.get('email')) else ''
                    code_poste = str(row.get('code_poste', '')).strip().upper() if pd.notna(row.get('code_poste')) else ''
                    
                    # Vérifications
                    if not matricule or matricule.lower() == 'nan':
                        continue
                    
                    if not nom_complet or nom_complet.lower() == 'nan':
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Nom complet manquant pour matricule '{matricule}'",
                            'type': 'nom_manquant'
                        })
                        nb_erreurs += 1
                        continue
                    
                    if not telephone or telephone.lower() == 'nan':
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Téléphone manquant pour '{nom_complet}'",
                            'type': 'telephone_manquant'
                        })
                        nb_erreurs += 1
                        continue
                    
                    if not habilitation_raw or habilitation_raw.lower() == 'nan':
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Habilitation manquante pour '{nom_complet}'",
                            'type': 'habilitation_manquante'
                        })
                        nb_erreurs += 1
                        continue
                    
                    # Normaliser l'habilitation
                    habilitation = HABILITATIONS_IMPORT_MAPPING.get(habilitation_raw)
                    
                    if not habilitation:
                        habilitations_non_trouvees.add(habilitation_raw)
                        erreurs_detail.append({
                            'ligne': ligne_num,
                            'erreur': f"Habilitation '{habilitation_raw}' invalide pour '{nom_complet}'",
                            'type': 'habilitation_invalide'
                        })
                        nb_erreurs += 1
                        continue
                    
                    # Nettoyer le téléphone
                    telephone = _nettoyer_telephone(telephone)
                    
                    # Chercher le poste si spécifié
                    poste_affectation = None
                    if code_poste:
                        poste_affectation = Poste.objects.filter(code=code_poste).first()
                        if not poste_affectation:
                            erreurs_detail.append({
                                'ligne': ligne_num,
                                'erreur': f"Poste '{code_poste}' introuvable pour '{nom_complet}'",
                                'type': 'poste_invalide'
                            })
                            # Continuer quand même sans poste
                    
                    # Vérifier si l'utilisateur existe
                    utilisateur_existant = UtilisateurSUPPER.objects.filter(username=matricule).first()
                    
                    if utilisateur_existant:
                        if action_doublon == 'ecraser':
                            # Mise à jour
                            utilisateur_existant.nom_complet = nom_complet
                            utilisateur_existant.telephone = telephone
                            utilisateur_existant.habilitation = habilitation
                            
                            if email:
                                utilisateur_existant.email = email
                            if poste_affectation:
                                utilisateur_existant.poste_affectation = poste_affectation
                            
                            utilisateur_existant.save()
                            utilisateurs_modifies.append({
                                'matricule': matricule,
                                'nom': nom_complet,
                                'habilitation': get_habilitation_label(habilitation)
                            })
                            nb_modifies += 1
                        else:
                            nb_sautes += 1
                    else:
                        # Créer nouvel utilisateur
                        nouvel_utilisateur = UtilisateurSUPPER.objects.create_user(
                            username=matricule,
                            password=mot_de_passe_defaut,
                            nom_complet=nom_complet,
                            telephone=telephone,
                            email=email if email else None,
                            habilitation=habilitation,
                            poste_affectation=poste_affectation,
                            cree_par=user,
                            is_active=True
                        )
                        utilisateurs_crees.append({
                            'matricule': matricule,
                            'nom': nom_complet,
                            'habilitation': get_habilitation_label(habilitation),
                            'poste': poste_affectation.nom if poste_affectation else 'Aucun'
                        })
                        nb_crees += 1
                        
                except Exception as e:
                    erreurs_detail.append({
                        'ligne': ligne_num,
                        'erreur': str(e),
                        'type': 'exception'
                    })
                    nb_erreurs += 1
        
        return {
            'success': True,
            'nb_crees': nb_crees,
            'nb_modifies': nb_modifies,
            'nb_sautes': nb_sautes,
            'nb_erreurs': nb_erreurs,
            'erreurs_detail': erreurs_detail[:50],
            'utilisateurs_crees': utilisateurs_crees[:20],
            'utilisateurs_modifies': utilisateurs_modifies[:20],
            'habilitations_non_trouvees': list(habilitations_non_trouvees),
            'total_lignes': len(df),
            'mot_de_passe_defaut': mot_de_passe_defaut
        }
        
    except Exception as e:
        return {
            'success': False,
            'erreur': str(e),
            'nb_crees': 0,
            'nb_modifies': 0,
            'nb_sautes': 0,
            'nb_erreurs': 0
        }


def _nettoyer_telephone(telephone):
    """Nettoie et formate un numéro de téléphone camerounais"""
    # Enlever tout sauf les chiffres et le +
    telephone = ''.join(c for c in str(telephone) if c.isdigit() or c == '+')
    
    # Si commence par 237, ajouter le +
    if telephone.startswith('237') and not telephone.startswith('+'):
        telephone = '+' + telephone
    
    # Si 9 chiffres sans indicatif, ajouter +237
    if len(telephone) == 9 and telephone[0] in '6789':
        telephone = '+237' + telephone
    
    return telephone


@login_required
@require_any_permission('peut_creer_utilisateur', 'peut_gerer_utilisateurs')
def telecharger_modele_utilisateurs_excel(request):
    """
    Génère et télécharge un modèle Excel pour l'import d'utilisateurs
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs à importer"
    
    # Styles
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = [
        ('MATRICULE', 15, 'Matricule unique'),
        ('NOM_COMPLET', 35, 'Nom et prénoms'),
        ('TELEPHONE', 18, '+237XXXXXXXXX'),
        ('HABILITATION', 25, 'Rôle dans le système'),
        ('EMAIL', 30, 'Email (optionnel)'),
        ('CODE_POSTE', 15, 'Code poste (optionnel)'),
    ]
    
    for col, (header, width, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Descriptions
    for col, (_, _, desc) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = desc
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal="center")
    
    # Exemples
    exemples = [
        ['ABC123456', 'MBARGA Achille Jean', '+237699123456', 'CHEF_PEAGE', 'achille.mbarga@email.cm', 'PG001'],
        ['DEF789012', 'NGONO Marie Claire', '+237677654321', 'AGENT_INVENTAIRE', '', 'PG001'],
        ['GHI345678', 'FOTSO Jean Pierre', '+237655987654', 'CHEF_STATION_PESAGE', 'jp.fotso@email.cm', 'PS001'],
        ['JKL901234', 'KAMGA Paul André', '+237688112233', 'CISOP_PEAGE', '', ''],
    ]
    
    for row_num, exemple in enumerate(exemples, 3):
        for col, value in enumerate(exemple, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.border = border
    
    # Feuille des habilitations
    ws_hab = wb.create_sheet("Habilitations")
    ws_hab.column_dimensions['A'].width = 30
    ws_hab.column_dimensions['B'].width = 40
    ws_hab.column_dimensions['C'].width = 50
    
    ws_hab.cell(row=1, column=1).value = "CODE"
    ws_hab.cell(row=1, column=2).value = "LIBELLÉ"
    ws_hab.cell(row=1, column=3).value = "DESCRIPTION"
    
    for col in range(1, 4):
        ws_hab.cell(row=1, column=col).font = header_font
        ws_hab.cell(row=1, column=col).fill = header_fill
    
    habilitations_info = [
        ('admin_principal', 'Administrateur Principal', 'Accès complet au système'),
        ('coord_psrr', 'Coordonnateur PSRR', 'Coordination générale'),
        ('serv_info', 'Service Informatique', 'Maintenance et support'),
        ('serv_emission', 'Service Émission', 'Gestion des tickets et émissions'),
        ('chef_ag', 'Chef Affaires Générales', 'Gestion administrative'),
        ('serv_controle', 'Service Contrôle', 'Contrôle et validation'),
        ('serv_ordre', 'Service Ordre', 'Secrétariat et ordre'),
        ('cisop_peage', 'CISOP Péage', 'Intervention péage'),
        ('cisop_pesage', 'CISOP Pesage', 'Intervention pesage'),
        ('chef_peage', 'Chef de Poste Péage', 'Gestion poste péage'),
        ('chef_station_pesage', 'Chef Station Pesage', 'Gestion station pesage'),
        ('regisseur_pesage', 'Régisseur Pesage', 'Régie station pesage'),
        ('chef_equipe_pesage', 'Chef Équipe Pesage', 'Chef d\'équipe pesage'),
        ('agent_inventaire', 'Agent Inventaire', 'Saisie inventaires'),
        ('caissier', 'Caissier', 'Encaissement'),
        ('focal_regional', 'Point Focal Régional', 'Coordination régionale'),
    ]
    
    for row_num, (code, libelle, desc) in enumerate(habilitations_info, 2):
        ws_hab.cell(row=row_num, column=1).value = code
        ws_hab.cell(row=row_num, column=2).value = libelle
        ws_hab.cell(row=row_num, column=3).value = desc
    
    # Feuille instructions
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 80
    
    instructions = [
        ("INSTRUCTIONS POUR L'IMPORT D'UTILISATEURS", Font(bold=True, size=14, color="2E7D32")),
        ("", None),
        ("📋 COLONNES OBLIGATOIRES :", Font(bold=True, size=12)),
        ("", None),
        ("  • MATRICULE : Identifiant unique (6-20 caractères alphanumériques)", None),
        ("  • NOM_COMPLET : Nom et prénoms complets", None),
        ("  • TELEPHONE : Format +237XXXXXXXXX ou 6XXXXXXXX", None),
        ("  • HABILITATION : Code du rôle (voir feuille 'Habilitations')", None),
        ("", None),
        ("📋 COLONNES OPTIONNELLES :", Font(bold=True, size=12)),
        ("", None),
        ("  • EMAIL : Adresse email (pour récupération de mot de passe)", None),
        ("  • CODE_POSTE : Code du poste d'affectation (ex: PG001)", None),
        ("", None),
        ("⚠️ IMPORTANT :", Font(bold=True, size=12, color="CC0000")),
        ("", None),
        ("  • Le mot de passe par défaut sera celui défini dans le formulaire", None),
        ("  • Les utilisateurs devront changer leur mot de passe à la première connexion", None),
        ("  • Les matricules doivent être UNIQUES", None),
        ("  • Vérifiez les codes d'habilitation dans la feuille 'Habilitations'", None),
    ]
    
    for row_num, (text, font) in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_num, column=1)
        cell.value = text
        if font:
            cell.font = font
    
    # Journaliser
    log_user_action(
        request.user,
        "TELECHARGEMENT_MODELE",
        "Téléchargement du modèle Excel pour import d'utilisateurs",
        request,
        type_modele="utilisateurs"
    )
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=modele_import_utilisateurs.xlsx'
    
    return response


# ===================================================================
# VUES D'EXPORT
# ===================================================================

@login_required
@require_any_permission('peut_voir_liste_stocks_peage', 'peut_gerer_postes')
def export_postes_excel(request):
    """
    Exporte la liste des postes au format Excel
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Postes SUPPER"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # En-têtes
    headers = ['Code', 'Nom', 'Type', 'Région', 'Département', 'Axe Routier', 'Actif', 'Date Création']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Données
    postes = Poste.objects.all().order_by('region', 'type', 'nom')
    
    for row_num, poste in enumerate(postes, 2):
        ws.cell(row=row_num, column=1).value = poste.code
        ws.cell(row=row_num, column=2).value = poste.nom
        ws.cell(row=row_num, column=3).value = 'Péage' if poste.type == 'peage' else 'Pesage'
        ws.cell(row=row_num, column=4).value = poste.region
        ws.cell(row=row_num, column=5).value = poste.departement or ''
        ws.cell(row=row_num, column=6).value = getattr(poste, 'axe_routier', '') or ''
        ws.cell(row=row_num, column=7).value = 'Oui' if poste.is_active else 'Non'
        ws.cell(row=row_num, column=8).value = poste.date_creation.strftime('%d/%m/%Y')
    
    # Ajuster largeurs
    widths = [15, 40, 12, 18, 20, 30, 8, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Journaliser
    log_user_action(
        request.user,
        "EXPORT_POSTES",
        f"Export Excel de {postes.count()} postes",
        request,
        nb_postes=postes.count(),
        format="Excel"
    )
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=export_postes_{timestamp}.xlsx'
    
    return response


@login_required
@require_any_permission('peut_gerer_utilisateurs', 'peut_voir_journal_audit')
def export_utilisateurs_excel(request):
    """
    Exporte la liste des utilisateurs au format Excel
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs SUPPER"
    
    # Styles
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # En-têtes
    headers = ['Matricule', 'Nom Complet', 'Téléphone', 'Email', 'Habilitation', 
               'Poste Affectation', 'Actif', 'Dernière Connexion']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Données
    utilisateurs = UtilisateurSUPPER.objects.all().select_related('poste_affectation').order_by('nom_complet')
    
    for row_num, user in enumerate(utilisateurs, 2):
        ws.cell(row=row_num, column=1).value = user.username
        ws.cell(row=row_num, column=2).value = user.nom_complet
        ws.cell(row=row_num, column=3).value = user.telephone
        ws.cell(row=row_num, column=4).value = user.email or ''
        ws.cell(row=row_num, column=5).value = get_habilitation_label(user.habilitation)
        ws.cell(row=row_num, column=6).value = user.poste_affectation.nom if user.poste_affectation else ''
        ws.cell(row=row_num, column=7).value = 'Oui' if user.is_active else 'Non'
        ws.cell(row=row_num, column=8).value = user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else 'Jamais'
    
    # Ajuster largeurs
    widths = [15, 35, 18, 30, 25, 30, 8, 18]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Journaliser
    log_user_action(
        request.user,
        "EXPORT_UTILISATEURS",
        f"Export Excel de {utilisateurs.count()} utilisateurs",
        request,
        nb_utilisateurs=utilisateurs.count(),
        format="Excel"
    )
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=export_utilisateurs_{timestamp}.xlsx'
    
    return response


# ===================================================================
# API AJAX POUR VALIDATION
# ===================================================================

@login_required
@require_http_methods(["POST"])
def valider_fichier_import_ajax(request):
    """
    API AJAX pour valider un fichier avant import complet
    Retourne un aperçu des données et erreurs potentielles
    """
    
    fichier = request.FILES.get('fichier')
    type_import = request.POST.get('type', 'postes')
    
    if not fichier:
        return JsonResponse({'success': False, 'erreur': 'Aucun fichier fourni'})
    
    try:
        df = pd.read_excel(fichier)
        df.columns = df.columns.str.strip().str.lower()
        
        if type_import == 'postes':
            colonnes_requises = ['nom', 'region', 'code', 'type']
        else:
            colonnes_requises = ['matricule', 'nom_complet', 'telephone', 'habilitation']
        
        colonnes_manquantes = [col for col in colonnes_requises if col not in df.columns]
        
        if colonnes_manquantes:
            return JsonResponse({
                'success': False,
                'erreur': f"Colonnes manquantes: {', '.join(colonnes_manquantes)}"
            })
        
        # Aperçu des premières lignes
        apercu = df.head(5).fillna('').to_dict('records')
        
        # Statistiques
        stats = {
            'total_lignes': len(df),
            'colonnes_presentes': list(df.columns),
            'colonnes_manquantes': colonnes_manquantes,
        }
        
        return JsonResponse({
            'success': True,
            'apercu': apercu,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'erreur': f"Erreur lecture fichier: {str(e)}"
        })