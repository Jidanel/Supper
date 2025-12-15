# ===================================================================
# inventaire/views.py - VUES COMPLÈTES POUR LE MODULE INVENTAIRE
# Application SUPPER - Suivi des Péages et Pesages Routiers
# ===================================================================

from math import e
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator
from decimal import Decimal, InvalidOperation
import decimal
from django.db.models import Sum, Avg, Count, Q
from django.utils import timezone
from django.db import models
from django.db import transaction
from datetime import datetime, date, timedelta
import json
from django.views.decorators.csrf import csrf_exempt
import logging

from .forms import *
from .models import *
from django.forms import formset_factory

logger = logging.getLogger('supper')

# ===================================================================
# IMPORTS DES DÉCORATEURS ET PERMISSIONS GRANULAIRES
# Ces imports proviennent de common/decorators.py et common/permissions.py
# ===================================================================
from common.decorators import *
from common.permissions import *
# Import de la fonction de logging depuis common/utils.py
from common.utils import log_user_action, require_permission


# ===================================================================
# FONCTIONS UTILITAIRES MISES À JOUR
# ===================================================================

def _check_admin_permission(user):
    """
    Vérifie les permissions administrateur avec les nouvelles habilitations.
    MISE À JOUR: Utilise is_admin_user() de common/permissions.py
    """
    return is_admin_user(user)


def _check_inventaire_view_permission(user):
    """
    Vérifie si l'utilisateur peut voir les inventaires.
    Permission: peut_voir_liste_inventaires OU peut_voir_liste_inventaires_admin
    """
    return has_any_permission(user, [
        'peut_voir_liste_inventaires',
        'peut_voir_liste_inventaires_admin'
    ])


def _check_inventaire_edit_permission(user):
    """
    Vérifie si l'utilisateur peut saisir/modifier des inventaires.
    Permission: peut_saisir_inventaire_normal OU peut_saisir_inventaire_admin
    """
    return has_any_permission(user, [
        'peut_saisir_inventaire_normal',
        'peut_saisir_inventaire_admin'
    ])


def _check_recette_view_permission(user):
    """
    Vérifie si l'utilisateur peut voir les recettes péage.
    Permission: peut_voir_liste_recettes_peage
    """
    return has_permission(user, 'peut_voir_liste_recettes_peage')


def _check_recette_edit_permission(user):
    """
    Vérifie si l'utilisateur peut saisir des recettes péage.
    Permission: peut_saisir_recette_peage
    """
    return has_permission(user, 'peut_saisir_recette_peage')


def _check_programmation_permission(user):
    """
    Vérifie si l'utilisateur peut programmer des inventaires.
    Permission: peut_programmer_inventaire
    """
    return has_permission(user, 'peut_programmer_inventaire')


def _check_stats_deperdition_permission(user):
    """
    Vérifie si l'utilisateur peut voir les stats de déperdition.
    Permission: peut_voir_stats_deperdition
    """
    return has_permission(user, 'peut_voir_stats_deperdition')


def _check_jours_impertinents_permission(user):
    """
    Vérifie si l'utilisateur peut voir les jours impertinents.
    Permission: peut_voir_jours_impertinents
    """
    return has_permission(user, 'peut_voir_jours_impertinents')


def _check_quittance_permission(user):
    """
    Vérifie si l'utilisateur peut saisir des quittances péage.
    Permission: peut_saisir_quittance_peage
    """
    return has_permission(user, 'peut_saisir_quittance_peage')


def _check_comptabilisation_permission(user):
    """
    Vérifie si l'utilisateur peut comptabiliser les quittances péage.
    Permission: peut_comptabiliser_quittances_peage
    """
    return has_permission(user, 'peut_comptabiliser_quittances_peage')


# ===================================================================
# MIXINS ET FONCTIONS UTILITAIRES
# ===================================================================

class InventaireMixin(LoginRequiredMixin):
    """
    Mixin de base pour les vues inventaire avec permissions granulaires.
    MISE À JOUR: Vérifie peut_voir_liste_inventaires au lieu de peut_gerer_inventaire
    """
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        
        # Vérifier permission de visualisation d'inventaires
        if not _check_inventaire_view_permission(request.user):
            log_user_action(
                request.user,
                "ACCÈS REFUSÉ - Liste inventaires",
                f"Permission peut_voir_liste_inventaires manquante | IP: {request.META.get('REMOTE_ADDR')}",
                request
            )
            messages.error(request, _("Vous n'avez pas les permissions pour accéder aux inventaires."))
            return redirect('common:dashboard')
        
        return super().dispatch(request, *args, **kwargs)


class AdminRequiredMixin(UserPassesTestMixin):
    """
    Mixin pour les vues nécessitant des droits admin.
    MISE À JOUR: Utilise is_admin_user() de common/permissions.py
    """
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def handle_no_permission(self):
        log_user_action(
            self.request.user,
            "ACCÈS REFUSÉ - Zone admin",
            f"Habilitation: {getattr(self.request.user, 'habilitation', 'N/A')} | IP: {self.request.META.get('REMOTE_ADDR')}",
            self.request
        )
        messages.error(self.request, _("Accès non autorisé à cette fonctionnalité."))
        return redirect('common:dashboard')

class RecetteMixin(LoginRequiredMixin):
    """
    Mixin pour les vues recettes avec permissions granulaires.
    NOUVEAU: Vérifie peut_voir_liste_recettes_peage
    """
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        
        # Vérifier permission de visualisation des recettes
        if not _check_recette_view_permission(request.user):
            log_user_action(
                request.user,
                "ACCÈS REFUSÉ - Liste recettes",
                f"Permission peut_voir_liste_recettes_peage manquante | IP: {request.META.get('REMOTE_ADDR')}",
                request
            )
            messages.error(request, _("Vous n'avez pas les permissions pour accéder aux recettes."))
            return redirect('common:dashboard')
        
        return super().dispatch(request, *args, **kwargs)


def _check_admin_permission(user):
    """Vérifier les permissions administrateur - FONCTION CORRIGÉE"""
    return (user.is_authenticated and (
        user.is_superuser or 
        user.is_staff or 
        hasattr(user, 'habilitation') and user.habilitation in [
            'admin_principal', 'coord_psrr', 'serv_info', 'serv_emission'
        ]
    ))


def _log_inventaire_action(request, action, details=""):
    """Journaliser une action inventaire"""
    try:
        JournalAudit.objects.create(
            utilisateur=request.user,
            action=action,
            details=details,
            adresse_ip=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
            session_key=request.session.session_key,
            url_acces=request.path,
            methode_http=request.method,
            succes=True
        )
        logger.info(f"INVENTAIRE - {request.user.username} - {action}")
    except Exception as e:
        logger.error(f"Erreur journalisation inventaire: {str(e)}")


# ===================================================================
# VUES PRINCIPALES D'INVENTAIRE
# ===================================================================

class InventaireListView(InventaireMixin, ListView):
    """
    Liste des inventaires avec filtres et recherche.
    MISE À JOUR: Permissions granulaires pour l'affichage
    """
    model = InventaireJournalier
    template_name = 'inventaire/inventaire_list.html'
    context_object_name = 'inventaires'
    paginate_by = 20
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        
        # Vérifier permission avec système granulaire
        if not _check_inventaire_view_permission(request.user):
            log_user_action(
                request.user,
                "ACCÈS REFUSÉ - Liste inventaires",
                f"Permission manquante: peut_voir_liste_inventaires | "
                f"Habilitation: {getattr(request.user, 'habilitation', 'N/A')} | "
                f"IP: {request.META.get('REMOTE_ADDR')}",
                request
            )
            messages.error(request, "Vous n'avez pas les permissions pour accéder aux inventaires.")
            return redirect('common:dashboard')
        
        # Log de l'accès autorisé
        log_user_action(
            request.user,
            "Consultation liste inventaires",
            f"Accès autorisé | Habilitation: {getattr(request.user, 'habilitation', 'N/A')}",
            request
        )
        
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = InventaireJournalier.objects.filter(
            type_inventaire='normal'
        ).select_related(
            'poste', 'agent_saisie',
        ).prefetch_related('details_periodes')
        
        # Filtrer selon les postes accessibles à l'utilisateur
        # Utilisation des permissions granulaires
        user = self.request.user
        
        # Si l'utilisateur a peut_voir_liste_inventaires_admin, il voit tout
        if has_permission(user, 'peut_voir_liste_inventaires_admin'):
            pass  # Pas de filtre
        elif hasattr(user, 'get_postes_accessibles'):
            queryset = queryset.filter(
                poste__in=user.get_postes_accessibles()
            )
        elif user.poste_affectation:
            queryset = queryset.filter(poste=user.poste_affectation)
        else:
            queryset = queryset.none()
        
        # Filtres de recherche (inchangé)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(poste__nom__icontains=search) |
                Q(poste__code__icontains=search) |
                Q(agent_saisie__nom_complet__icontains=search)
            )
        
        # Filtre par poste
        poste_id = self.request.GET.get('poste')
        if poste_id:
            queryset = queryset.filter(poste_id=poste_id)
        
        # Filtre par date
        date_debut = self.request.GET.get('date_debut')
        date_fin = self.request.GET.get('date_fin')
        
        if date_debut:
            try:
                date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
                queryset = queryset.filter(date__gte=date_debut)
            except ValueError:
                pass
        
        if date_fin:
            try:
                date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
                queryset = queryset.filter(date__lte=date_fin)
            except ValueError:
                pass
        
        return queryset.order_by('-date', 'poste__nom')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Permissions granulaires pour le contexte
        context['can_admin'] = has_permission(user, 'peut_voir_liste_inventaires_admin')
        context['can_edit'] = _check_inventaire_edit_permission(user)
        context['can_program'] = _check_programmation_permission(user)
        context['can_view_stats'] = _check_stats_deperdition_permission(user)
        
        # Postes accessibles pour le filtre
        if context['can_admin']:
            context['postes'] = Poste.objects.filter(is_active=True, type='peage').order_by('nom')
        elif hasattr(user, 'get_postes_accessibles'):
            context['postes'] = user.get_postes_accessibles()
        else:
            context['postes'] = Poste.objects.none()
        
        # Statistiques rapides
        total_inventaires = self.get_queryset().count()
        inventaires_today = self.get_queryset().filter(date=timezone.now().date()).count()
        
        context.update({
            'total_inventaires': total_inventaires,
            'inventaires_today': inventaires_today,
            'current_filters': {
                'search': self.request.GET.get('search', ''),
                'poste': self.request.GET.get('poste', ''),
                'date_debut': self.request.GET.get('date_debut', ''),
                'date_fin': self.request.GET.get('date_fin', ''),
                'statut': self.request.GET.get('statut', ''),
            }
        })
        
        return context


class InventaireDetailView(InventaireMixin, DetailView):
    """
    Détail d'un inventaire avec calculs et historique.
    MISE À JOUR: Permissions granulaires
    """
    model = InventaireJournalier
    template_name = 'inventaire/inventaire_detail.html'
    context_object_name = 'inventaire'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        
        # Vérifier permission de visualisation
        if not _check_inventaire_view_permission(request.user):
            log_user_action(
                request.user,
                "ACCÈS REFUSÉ - Détail inventaire",
                f"Permission manquante | Inventaire ID: {kwargs.get('pk')}",
                request
            )
            messages.error(request, "Vous n'avez pas accès à ce détail d'inventaire.")
            return redirect('common:dashboard')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user
        
        # Vérifier l'accès au poste avec permissions granulaires
        if not has_permission(user, 'peut_voir_liste_inventaires_admin'):
            if not check_poste_access(user, obj.poste):
                log_user_action(
                    user,
                    "ACCÈS REFUSÉ - Détail inventaire (poste)",
                    f"Poste non autorisé: {obj.poste.nom} | Inventaire ID: {obj.pk}",
                    self.request
                )
                raise PermissionError("Accès non autorisé à ce poste.")
        
        # Log de l'accès autorisé
        log_user_action(
            user,
            "Consultation détail inventaire",
            f"Poste: {obj.poste.nom} | Date: {obj.date} | Total véhicules: {obj.total_vehicules}",
            self.request
        )
        
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventaire = self.object
        user = self.request.user
        
        # Calculs principaux avec gestion sécurisée (inchangé)
        context['total_vehicules'] = inventaire.total_vehicules or 0
        
        try:
            context['moyenne_horaire'] = int(inventaire.calculer_moyenne_horaire())
        except:
            context['moyenne_horaire'] = 0
            
        try:
            context['estimation_24h'] = int(inventaire.estimer_total_24h())
        except:
            context['estimation_24h'] = 0
            
        try:
            recette_pot = inventaire.calculer_recette_potentielle()
            if isinstance(recette_pot, Decimal):
                context['recette_potentielle'] = float(str(recette_pot))
            else:
                context['recette_potentielle'] = float(recette_pot) if recette_pot else 0
        except:
            context['recette_potentielle'] = 0
        
        # Détails par période
        details_periodes = inventaire.details_periodes.all().order_by('periode')
        context['details_periodes'] = details_periodes
        
        # Préparer les données du graphique
        graph_periodes = []
        graph_vehicules = []
        for detail in details_periodes:
            graph_periodes.append(detail.get_periode_display())
            graph_vehicules.append(detail.nombre_vehicules)
        
        context['graph_data'] = {
            'periodes': graph_periodes,
            'vehicules': graph_vehicules,
        }
        context['graph_data_json'] = json.dumps(context['graph_data'])
        
        # Recette associée
        try:
            recette = RecetteJournaliere.objects.get(
                poste=inventaire.poste,
                date=inventaire.date
            )
            context['recette'] = recette
        except RecetteJournaliere.DoesNotExist:
            context['recette'] = None
        
        # Permissions granulaires pour les actions
        context['can_edit'] = (
            has_permission(user, 'peut_saisir_inventaire_admin') or 
            (has_permission(user, 'peut_saisir_inventaire_normal') and inventaire.agent_saisie == user)
        )
        context['can_delete'] = has_permission(user, 'peut_saisir_inventaire_admin')
        context['can_view_stats'] = _check_stats_deperdition_permission(user)
        
        # Vérifier si jour impertinent
        try:
            from inventaire.models import ConfigurationJour, StatutJour
            config = ConfigurationJour.objects.filter(date=inventaire.date).first()
            context['jour_impertinent'] = config and config.statut == StatutJour.IMPERTINENT
        except:
            context['jour_impertinent'] = False
        
        return context
    
class SaisieInventaireView(LoginRequiredMixin, View):
    """
    Vue pour la saisie d'inventaire par les agents.
    MISE À JOUR: Permissions granulaires peut_saisir_inventaire_normal
                 et peut_saisir_inventaire_admin
    """
    template_name = 'inventaire/saisie_inventaire.html'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Vous devez être connecté.")
            return redirect('accounts:login')
        
        # Vérifier les permissions de saisie avec système granulaire
        user = request.user
        can_saisie_normal = has_permission(user, 'peut_saisir_inventaire_normal')
        can_saisie_admin = has_permission(user, 'peut_saisir_inventaire_admin')
        
        if not (can_saisie_normal or can_saisie_admin):
            log_user_action(
                user,
                "ACCÈS REFUSÉ - Saisie inventaire",
                f"Permissions manquantes: peut_saisir_inventaire_normal/admin | "
                f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | "
                f"IP: {request.META.get('REMOTE_ADDR')}",
                request
            )
            messages.error(request, "Vous n'avez pas la permission de saisir des inventaires.")
            return HttpResponseForbidden("Accès non autorisé")
        
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, poste_id=None, date_str=None):
        """Affichage du formulaire de saisie"""
        from datetime import date, datetime
        
        user = request.user
        poste = None
        
        # Déterminer si l'utilisateur a des droits admin
        is_admin = has_permission(user, 'peut_saisir_inventaire_admin')
        
        # 1. GESTION ADMIN (permission peut_saisir_inventaire_admin)
        if is_admin:
            if not poste_id:
                # Admin sans poste_id : afficher la sélection
                mois_actuel = date.today().replace(day=1)
                
                postes_programmes = Poste.objects.filter(
                    programmations_inventaire__mois=mois_actuel,
                    programmations_inventaire__actif=True,
                    is_active=True, type='peage'
                ).distinct().order_by('nom')
                
                log_user_action(
                    user,
                    "Accès sélection poste inventaire",
                    f"Mode admin | Postes programmés: {postes_programmes.count()}",
                    request
                )
                
                return render(request, 'inventaire/choix_poste.html', {
                    'postes': postes_programmes,
                    'date_str': date_str,
                    'mois_actuel': mois_actuel
                })
            else:
                # Admin avec poste_id
                try:
                    poste = Poste.objects.get(id=poste_id, is_active=True, type='peage')
                except Poste.DoesNotExist:
                    messages.error(request, "Poste introuvable ou inactif.")
                    return redirect('inventaire:saisie_inventaire')
        
        # 2. GESTION NON-ADMIN (permission peut_saisir_inventaire_normal)
        else:
            if poste_id:
                try:
                    poste = Poste.objects.get(id=poste_id, is_active=True, type='peage')
                    # Vérifier que c'est bien son poste
                    if user.poste_affectation and poste.id != user.poste_affectation.id:
                        log_user_action(
                            user,
                            "Redirection poste inventaire",
                            f"Tentative accès poste {poste.nom} | Redirigé vers {user.poste_affectation.nom}",
                            request
                        )
                        messages.warning(request, f"Redirection vers votre poste: {user.poste_affectation.nom}")
                        return redirect('inventaire:saisie_inventaire_avec_poste', poste_id=user.poste_affectation.id)
                except Poste.DoesNotExist:
                    messages.error(request, "Poste introuvable.")
                    return redirect('inventaire:inventaire_list')
            else:
                if not user.poste_affectation:
                    messages.error(request, "Aucun poste d'affectation configuré.")
                    return redirect('common:dashboard')
                poste = user.poste_affectation
            
            # Vérification des droits d'accès pour non-admin
            if not check_poste_access(user, poste):
                log_user_action(
                    user,
                    "ACCÈS REFUSÉ - Saisie inventaire (poste)",
                    f"Poste non autorisé: {poste.nom}",
                    request
                )
                messages.error(request, "Accès non autorisé à ce poste.")
                return redirect('inventaire:inventaire_list')
        
        # 3. À CE STADE, ON A FORCÉMENT UN POSTE
        if not poste:
            messages.error(request, "Aucun poste sélectionné.")
            return redirect('inventaire:inventaire_list')
        
        # 4. DÉTERMINER LA DATE
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Format de date invalide.")
                return redirect('inventaire:saisie_inventaire_avec_poste', poste_id=poste.id)
        else:
            target_date = timezone.now().date()
        
        # 5. VÉRIFIER LA PROGRAMMATION
        mois_inventaire = target_date.replace(day=1)
        programmation_existe = ProgrammationInventaire.objects.filter(
            poste=poste,
            mois=mois_inventaire,
            actif=True
        ).exists()
        
        if not programmation_existe:
            log_user_action(
                user,
                "Saisie inventaire - Poste non programmé",
                f"Poste: {poste.nom} | Mois: {target_date.strftime('%B %Y')}",
                request
            )
            messages.error(
                request, 
                f"Le poste {poste.nom} n'est pas programmé pour {target_date.strftime('%B %Y')}. "
                "Veuillez d'abord programmer l'inventaire mensuel."
            )
            return redirect('inventaire:programmer_inventaire')
        
        # 6. CRÉER/RÉCUPÉRER L'INVENTAIRE
        inventaire, created = InventaireJournalier.objects.get_or_create(
            poste=poste,
            date=target_date,
            defaults={'agent_saisie': user}
        )
        
        # Log de l'action
        action = "Création inventaire" if created else "Accès inventaire existant"
        log_user_action(
            user,
            action,
            f"Poste: {poste.nom} | Date: {target_date} | Type: {'Normal' if not is_admin else 'Admin'}",
            request
        )
        
        # 7. PRÉPARER LES DONNÉES
        details_existants = {
            d.periode: d for d in inventaire.details_periodes.all()
        }
        
        periodes_data = []
        for periode_choice in PeriodeHoraire.choices:
            periode_code, periode_display = periode_choice
            detail = details_existants.get(periode_code)
            periodes_data.append({
                'code': periode_code,
                'display': periode_display,
                'nombre_vehicules': detail.nombre_vehicules if detail else 0,
                'observations': detail.observations_periode if detail else '',
                'has_data': detail is not None,
            })
        
        context = {
            'inventaire': inventaire,
            'poste': poste,
            'target_date': target_date,
            'periodes_data': periodes_data,
            'is_new': created,
            'is_admin': is_admin,
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request, poste_id=None, date_str=None):
        """Traitement de la saisie d'inventaire"""
        user = request.user
        
        # Récupérer les paramètres
        if poste_id:
            poste = get_object_or_404(Poste, id=poste_id)
        else:
            poste = user.poste_affectation
            if not poste:
                messages.error(request, "Aucun poste d'affectation configuré.")
                return redirect('common:dashboard')
        
        if date_str:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            target_date = timezone.now().date()
        
        # Vérifications de sécurité avec permissions granulaires
        is_admin = has_permission(user, 'peut_saisir_inventaire_admin')
        
        if not is_admin and not check_poste_access(user, poste):
            log_user_action(
                user,
                "ACCÈS REFUSÉ - POST Saisie inventaire",
                f"Poste non autorisé: {poste.nom} | Date: {target_date}",
                request
            )
            return JsonResponse({'error': 'Accès non autorisé'}, status=403)
        
        try:
            # Récupérer ou créer l'inventaire
            inventaire, created = InventaireJournalier.objects.get_or_create(
                poste=poste,
                date=target_date,
                type_inventaire='normal',
                defaults={
                    'agent_saisie': user,
                    'type_inventaire': 'normal'
                }
            )
            
            # Vérifier si peut être modifié (permissions granulaires)
            can_modify = is_admin or inventaire.peut_etre_modifie_par(user)
            if not can_modify:
                log_user_action(
                    user,
                    "Modification inventaire refusée",
                    f"Inventaire ID: {inventaire.pk} | Agent original: {inventaire.agent_saisie}",
                    request
                )
                messages.error(request, 
                    "Cet inventaire a déjà été saisi et ne peut être modifié que par un administrateur.")
                return redirect('inventaire:inventaire_detail', pk=inventaire.pk)
            
            # Traiter les données de période
            total_vehicules = 0
            details_saved = 0
            
            for periode_choice in PeriodeHoraire.choices:
                periode_code, _ = periode_choice
                field_name = f'periode_{periode_code.replace("-", "_")}'
                value = request.POST.get(field_name)
                
                if value and value.strip():
                    try:
                        nombre = int(value)
                        if nombre >= 0:
                            DetailInventairePeriode.objects.update_or_create(
                                inventaire=inventaire,
                                periode=periode_code,
                                defaults={'nombre_vehicules': nombre}
                            )
                            total_vehicules += nombre
                            details_saved += 1
                    except ValueError:
                        continue
            
            # Mettre à jour le total
            inventaire.total_vehicules = total_vehicules
            inventaire.derniere_modification_par = user
            inventaire.save()
            
            # Log détaillé de la saisie
            log_user_action(
                user,
                "Saisie inventaire réussie",
                f"Poste: {poste.nom} | Date: {target_date} | "
                f"Total véhicules: {total_vehicules} | Périodes saisies: {details_saved} | "
                f"Mode: {'Admin' if is_admin else 'Normal'}",
                request
            )
            
            messages.success(request, "Inventaire sauvegardé avec succès.")
            return redirect('inventaire:inventaire_detail', pk=inventaire.pk)
        
        except Exception as e:
            logger.error(f"Erreur saisie inventaire: {str(e)}")
            log_user_action(
                user,
                "ERREUR - Saisie inventaire",
                f"Poste: {poste.nom} | Date: {target_date} | Erreur: {str(e)}",
                request
            )
            messages.error(request, "Erreur lors de la sauvegarde de l'inventaire.")
            return redirect('inventaire:saisie_inventaire')

@login_required
@require_permission('peut_gerer_inventaire')
def modifier_inventaire(request, pk):
    """
    Permet de modifier un inventaire existant
    Pas de vérification de verrouillage car plus de verrouillage
    """
    inventaire = get_object_or_404(InventaireJournalier, pk=pk)
    
    # Vérifier que l'utilisateur peut accéder à ce poste
    if not request.user.peut_acceder_poste(inventaire.poste):
        messages.error(request, "Vous n'avez pas accès à ce poste.")
        return HttpResponseForbidden("Accès non autorisé")
    
    # Vérifier si l'utilisateur peut modifier (admin ou agent si modifiable_par_agent)
    if not inventaire.peut_etre_modifie_par(request.user):
        messages.error(request, "Vous n'avez pas la permission de modifier cet inventaire.")
        return HttpResponseForbidden("Permission insuffisante")
    
    if request.method == 'POST':
        form = InventaireJournalierForm(
            request.POST, 
            instance=inventaire,
            user=request.user
        )
        
        if form.is_valid():
            inventaire = form.save(commit=False)
            inventaire.derniere_modification_par = request.user
            inventaire.save()
            
            # Traiter les détails par période
            periodes_data = {}
            for key, value in request.POST.items():
                if key.startswith('periode_'):
                    periode = key.replace('periode_', '').replace('_', '-')
                    if value and value.isdigit():
                        periodes_data[periode] = int(value)
            
            # Mettre à jour ou créer les détails
            for periode, nombre in periodes_data.items():
                DetailInventairePeriode.objects.update_or_create(
                    inventaire=inventaire,
                    periode=periode,
                    defaults={'nombre_vehicules': nombre}
                )
            
            # Supprimer les périodes non soumises
            periodes_soumises = list(periodes_data.keys())
            inventaire.details_periodes.exclude(periode__in=periodes_soumises).delete()
            
            # Recalculer les totaux
            inventaire.recalculer_totaux()
            
            # Recalculer la recette associée si elle existe
            try:
                if hasattr(inventaire, 'recette'):
                    inventaire.recette.calculer_indicateurs()
                    inventaire.recette.save()
            except:
                pass
            
            # Journaliser l'action
            # log_user_action(
            #     request.user,
            #     "Modification inventaire",
            #     f"Inventaire modifié: {inventaire.poste.nom} - {inventaire.date}",
            #     request
            # )
            
            messages.success(request, "L'inventaire a été modifié avec succès.")
            return redirect('inventaire_detail', pk=inventaire.pk)
    else:
        form = InventaireJournalierForm(
            instance=inventaire,
            user=request.user
        )
        
        # Récupérer les détails existants
        details = inventaire.details_periodes.all()
        periodes_data = {d.periode: d.nombre_vehicules for d in details}
    
    context = {
        'form': form,
        'inventaire': inventaire,
        'periodes_data': periodes_data,
        'periodes_choices': [
            ('08h-09h', '08h-09h'), ('09h-10h', '09h-10h'), 
            ('10h-11h', '10h-11h'), ('11h-12h', '11h-12h'),
            ('12h-13h', '12h-13h'), ('13h-14h', '13h-14h'), 
            ('14h-15h', '14h-15h'), ('15h-16h', '15h-16h'),
            ('16h-17h', '16h-17h'), ('17h-18h', '17h-18h')
        ],
        'title': f'Modifier inventaire du {inventaire.date}'
    }
    
    return render(request, 'inventaire/modifier_inventaire.html', context)

@login_required
@require_permission('peut_gerer_inventaire')
def supprimer_inventaire(request, pk):
    """
    Permet de supprimer un inventaire
    Seuls les admins  peuvent supprimer
    """
    inventaire = get_object_or_404(InventaireJournalier, pk=pk)
    
    # Vérifier les permissions
    if not request.user.peut_acceder_poste(inventaire.poste):
        messages.error(request, "Vous n'avez pas accès à ce poste.")
        return HttpResponseForbidden("Accès non autorisé")
    
    # Seuls admin et chef de poste peuvent supprimer
    if not (request.user.is_admin):
        messages.error(request, "Seuls les administrateurs peuvent supprimer des inventaires.")
        return HttpResponseForbidden("Permission insuffisante")
    
    if request.method == 'POST':
        # Sauvegarder les infos pour le log
        info_inventaire = f"{inventaire.poste.nom} - {inventaire.date}"
        
        # Vérifier s'il y a une recette associée
        recette_associee = None
        try:
            recette_associee = inventaire.recette
        except:
            pass
        
        if recette_associee:
            # Dissocier la recette
            recette_associee.inventaire_associe = None
            recette_associee.recette_potentielle = None
            recette_associee.ecart = None
            recette_associee.taux_deperdition = None
            recette_associee.save()
            messages.info(request, "La recette associée a été dissociée.")
        
        # Supprimer l'inventaire et ses détails (cascade)
        inventaire.delete()
        
        # Journaliser l'action
        log_user_action(
            request.user,
            "Suppression inventaire",
            f"Inventaire supprimé: {info_inventaire}",
            request
        )
        
        messages.success(request, "L'inventaire a été supprimé avec succès.")
        return redirect('liste_inventaires')
    
    # Vérifier s'il y a une recette associée pour avertir
    has_recette = False
    try:
        has_recette = hasattr(inventaire, 'recette') and inventaire.recette is not None
    except:
        pass
    
    context = {
        'inventaire': inventaire,
        'has_recette': has_recette,
        'title': 'Confirmer la suppression'
    }
    
    return render(request, 'inventaire/confirmer_suppression.html', context)



@login_required
def saisir_recette_avec_tickets(request):
    """
    Version CORRIGÉE avec préservation de la date lors du rechargement automatique
    """
    
    # Vérifier permissions
    if not (request.user.is_chef_poste or request.user.is_admin):
        messages.error(request, "Vous n'avez pas la permission de saisir des recettes.")
        return HttpResponseForbidden("Accès non autorisé")
    
    # Déterminer les postes accessibles
    if hasattr(request.user, 'get_postes_accessibles'):
        postes = request.user.get_postes_accessibles()
    else:
        if request.user.acces_tous_postes or request.user.is_admin:
            postes = Poste.objects.filter(is_active=True, type='peage')
        elif request.user.poste_affectation:
            postes = Poste.objects.filter(id=request.user.poste_affectation.id)
        else:
            postes = Poste.objects.none()
    
    # ===================================================================
    # CORRECTION : Récupération améliorée du poste et date sélectionnés
    # ===================================================================
    
    poste_selectionne = None
    date_selectionnee = None
    stock_disponible = []
    afficher_formulaire_complet = False
    
    # Vérifier les paramètres GET pour le poste et la date
    if request.method == 'GET':
        poste_id = request.GET.get('poste')
        date_str = request.GET.get('date')
        
        # Traiter le poste sélectionné
        if poste_id:
            try:
                poste_selectionne = Poste.objects.get(id=poste_id)
                if not request.user.peut_acceder_poste(poste_selectionne):
                    messages.error(request, "Vous n'avez pas accès à ce poste")
                    poste_selectionne = None
            except (Poste.DoesNotExist, ValueError):
                pass
        
        # Traiter la date sélectionnée
        if date_str:
            try:
                from datetime import datetime
                date_selectionnee = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                # Si format invalide, utiliser date du jour
                date_selectionnee = timezone.now().date()
        
        # Afficher le formulaire complet si poste ET date sont sélectionnés
        if poste_selectionne and date_selectionnee:
            afficher_formulaire_complet = True
            stock_disponible = _obtenir_stock_disponible(poste_selectionne)
    
    # ===================================================================
    # TRAITEMENT POST (soumission du formulaire)
    # ===================================================================
    
    if request.method == 'POST':
        # Vérifier si c'est une confirmation
        if request.POST.get('action') == 'confirmer':
            return traiter_confirmation_recette_tickets(request)
        
        # Sinon, c'est une soumission du formulaire
        form = RecetteAvecTicketsForm(request.POST, user=request.user)
        formset = DetailVenteTicketFormSet(request.POST, prefix='tickets')
        
        if form.is_valid() and formset.is_valid():
            poste = form.cleaned_data['poste']
            date_recette = form.cleaned_data['date']
            observations = form.cleaned_data.get('observations', '')
            
            # Vérifier qu'une recette n'existe pas déjà
            if RecetteJournaliere.objects.filter(poste=poste, date=date_recette).exists():
                messages.error(request, f"Une recette existe déjà pour {poste.nom} le {date_recette}")
                return redirect('inventaire:liste_recettes')
            
            # Validation des séries vendues
            details_ventes = []
            montant_total_calcule = Decimal('0')
            erreurs_validation = []
            
            for i, form_detail in enumerate(formset):
                if form_detail.cleaned_data and not form_detail.cleaned_data.get('DELETE', False):
                    couleur = form_detail.cleaned_data['couleur']
                    num_premier = form_detail.cleaned_data['numero_premier']
                    num_dernier = form_detail.cleaned_data['numero_dernier']
                    
                    # Vérifier disponibilité
                    disponible, msg, tickets_prob = SerieTicket.verifier_disponibilite_serie_complete(
                        poste, couleur, num_premier, num_dernier
                    )
                    
                    if not disponible:
                        erreurs_validation.append(f"❌ Ligne {i+1}: {msg}")
                        
                        if tickets_prob:
                            for ticket in tickets_prob:
                                erreurs_validation.append(
                                    f"   → Série #{ticket['premier']}-{ticket['dernier']} "
                                    f"vendue le {ticket['date_vente'].strftime('%d/%m/%Y')} "
                                    f"au poste {ticket['poste']}"
                                )
                        continue
                                       
                    # Si OK, ajouter aux détails
                    nombre = num_dernier - num_premier + 1
                    montant = Decimal(nombre) * Decimal('500')
                    montant_total_calcule += montant
                    
                    details_ventes.append({
                        'couleur': couleur,
                        'numero_premier': num_premier,
                        'numero_dernier': num_dernier,
                        'nombre_tickets': nombre,
                        'montant': montant,
                        'ordre': i + 1
                    })
            
            # Vérifier qu'il y a au moins une série valide
            if not details_ventes:
                if not erreurs_validation:
                    erreurs_validation.append("Vous devez saisir au moins une série de tickets vendus")
            
            # Si erreurs, réafficher avec stock
            if erreurs_validation:
                for erreur in erreurs_validation:
                    messages.error(request, erreur)
                
                # Recharger le stock pour réaffichage
                stock_disponible = _obtenir_stock_disponible(poste)
                
                return render(request, 'inventaire/saisir_recette_tickets.html', {
                    'form': form,
                    'formset': formset,
                    'postes': postes,
                    'poste_selectionne': poste,
                    'date_selectionnee': date_recette,
                    'stock_disponible': stock_disponible,
                    'afficher_formulaire_complet': True,
                    'title': 'Saisir une recette journalière',
                    'recettes_recentes': RecetteJournaliere.objects.filter(
                        chef_poste=request.user
                    ).select_related('poste').order_by('-date')[:10]
                })
            
            # Vérifier le stock global
            stock_actuel = Decimal('0')
            try:
                stock = GestionStock.objects.get(poste=poste)
                stock_actuel = stock.valeur_monetaire
            except GestionStock.DoesNotExist:
                messages.warning(
                    request,
                    "⚠️ Aucun stock enregistré pour ce poste. "
                    "Le stock sera initialisé en négatif après cette vente."
                )
            
            # Vérifier si un inventaire existe
            inventaire_existe = InventaireJournalier.objects.filter(
                poste=poste,
                date=date_recette
            ).exists()
            
            if not inventaire_existe:
                messages.info(
                    request,
                    "ℹ️ Aucun inventaire n'a été saisi pour cette date. "
                    "Le taux de déperdition ne pourra pas être calculé."
                )
            
            # Stocker en session pour confirmation
            request.session['recette_tickets_confirmation'] = {
                'poste_id': poste.id,
                'date': date_recette.isoformat(),
                'montant_total': str(montant_total_calcule),
                'observations': observations,
                'inventaire_existe': inventaire_existe,
                'details_ventes': [
                    {
                        'couleur_id': d['couleur'].id,
                        'couleur_libelle': d['couleur'].libelle_affichage,
                        'numero_premier': d['numero_premier'],
                        'numero_dernier': d['numero_dernier'],
                        'nombre_tickets': d['nombre_tickets'],
                        'montant': str(d['montant']),
                        'ordre': d['ordre']
                    }
                    for d in details_ventes
                ],
                'stock_actuel': str(stock_actuel)
            }
            
            return redirect('inventaire:confirmation_recette_tickets')
    
    # ===================================================================
    # AFFICHAGE FORMULAIRE (GET)
    # ===================================================================
    else:
        # CORRECTION : Préparer les données initiales en priorité depuis GET
        initial_data = {}
        
        # Priorité 1 : Paramètres GET (pour préserver lors du rechargement)
        if poste_selectionne:
            initial_data['poste'] = poste_selectionne
        elif request.user.poste_affectation:
            initial_data['poste'] = request.user.poste_affectation
        
        if date_selectionnee:
            initial_data['date'] = date_selectionnee
        else:
            # Par défaut, utiliser la date du jour
            initial_data['date'] = timezone.now().date()
        
        form = RecetteAvecTicketsForm(initial=initial_data, user=request.user)
        
        # CORRECTION : Modifier le formulaire pour activer le rechargement auto
        # IMPORTANT : Utiliser JavaScript pour soumettre avec TOUS les champs
        form.fields['poste'].widget.attrs.update({
            'onchange': 'submitFormWithAllFields()',
            'class': 'form-select'
        })
        form.fields['date'].widget.attrs.update({
            'onchange': 'submitFormWithAllFields()',
            'class': 'form-control',
            'type': 'date'  # Important pour le format HTML5
        })
        
        # Créer le formset uniquement si formulaire complet affiché
        if afficher_formulaire_complet and stock_disponible:
            formset = DetailVenteTicketFormSet(prefix='tickets')
            
            # Filtrer les couleurs disponibles selon le stock
            couleurs_disponibles = CouleurTicket.objects.filter(
                id__in=[s['couleur'].id for s in stock_disponible]
            ).order_by('code_normalise')
            
            for form_detail in formset:
                form_detail.fields['couleur'].queryset = couleurs_disponibles
        else:
            formset = None
    
    # Statistiques
    from django.db.models import Sum
    recettes_query = RecetteJournaliere.objects.filter(
        chef_poste=request.user
    ).select_related('poste').order_by('-date')
    
    stats = {'total_mois': 0, 'moyenne_taux': 0}
    
    if recettes_query.exists():
        recettes_mois = recettes_query.filter(
            date__month=timezone.now().month,
            date__year=timezone.now().year
        )
        
        total_result = recettes_mois.aggregate(total=Sum('montant_declare'))['total']
        if total_result:
            stats['total_mois'] = float(total_result)
        
        taux_values = []
        for recette in recettes_query.filter(taux_deperdition__isnull=False):
            if recette.taux_deperdition is not None:
                try:
                    taux_values.append(float(recette.taux_deperdition))
                except (TypeError, ValueError, InvalidOperation):
                    continue
        
        if taux_values:
            stats['moyenne_taux'] = sum(taux_values) / len(taux_values)
    
    recettes_recentes = recettes_query[:10]
    
    context = {
        'form': form,
        'formset': formset,
        'postes': postes,
        'poste_selectionne': poste_selectionne,
        'date_selectionnee': date_selectionnee,
        'stock_disponible': stock_disponible,
        'afficher_formulaire_complet': afficher_formulaire_complet,
        'recettes_recentes': recettes_recentes,
        'stats': stats,
        'title': 'Saisir une recette journalière'
    }
    
    return render(request, 'inventaire/saisir_recette_tickets.html', context)

# ===================================================================
# FONCTION UTILITAIRE : Obtenir le stock disponible
# ===================================================================

def _obtenir_stock_disponible(poste):
    """
    Obtient le stock de tickets disponibles pour un poste
    Groupé par couleur avec les séries détaillées
    
    Returns:
        list: Liste de dictionnaires avec couleur, séries, totaux
    """
    from collections import defaultdict
    
    # Récupérer toutes les séries en stock pour ce poste
    series_stock = SerieTicket.objects.filter(
        poste=poste,
        statut='stock'
    ).select_related('couleur').order_by('couleur__code_normalise', 'numero_premier')
    
    # Grouper par couleur
    stock_par_couleur = defaultdict(lambda: {
        'couleur': None,
        'series': [],
        'total_tickets': 0,
        'valeur_totale': Decimal('0')
    })
    
    for serie in series_stock:
        couleur_code = serie.couleur.code_normalise
        
        if stock_par_couleur[couleur_code]['couleur'] is None:
            stock_par_couleur[couleur_code]['couleur'] = serie.couleur
        
        stock_par_couleur[couleur_code]['series'].append({
            'numero_premier': serie.numero_premier,
            'numero_dernier': serie.numero_dernier,
            'nombre_tickets': serie.nombre_tickets,
            'valeur': serie.valeur_monetaire
        })
        
        stock_par_couleur[couleur_code]['total_tickets'] += serie.nombre_tickets
        stock_par_couleur[couleur_code]['valeur_totale'] += serie.valeur_monetaire
    
    # Convertir en liste triée
    stock_liste = sorted(
        stock_par_couleur.values(),
        key=lambda x: x['couleur'].code_normalise if x['couleur'] else ''
    )
    
    return stock_liste


def traiter_confirmation_recette_tickets(request):
    """
    VERSION AMÉLIORÉE avec liaison des séries de tickets à l'historique
    Traite la confirmation de la recette avec tickets
    """
    
    # Récupérer les données de session (code existant conservé)
    data = request.session.get('recette_tickets_confirmation')
    
    if not data:
        messages.error(request, "Aucune recette en attente de confirmation")
        return redirect('inventaire:saisir_recette_avec_tickets')
    
    try:
        poste = Poste.objects.get(id=data['poste_id'])
        date_recette = datetime.fromisoformat(data['date']).date()
        montant_total = Decimal(data['montant_total'])
        
        with transaction.atomic():
            # 1. Créer la recette (code existant conservé)
            recette = RecetteJournaliere.objects.create(
                poste=poste,
                date=date_recette,
                montant_declare=montant_total,
                chef_poste=request.user,
                modifiable_par_chef=False,
                observations=data['observations']
            )
            
            # Associer l'inventaire s'il existe
            try:
                inventaire = InventaireJournalier.objects.get(
                    poste=poste,
                    date=date_recette
                )
                recette.inventaire_associe = inventaire
                recette.save()
            except InventaireJournalier.DoesNotExist:
                pass
            
            # ===== NOUVELLE LOGIQUE : Tracking des séries =====
            series_vendues = []  # Pour liaison à l'historique
            
            # 2. Créer les détails de vente et consommer les séries
            for detail_data in data['details_ventes']:
                couleur = CouleurTicket.objects.get(id=detail_data['couleur_id'])
                
                # Créer le détail de vente
                DetailVenteTicket.objects.create(
                    recette=recette,
                    couleur=couleur,
                    numero_premier=detail_data['numero_premier'],
                    numero_dernier=detail_data['numero_dernier'],
                    ordre=detail_data['ordre']
                )
                
                # Consommer la série de tickets
                success, msg, series = SerieTicket.consommer_serie(
                    poste,
                    couleur,
                    detail_data['numero_premier'],
                    detail_data['numero_dernier'],
                    recette
                )
                
                if not success:
                    raise Exception(f"Erreur consommation série: {msg}")
                
                # ===== NOUVEAU : Collecter les séries vendues =====
                series_vendues.extend(series)
            
            # 3. Mettre à jour le stock global (code existant conservé)
            stock, _ = GestionStock.objects.get_or_create(
                poste=poste,
                defaults={'valeur_monetaire': Decimal('0')}
            )
            
            stock_avant = stock.valeur_monetaire
            stock.valeur_monetaire -= montant_total
            stock.save()
            
            # 4. Créer l'historique
            historique = HistoriqueStock.objects.create(
                poste=poste,
                type_mouvement='DEBIT',
                montant=montant_total,
                nombre_tickets=int(montant_total / 500),
                stock_avant=stock_avant,
                stock_apres=stock.valeur_monetaire,
                effectue_par=request.user,
                reference_recette=recette,
                commentaire=f"Vente du {date_recette.strftime('%d/%m/%Y')} - {len(data['details_ventes'])} série(s)"
            )
            StockEvent.objects.create(
                    poste=poste,
                    event_type='VENTE',
                    event_datetime=datetime.combine(date_recette, datetime.now().time()),
                    montant_variation=-montant_total,  # NÉGATIF car c'est une sortie
                    nombre_tickets_variation=-int(montant_total / 500),  # NÉGATIF
                    stock_resultant=stock.valeur_monetaire,  # Stock APRÈS la vente
                    tickets_resultants=int(stock.valeur_monetaire / 500),
                    effectue_par=request.user,
                    reference_id=str(recette.id),
                    reference_type='RecetteJournaliere',
                    metadata={
                        'date_recette': str(date_recette),
                        'montant_declare': str(montant_total),
                        'series_vendues': [
                            {
                                'couleur': detail_data['couleur_libelle'],
                                'numero_premier': detail_data['numero_premier'],
                                'numero_dernier': detail_data['numero_dernier'],
                                'nombre_tickets': detail_data['nombre_tickets'],
                                'montant': detail_data['montant']
                            }
                            for detail_data in data['details_ventes']
                        ]
                    },
                    commentaire=f"Vente du {date_recette.strftime('%d/%m/%Y')} - {len(data['details_ventes'])} série(s)"
                )

            # ===== NOUVEAU : Associer les séries à l'historique =====
            if series_vendues:
                historique.associer_series_tickets(series_vendues)
            
            # 5. Journaliser l'action (code existant conservé)
            log_user_action(
                request.user,
                "Saisie recette avec tickets",
                f"Recette de {montant_total:,.0f} FCFA saisie pour {poste.nom} "
                f"le {date_recette.strftime('%d/%m/%Y')} - "
                f"{len(data['details_ventes'])} série(s) de tickets",
                request
            )
            
            # 6. Notifications aux administrateurs
            from accounts.models import UtilisateurSUPPER
            admins = UtilisateurSUPPER.objects.filter(
                    habilitation__in=[
                        'admin_principal',
                        'coord_psrr',
                        'serv_info',
                        'serv_emission'
                    ],
                is_active=True
            )
            
            # for admin in admins:
            #     NotificationUtilisateur.objects.create(
            #         destinataire=admin,
            #         expediteur=request.user,
            #         titre="Nouvelle recette saisie",
            #         message=(
            #             f"Recette de {montant_total:,.0f} FCFA saisie par "
            #             f"{request.user.nom_complet} pour le poste {poste.nom} "
            #             f"le {date_recette.strftime('%d/%m/%Y')}"
            #         ),
            #         type_notification='info'
            #     )
            
            # 7. Nettoyer la session
            del request.session['recette_tickets_confirmation']
            
            messages.success(
                request,
                f"✅ Recette enregistrée avec succès : {montant_total:,.0f} FCFA "
                f"({len(data['details_ventes'])} série(s) de tickets). "
                f"Stock restant: {stock.valeur_monetaire:,.0f} FCFA"
            )
            
            # Redirection selon le type d'utilisateur
            if request.user.is_admin:
                return redirect('inventaire:liste_recettes')
            else:
                return redirect(f"{reverse('inventaire:liste_recettes')}?poste={poste.id}")
    
    except Exception as e:
        logger.error(f"Erreur confirmation recette tickets: {str(e)}", exc_info=True)
        messages.error(request, f"❌ Erreur lors de l'enregistrement: {str(e)}")
        return redirect('inventaire:saisir_recette_avec_tickets')


@login_required
def confirmation_recette_tickets(request):
    """
    Page de confirmation avant validation de la recette avec tickets
    """
    
    data = request.session.get('recette_tickets_confirmation')
    
    if not data:
        messages.error(request, "Aucune recette en attente")
        return redirect('inventaire:saisir_recette_avec_tickets')
    
    poste = Poste.objects.get(id=data['poste_id'])
    montant_total = Decimal(data['montant_total'])
    stock_actuel = Decimal(data['stock_actuel'])
    
    # Préparer les détails enrichis
    details_enrichis = []
    for detail in data['details_ventes']:
        couleur = CouleurTicket.objects.get(id=detail['couleur_id'])
        details_enrichis.append({
            'couleur': couleur,
            'numero_premier': detail['numero_premier'],
            'numero_dernier': detail['numero_dernier'],
            'nombre_tickets': detail['nombre_tickets'],
            'montant': Decimal(detail['montant'])
        })
    
    context = {
        'poste': poste,
        'date': datetime.fromisoformat(data['date']).date(),
        'montant_total': montant_total,
        'observations': data['observations'],
        'details_ventes': details_enrichis,
        'stock_actuel': stock_actuel,
        'stock_apres': stock_actuel - montant_total,
        'stock_suffisant': stock_actuel >= montant_total,
        'is_admin': request.user.is_admin,
        'title': 'Confirmation de la recette'
    }
    
    return render(request, 'inventaire/confirmation_recette_tickets.html', context)

@login_required
def modifier_recette(request, pk):
    """
    Permet de modifier une recette existante
    Bloque les modifications par admin et après validation
    """
    recette = get_object_or_404(RecetteJournaliere, pk=pk)
    
    # BLOQUER modification pour les administrateurs
    if request.user.is_admin:
        messages.error(request, 
            "Les administrateurs ne peuvent pas modifier les recettes. "
            "Seuls les chefs de poste peuvent modifier leurs propres recettes non validées.")
        return redirect('inventaire:recette_detail', pk=pk)
    
    # Vérifier que c'est bien le chef qui a saisi
    if recette.chef_poste != request.user:
        messages.error(request, "Vous ne pouvez modifier que vos propres recettes.")
        return redirect('inventaire:recette_detail', pk=pk)
    
    # Vérifier si la recette est encore modifiable
    if not recette.modifiable_par_chef:
        messages.error(request, 
            "Cette recette a été validée et ne peut plus être modifiée. "
            "Les recettes validées sont définitives pour garantir l'intégrité des données.")
        return redirect('inventaire:recette_detail', pk=pk)
    
    # Vérifier l'accès au poste
    if not request.user.peut_acceder_poste(recette.poste):
        messages.error(request, "Vous n'avez pas accès à ce poste.")
        return HttpResponseForbidden("Accès non autorisé")
    
    if request.method == 'POST':
        form = RecetteJournaliereForm(
            request.POST,
            instance=recette,
            user=request.user
        )
        
        if form.is_valid():
            # Empêcher la modification du montant si déjà déduit du stock
            ancien_montant = recette.montant_declare
            nouveau_montant = form.cleaned_data['montant_declare']
            
            if ancien_montant != nouveau_montant:
                messages.error(request, 
                    "Le montant ne peut pas être modifié après validation. "
                    "Veuillez contacter un administrateur si nécessaire.")
                return redirect('inventaire:modifier_recette', pk=pk)
            
            recette = form.save(commit=False)
            recette.derniere_modification_par = request.user
            
            # Re-lier l'inventaire si demandé
            if 'lier_inventaire' in request.POST:
                try:
                    inventaire = InventaireJournalier.objects.get(
                        poste=recette.poste,
                        date=recette.date
                    )
                    recette.inventaire_associe = inventaire
                except InventaireJournalier.DoesNotExist:
                    messages.info(request, "Aucun inventaire trouvé pour cette date.")
            
            recette.save()
            
            # Journaliser
            # log_user_action(
            #     request.user,
            #     "Modification recette",
            #     f"Recette modifiée: {recette.poste.nom} - {recette.date}",
            #     request
            # )
            
            messages.success(request, "La recette a été modifiée avec succès.")
            
            if request.user.is_admin:
                return redirect('inventaire:liste_recettes')
            else:
                return redirect(f"{reverse('inventaire:liste_recettes')}?poste={recette.poste.id}")
    else:
        form = RecetteJournaliereForm(
            instance=recette,
            user=request.user
        )
    
    context = {
        'form': form,
        'recette': recette,
        'title': f'Modifier recette du {recette.date}',
        'is_modifiable': recette.modifiable_par_chef,
        'warning_message': "Attention : Les modifications sont limitées après validation."
    }
    
    return render(request, 'inventaire/modifier_recette.html', context)
@login_required
def supprimer_recette(request, pk):
    """
    Permet de supprimer une recette
    """
    recette = get_object_or_404(RecetteJournaliere, pk=pk)
    
    # Vérifier les permissions
    if not (request.user.is_admin):
        messages.error(request, "Seuls les administrateurs  peuvent supprimer des recettes.")
        return HttpResponseForbidden("Permission insuffisante")
    
    if not request.user.peut_acceder_poste(recette.poste):
        messages.error(request, "Vous n'avez pas accès à ce poste.")
        return HttpResponseForbidden("Accès non autorisé")
    
    if request.method == 'POST':
        # Sauvegarder les infos
        info_recette = f"{recette.poste.nom} - {recette.date} - {recette.montant_declare} FCFA"
        
        # Supprimer
        recette.delete()
        
        # Journaliser
        log_user_action(
            request.user,
            "Suppression recette",
            f"Recette supprimée: {info_recette}",
            request
        )
        
        messages.success(request, "La recette a été supprimée avec succès.")
        if request.user.is_admin:
                    return redirect('inventaire:liste_recettes')
        else:
                    # Rediriger vers la liste filtrée sur son poste
                    return redirect(f"{reverse('inventaire:liste_recettes')}?poste={recette.poste.id}")
    
    context = {
        'recette': recette,
        'title': 'Confirmer la suppression'
    }
    
    return render(request, 'inventaire/confirmer_suppression_recette.html', context)

from django.views.generic import ListView, DetailView

class RecetteListView(LoginRequiredMixin, ListView):
    """
    Vue pour lister les recettes avec filtres avancés
    
    MISE À JOUR - Permissions granulaires:
    - Permission requise: peut_voir_liste_recettes_peage
    - Remplace: self.request.user.is_admin
    - Par: user_has_acces_tous_postes(user) et is_admin_user(user)
    - Ajout de logs détaillés pour chaque action
    
    RÈGLES MÉTIER:
    - Admin/Services centraux: voient tous les postes
    - Chef de poste: voit uniquement son poste d'affectation
    - Agent inventaire: voit uniquement son poste d'affectation
    """
    model = RecetteJournaliere
    template_name = 'inventaire/liste_recettes.html'
    context_object_name = 'recettes'
    paginate_by = 25
    
    def dispatch(self, request, *args, **kwargs):
        """
        Vérification des permissions avant d'accéder à la vue
        NOUVEAU: Ajout de la vérification granulaire
        """
        user = request.user
        
        # Vérifier l'authentification (géré par LoginRequiredMixin mais on double-check)
        if not user.is_authenticated:
            return redirect('accounts:login')
        
        # =========================================
        # VÉRIFICATION PERMISSION GRANULAIRE
        # =========================================
        if not has_permission(user, 'peut_voir_liste_recettes_peage'):
            log_user_action(
                user,
                "ACCÈS REFUSÉ - Liste recettes",
                f"Permission manquante: peut_voir_liste_recettes_peage | "
                f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | "
                f"IP: {request.META.get('REMOTE_ADDR')}",
                request
            )
            messages.error(request, _("Vous n'avez pas la permission de voir les recettes."))
            return redirect('common:dashboard')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        """
        Récupère les recettes filtrées selon les permissions et les filtres GET
        
        MISE À JOUR:
        - Remplace: self.request.user.is_admin
        - Par: user_has_acces_tous_postes(self.request.user)
        """
        user = self.request.user
        queryset = super().get_queryset()
        
        # =========================================
        # FILTRAGE SELON LES PERMISSIONS
        # (Remplace: if not self.request.user.is_admin)
        # =========================================
        
        if not user_has_acces_tous_postes(user):
            # Utilisateur avec accès limité (chef_peage, agent_inventaire, etc.)
            poste_filter = self.request.GET.get('poste')
            
            if not poste_filter:
                # Par défaut, montrer seulement les recettes de son poste
                if user.poste_affectation:
                    queryset = queryset.filter(poste=user.poste_affectation)
                else:
                    # Si pas de poste d'affectation, montrer ses propres recettes saisies
                    queryset = queryset.filter(chef_poste=user)
            else:
                # Vérifier que le poste demandé est bien accessible
                if user.poste_affectation and str(user.poste_affectation.id) == poste_filter:
                    queryset = queryset.filter(poste_id=poste_filter)
                else:
                    # Accès non autorisé à ce poste - filtrer sur son propre poste
                    log_user_action(
                        user,
                        "TENTATIVE ACCÈS - Recettes poste non autorisé",
                        f"Poste demandé: {poste_filter} | "
                        f"Poste affectation: {getattr(user.poste_affectation, 'id', 'Aucun')}",
                        self.request
                    )
                    if user.poste_affectation:
                        queryset = queryset.filter(poste=user.poste_affectation)
                    else:
                        queryset = queryset.filter(chef_poste=user)
        
        # Jointures pour optimiser les requêtes
        queryset = queryset.select_related('poste', 'chef_poste', 'inventaire_associe')
        
        # =========================================
        # APPLICATION DES FILTRES GET
        # =========================================
        
        filters = self.request.GET
        
        # Filtre par poste (pour les admins uniquement, sinon déjà filtré ci-dessus)
        if user_has_acces_tous_postes(user) and filters.get('poste'):
            queryset = queryset.filter(poste_id=filters.get('poste'))
        
        # Filtre par période
        periode = filters.get('periode', 'all')
        if periode == 'jour':
            date_str = filters.get('date')
            if date_str:
                try:
                    date_filtre = datetime.strptime(date_str, '%Y-%m-%d').date()
                    queryset = queryset.filter(date=date_filtre)
                except ValueError:
                    pass
        elif periode == 'semaine':
            debut_semaine = timezone.now().date() - timedelta(days=timezone.now().weekday())
            queryset = queryset.filter(date__gte=debut_semaine)
        elif periode == 'mois':
            mois = filters.get('mois')
            annee = filters.get('annee')
            if mois and annee:
                queryset = queryset.filter(date__month=mois, date__year=annee)
        
        # Filtre par taux de déperdition
        taux_filtre = filters.get('taux_filtre')
        if taux_filtre == 'bon':
            queryset = queryset.filter(taux_deperdition__gt=-10)
        elif taux_filtre == 'moyen':
            queryset = queryset.filter(taux_deperdition__lte=-10, taux_deperdition__gt=-30)
        elif taux_filtre == 'mauvais':
            queryset = queryset.filter(taux_deperdition__lte=-30)
        
        # Recherche textuelle
        search = filters.get('search')
        if search:
            queryset = queryset.filter(
                Q(poste__nom__icontains=search) |
                Q(poste__code__icontains=search) |
                Q(chef_poste__nom_complet__icontains=search)
            )
        
        # Tri
        order = filters.get('order', '-date')
        queryset = queryset.order_by(order)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """
        Prépare le contexte pour le template
        
        MISE À JOUR:
        - Remplace: self.request.user.is_admin
        - Par: user_has_acces_tous_postes(self.request.user)
        - PRÉSERVE toutes les variables de contexte existantes
        """
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # =========================================
        # CONTEXTE POUR VUE LIMITÉE AU POSTE
        # (Remplace: if not self.request.user.is_admin)
        # =========================================
        
        if not user_has_acces_tous_postes(user) and user.poste_affectation:
            poste_filter = self.request.GET.get('poste')
            if poste_filter == str(user.poste_affectation.id):
                context['viewing_own_poste'] = True
                context['poste_name'] = user.poste_affectation.nom
        
        # Ajouter les filtres actuels
        context['current_filters'] = self.request.GET.dict()
        context['jours_non_declares'] = self.calculer_jours_non_declares()
        
        # =========================================
        # STATISTIQUES GLOBALES
        # =========================================
        
        all_recettes = self.get_queryset()
        
        # Calcul du total montant sécurisé
        total_montant = all_recettes.aggregate(
            Sum('montant_declare')
        )['montant_declare__sum']
        
        # Calcul de la moyenne des taux sécurisé
        moyenne_taux = None
        taux_queryset = all_recettes.filter(
            taux_deperdition__isnull=False
        ).values_list('taux_deperdition', flat=True)
        
        if taux_queryset:
            taux_values = []
            for taux in taux_queryset:
                if taux is not None:
                    try:
                        taux_values.append(float(taux))
                    except (TypeError, ValueError, decimal.InvalidOperation):
                        continue
            
            if taux_values:
                moyenne_taux = sum(taux_values) / len(taux_values)
        
        # Variable stats - PRÉSERVÉE pour compatibilité template
        context['stats'] = {
            'total_recettes': all_recettes.count(),
            'total_montant': float(total_montant) if total_montant else 0,
            'moyenne_taux': moyenne_taux,
            'recettes_jour': all_recettes.filter(date=timezone.now().date()).count(),
        }
        
        # =========================================
        # LISTE DES POSTES POUR LE FILTRE
        # (Remplace: if self.request.user.is_admin)
        # =========================================
        
        if user_has_acces_tous_postes(user):
            # Admin et services centraux: tous les postes péage actifs
            context['postes'] = Poste.objects.filter(
                is_active=True, 
                type='peage'
            ).order_by('nom')
        else:
            # Utilisateurs limités: seulement leurs postes accessibles
            context['postes'] = get_postes_accessibles(user, type_poste='peage')
        
        # Mois disponibles pour le filtre
        dates = all_recettes.dates('date', 'month', order='DESC')
        context['mois_disponibles'] = [
            {'mois': d.month, 'annee': d.year, 'label': d.strftime('%B %Y')}
            for d in dates
        ]
        
        # =========================================
        # PERMISSIONS POUR AFFICHAGE CONDITIONNEL
        # =========================================
        
        context['can_edit'] = has_permission(user, 'peut_saisir_recette_peage')
        context['can_view_stats'] = has_permission(user, 'peut_voir_stats_recettes_peage')
        context['can_import'] = has_permission(user, 'peut_importer_recettes_peage')
        context['is_admin'] = user_has_acces_tous_postes(user)
        
        # =========================================
        # LOG DE LA CONSULTATION
        # =========================================
        
        filtres_actifs = []
        if self.request.GET.get('poste'):
            filtres_actifs.append(f"poste={self.request.GET.get('poste')}")
        if self.request.GET.get('periode'):
            filtres_actifs.append(f"periode={self.request.GET.get('periode')}")
        if self.request.GET.get('taux_filtre'):
            filtres_actifs.append(f"taux={self.request.GET.get('taux_filtre')}")
        
        log_user_action(
            user,
            "Consultation liste recettes",
            f"Total résultats: {context['stats']['total_recettes']} | "
            f"Filtres: {', '.join(filtres_actifs) if filtres_actifs else 'Aucun'} | "
            f"Accès: {'Global' if user_has_acces_tous_postes(user) else 'Limité à ' + str(getattr(user.poste_affectation, 'nom', 'N/A'))}",
            self.request
        )
        
        return context
    
    def calculer_jours_non_declares(self):
        """
        Calcule les jours sans déclaration ET estime les recettes manquantes
        UNIQUEMENT pour les jours PASSÉS (jusqu'à aujourd'hui)
        
        MISE À JOUR:
        - Utilise get_postes_accessibles() pour filtrer les postes selon les permissions
        """
        from django.db.models import Q
        from datetime import datetime, timedelta
        from inventaire.services.forecasting_service import ForecastingService
        import calendar
        
        user = self.request.user
        filters = self.request.GET
        periode = filters.get('periode', 'mois')
        poste_id = filters.get('poste')
        
        # Déterminer la période d'analyse
        if periode == 'jour':
            date_str = filters.get('date')
            if date_str:
                date_debut = date_fin = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                date_debut = date_fin = date.today()
        elif periode == 'semaine':
            date_fin = date.today()
            date_debut = date_fin - timedelta(days=7)
        elif periode == 'mois':
            mois = int(filters.get('mois', date.today().month))
            annee = int(filters.get('annee', date.today().year))
            date_debut = date(annee, mois, 1)
            # IMPORTANT : Ne pas dépasser aujourd'hui
            date_fin_mois = date(annee, mois, calendar.monthrange(annee, mois)[1])
            date_fin = min(date_fin_mois, date.today())
        else:  # all
            date_fin = date.today()
            date_debut = date_fin - timedelta(days=30)
        
        # S'assurer qu'on ne dépasse jamais la date du jour
        if date_fin > date.today():
            date_fin = date.today()
        
        # =========================================
        # FILTRAGE DES POSTES SELON PERMISSIONS
        # (Remplace la logique basée sur is_admin)
        # =========================================
        
        if user_has_acces_tous_postes(user):
            # Admin: tous les postes péage actifs
            postes_query = Poste.objects.filter(is_active=True, type='peage')
        else:
            # Utilisateur limité: seulement ses postes accessibles
            postes_query = get_postes_accessibles(user, type_poste='peage')
        
        # Appliquer le filtre de poste si spécifié
        if poste_id:
            # Vérifier que le poste demandé est accessible
            if user_has_acces_tous_postes(user):
                postes_query = postes_query.filter(id=poste_id)
            elif user.poste_affectation and str(user.poste_affectation.id) == poste_id:
                postes_query = postes_query.filter(id=poste_id)
            # Sinon, ignorer le filtre (l'utilisateur verra son poste par défaut)
        
        resultats = []
        total_estimation_manquante = Decimal('0')
        
        for poste in postes_query:
            # Jours avec recettes déclarées
            jours_declares = set(
                RecetteJournaliere.objects.filter(
                    poste=poste,
                    date__range=[date_debut, date_fin]
                ).values_list('date', flat=True)
            )
            
            # Tous les jours de la période (UNIQUEMENT jusqu'à aujourd'hui)
            jours_periode = set()
            current_date = date_debut
            while current_date <= date_fin:
                jours_periode.add(current_date)
                current_date += timedelta(days=1)
            
            # Jours manquants (seulement les jours passés)
            jours_manquants = sorted(jours_periode - jours_declares)
            
            if jours_manquants:
                estimation_manquante = Decimal('0')
                details_estimations = []
                
                try:
                    # Générer des prévisions RÉTROACTIVES pour les jours manquants
                    date_reference = min(jours_manquants) - timedelta(days=1)
                    nb_jours = (max(jours_manquants) - min(jours_manquants)).days + 1
                    
                    resultats_forecast = ForecastingService.prevoir_recettes(
                        poste,
                        nb_jours_future=nb_jours,
                        date_reference=date_reference
                    )
                    
                    if resultats_forecast['success']:
                        df_prev = resultats_forecast['predictions']
                        
                        for jour_manquant in jours_manquants:
                            prev_jour = df_prev[df_prev['date'].dt.date == jour_manquant]
                            
                            if not prev_jour.empty:
                                montant_estime = Decimal(str(prev_jour['montant_prevu'].values[0]))
                                estimation_manquante += montant_estime
                                
                                details_estimations.append({
                                    'date': jour_manquant,
                                    'montant_estime': float(montant_estime)
                                })
                            else:
                                moyenne_prev = Decimal(str(df_prev['montant_prevu'].mean()))
                                estimation_manquante += moyenne_prev
                                
                                details_estimations.append({
                                    'date': jour_manquant,
                                    'montant_estime': float(moyenne_prev)
                                })
                    
                    # Fallback si forecasting ne marche pas
                    if estimation_manquante == 0:
                        from django.db.models import Avg
                        
                        fin_moyenne = date_debut - timedelta(days=1)
                        debut_moyenne = fin_moyenne - timedelta(days=30)
                        
                        moyenne = RecetteJournaliere.objects.filter(
                            poste=poste,
                            date__range=[debut_moyenne, fin_moyenne]
                        ).aggregate(moy=Avg('montant_declare'))['moy'] or Decimal('0')
                        
                        estimation_manquante = moyenne * len(jours_manquants)
                        
                        for jour_manquant in jours_manquants:
                            details_estimations.append({
                                'date': jour_manquant,
                                'montant_estime': float(moyenne)
                            })
                        
                except Exception as e:
                    import logging
                    logger = logging.getLogger('supper')
                    logger.error(f"Erreur estimation recettes manquantes {poste.nom}: {str(e)}")
                    
                    # Fallback final : moyenne simple
                    from django.db.models import Avg
                    moyenne = RecetteJournaliere.objects.filter(
                        poste=poste,
                        date__lt=date_debut
                    ).order_by('-date')[:30].aggregate(moy=Avg('montant_declare'))['moy'] or Decimal('0')
                    
                    estimation_manquante = moyenne * len(jours_manquants)
                    
                    for jour_manquant in jours_manquants:
                        details_estimations.append({
                            'date': jour_manquant,
                            'montant_estime': float(moyenne)
                        })
                
                total_estimation_manquante += estimation_manquante
                
                resultats.append({
                    'poste': poste,
                    'nombre_jours': len(jours_manquants),
                    'jours': jours_manquants,
                    'pourcentage': (len(jours_manquants) / len(jours_periode)) * 100 if len(jours_periode) > 0 else 0,
                    'estimation_manquante': float(estimation_manquante),
                    'details_estimations': details_estimations,
                    'methode_calcul': 'forecasting' if details_estimations else 'moyenne'
                })
        
        # Trier par estimation décroissante
        resultats.sort(key=lambda x: x['estimation_manquante'], reverse=True)
        
        return {
            'par_poste': resultats,
            'total_jours_manquants': sum(r['nombre_jours'] for r in resultats),
            'total_estimation_manquante': float(total_estimation_manquante),
            'periode_debut': date_debut,
            'periode_fin': date_fin,
            'date_limite': date.today()
        }

class RecetteDetailView(LoginRequiredMixin, DetailView):
    """Vue pour afficher le détail d'une recette"""
    model = RecetteJournaliere
    template_name = 'inventaire/recette_detail.html'
    context_object_name = 'recette'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Vérifier les permissions de modification
        recette = self.object
        context['peut_modifier'] = (
            self.request.user.is_admin or
            recette.chef_poste == self.request.user
        )
        
        # Ajouter l'inventaire associé si existe
        if recette.inventaire_associe:
            context['inventaire'] = recette.inventaire_associe
            context['details_inventaire'] = recette.inventaire_associe.details_periodes.all()
        peut_voir_taux = False
        
        if self.request.user.is_admin:
            peut_voir_taux = True
        elif self.request.user.is_chef_poste:
                peut_voir_taux = True
        
        context['peut_voir_taux'] = peut_voir_taux
        
        # Si peut voir, calculer les statistiques hebdomadaires
        if peut_voir_taux and self.request.user.is_chef_poste:
            debut_semaine = recette.date - timedelta(days=recette.date.weekday())
            fin_semaine = debut_semaine + timedelta(days=6)
            
            recettes_semaine = RecetteJournaliere.objects.filter(
                poste=recette.poste,
                date__range=[debut_semaine, fin_semaine]
            ).exclude(
                # Exclure les jours impertinents
                date__in=ConfigurationJour.objects.filter(
                    statut='impertinent',
                    date__range=[debut_semaine, fin_semaine]
                ).values_list('date', flat=True)
            )
            
            # Calculer le taux moyen hebdomadaire
            taux_list = [r.taux_deperdition for r in recettes_semaine 
                        if r.taux_deperdition is not None]
            
            context['taux_hebdomadaire'] = {
                'moyenne': sum(taux_list) / len(taux_list) if taux_list else None,
                'nombre_jours': len(taux_list),
                'debut': debut_semaine,
                'fin': fin_semaine
            }
        
        return context
    
@login_required
@require_permission('peut_gerer_inventaire')
def liste_inventaires_mensuels(request):
    """
    Liste des inventaires mensuels programmés
    """
    # Filtrer selon les permissions
    if request.user.is_admin:
        inventaires = InventaireMensuel.objects.all()
    else:
        postes = request.user.get_postes_accessibles()
        inventaires = InventaireMensuel.objects.filter(
            poste__in=postes
        )
    
    # Filtres
    mois = request.GET.get('mois')
    annee = request.GET.get('annee')
    poste_id = request.GET.get('poste')
    motif = request.GET.get('motif')
    
    if mois:
        inventaires = inventaires.filter(mois=mois)
    if annee:
        inventaires = inventaires.filter(annee=annee)
    if poste_id:
        inventaires = inventaires.filter(poste_id=poste_id)
    if motif:
        inventaires = inventaires.filter(motif=motif)
    
    # Ordering
    inventaires = inventaires.select_related(
        'poste', 'cree_par', 'programmation'
    ).order_by('-annee', '-mois', 'poste__nom')
    
    # Statistiques
    stats = {
        'total': inventaires.count(),
        'actifs': inventaires.filter(actif=True).count(),
        'mois_courant': inventaires.filter(
            mois=str(timezone.now().month).zfill(2),
            annee=timezone.now().year
        ).count()
    }
    
    context = {
        'inventaires': inventaires,
        'stats': stats,
        'postes': Poste.objects.filter(is_active=True, type='peage'),
        'current_year': timezone.now().year,
        'title': 'Inventaires mensuels'
    }
    
    return render(request, 'inventaire/liste_inventaires_mensuels.html', context)


@login_required
@require_permission('peut_gerer_inventaire')
def detail_inventaire_mensuel(request, pk):
    """
    Détail d'un inventaire mensuel avec tous les inventaires journaliers associés
    """
    inventaire_mensuel = get_object_or_404(InventaireMensuel, pk=pk)
    
    # Vérifier l'accès au poste
    if not request.user.peut_acceder_poste(inventaire_mensuel.poste):
        messages.error(request, "Vous n'avez pas accès à ce poste.")
        return HttpResponseForbidden("Accès non autorisé")
    
    # Calculer les dates du mois
    annee = inventaire_mensuel.annee
    mois = int(inventaire_mensuel.mois)
    date_debut = date(annee, mois, 1)
    dernier_jour = calendar.monthrange(annee, mois)[1]
    date_fin = date(annee, mois, dernier_jour)
    
    # Récupérer les inventaires journaliers du mois
    inventaires_journaliers = InventaireJournalier.objects.filter(
        poste=inventaire_mensuel.poste,
        date__range=[date_debut, date_fin]
    ).select_related('agent_saisie').prefetch_related('details_periodes')
    
    # Récupérer les recettes du mois
    recettes = RecetteJournaliere.objects.filter(
        poste=inventaire_mensuel.poste,
        date__range=[date_debut, date_fin]
    ).select_related('chef_poste', 'inventaire_associe')
    
    # Calculer les statistiques
    stats = {
        'total_vehicules': inventaires_journaliers.aggregate(
            total=Sum('total_vehicules')
        )['total'] or 0,
        'total_recettes_declarees': recettes.aggregate(
            total=Sum('montant_declare')
        )['total'] or float('0'),
        'total_recettes_potentielles': recettes.aggregate(
            total=Sum('recette_potentielle')
        )['total'] or float('0'),
        'nombre_jours_saisis': inventaires_journaliers.count(),
        'nombre_jours_impertinents': ConfigurationJour.objects.filter(
            date__range=[date_debut, date_fin],
            statut='impertinent'
        ).count()
    }
    
    # Calculer le taux moyen
    if stats['total_recettes_potentielles'] > 0:
        ecart = stats['total_recettes_declarees'] - stats['total_recettes_potentielles']
        stats['taux_deperdition_moyen'] = (ecart / stats['total_recettes_potentielles']) * 100
    else:
        stats['taux_deperdition_moyen'] = None
    
    # Créer un calendrier du mois avec les données
    cal = calendar.monthcalendar(annee, mois)
    calendrier_data = []
    
    for semaine in cal:
        semaine_data = []
        for jour in semaine:
            if jour == 0:
                semaine_data.append(None)
            else:
                date_jour = date(annee, mois, jour)
                inventaire = inventaires_journaliers.filter(date=date_jour).first()
                recette = recettes.filter(date=date_jour).first()
                config = ConfigurationJour.objects.filter(date=date_jour).first()
                
                semaine_data.append({
                    'jour': jour,
                    'date': date_jour,
                    'inventaire': inventaire,
                    'recette': recette,
                    'config': config,
                    'is_impertinent': config and config.statut == 'impertinent'
                })
        calendrier_data.append(semaine_data)
    
    context = {
        'inventaire_mensuel': inventaire_mensuel,
        'inventaires_journaliers': inventaires_journaliers,
        'recettes': recettes,
        'stats': stats,
        'calendrier': calendrier_data,
        'title': f'Inventaire mensuel - {inventaire_mensuel.get_mois_display()} {annee}'
    }
    
    return render(request, 'inventaire/detail_inventaire_mensuel.html', context)
@login_required
@require_permission('peut_gerer_inventaire')
def consolider_inventaire_mensuel(request):
    """
    Consolide les données d'un mois pour créer/mettre à jour les statistiques
    """
    if request.method != 'POST':
        return redirect('inventaire:liste_inventaires_mensuels')
    
    poste_id = request.POST.get('poste_id')
    mois = request.POST.get('mois')
    annee = request.POST.get('annee')
    
    if not all([poste_id, mois, annee]):
        messages.error(request, "Paramètres manquants pour la consolidation.")
        return redirect('liste_inventaires_mensuels')
    
    try:
        poste = Poste.objects.get(pk=poste_id)
        
        # Vérifier l'accès
        if not request.user.peut_acceder_poste(poste):
            messages.error(request, "Vous n'avez pas accès à ce poste.")
            return HttpResponseForbidden("Accès non autorisé")
        
        # Calculer les dates
        annee_int = int(annee)
        mois_int = int(mois)
        date_debut = date(annee_int, mois_int, 1)
        dernier_jour = calendar.monthrange(annee_int, mois_int)[1]
        date_fin = date(annee_int, mois_int, dernier_jour)
        
        # Créer ou mettre à jour l'inventaire mensuel
        inventaire_mensuel, created = InventaireMensuel.objects.update_or_create(
            poste=poste,
            mois=str(mois_int).zfill(2),
            annee=annee_int,
            defaults={
                'titre': f"Inventaire {poste.nom} - {mois}/{annee}",
                'cree_par': request.user if created else None,
                'actif': True
            }
        )
        
        # Consolider les données
        inventaire_mensuel.consolider_donnees()
        
        # Créer les statistiques périodiques
        StatistiquesPeriodiques.calculer_statistiques_periode(
            poste=poste,
            type_periode='mensuelle',
            date_debut=date_debut,
            date_fin=date_fin
        )
        
        # Journaliser
        log_user_action(
            request.user,
            "Consolidation inventaire mensuel",
            f"Consolidation effectuée pour {poste.nom} - {mois}/{annee}",
            request
        )
        
        messages.success(request, f"Consolidation réussie pour {poste.nom} - {mois}/{annee}")
        return redirect('detail_inventaire_mensuel', pk=inventaire_mensuel.pk)
        
    except Poste.DoesNotExist:
        messages.error(request, "Poste introuvable.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la consolidation: {str(e)}")
    
    return redirect('liste_inventaires_mensuels')

class ConfigurationJourListView(AdminRequiredMixin, ListView):
    """Liste des configurations de jours"""
    model = ConfigurationJour
    template_name = 'inventaire/config_jour_list.html'
    context_object_name = 'configurations'
    paginate_by = 31  # Un mois par page
    
    def get_queryset(self):
        queryset = ConfigurationJour.objects.select_related('cree_par')
        
        # Filtre par mois/année
        mois = self.request.GET.get('mois')
        annee = self.request.GET.get('annee')
        
        if mois and annee:
            try:
                mois = int(mois)
                annee = int(annee)
                queryset = queryset.filter(date__month=mois, date__year=annee)
            except ValueError:
                pass
        
        return queryset.order_by('-date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Dates pour le filtre
        now = timezone.now()
        context.update({
            'current_month': now.month,
            'current_year': now.year,
            'months': [
                (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
                (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
                (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre')
            ],
            'years': range(now.year - 2, now.year + 2),
        })
        
        return context


# ===================================================================
# API VIEWS POUR LES CALCULS EN TEMPS RÉEL
# ===================================================================

class CalculAutomatiqueAPIView(InventaireMixin, View):
    """API pour les calculs automatiques d'inventaire"""
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Non authentifié'}, status=401)
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            periodes_data = data.get('periodes', {})
            
            total_vehicules = 0
            nombre_periodes = 0
            
            for periode_code, nombre_str in periodes_data.items():
                if nombre_str and nombre_str.strip():
                    try:
                        nombre = int(nombre_str)
                        if nombre > 0:
                            total_vehicules += nombre
                            nombre_periodes += 1
                    except ValueError:
                        continue
            
            moyenne_horaire = total_vehicules / nombre_periodes if nombre_periodes > 0 else 0
            estimation_24h = moyenne_horaire * 24
            
            from django.conf import settings
            config = getattr(settings, 'SUPPER_CONFIG', {})
            tarif = config.get('TARIF_VEHICULE_LEGER', 500)
            pourcentage_legers = config.get('POURCENTAGE_VEHICULES_LEGERS', 75)
            
            recette_potentielle = (estimation_24h * pourcentage_legers * tarif) / 100
            
            return JsonResponse({
                'success': True,
                'total_vehicules': total_vehicules,
                'nombre_periodes': nombre_periodes,
                'moyenne_horaire': round(moyenne_horaire, 2),
                'estimation_24h': round(estimation_24h, 2),
                'recette_potentielle': round(recette_potentielle, 2),
            })
            
        except Exception as e:
            logger.error(f"Erreur calcul automatique: {str(e)}")
            return JsonResponse({'error': 'Erreur de calcul'}, status=500)


# ===================================================================
# REDIRECTIONS VERS L'ADMIN DJANGO
# ===================================================================

@login_required
def redirect_to_inventaires_admin(request):
    """Redirection vers la gestion des inventaires dans l'admin Django"""
    
    if not _check_admin_permission(request.user):
        messages.error(request, "Accès non autorisé à la gestion des inventaires.")
        return redirect('inventaire:inventaire_list')
    
    log_user_action(request.user, "Accès admin inventaires", "", request)
    messages.info(request, "Redirection vers la gestion des inventaires.")
    
    return redirect('/admin/inventaire/inventairejournalier/')


@login_required
def redirect_to_recettes_admin(request):
    """Redirection vers la gestion des recettes dans l'admin Django"""
    
    if not _check_admin_permission(request.user):
        messages.error(request, "Accès non autorisé à la gestion des recettes.")
        return redirect('inventaire:inventaire_list')
    
    log_user_action(request.user, "Accès admin recettes", "", request)
    messages.info(request, "Redirection vers la gestion des recettes.")
    
    return redirect('/admin/inventaire/recettejournaliere/')


@login_required
def redirect_to_config_jours_admin(request):
    """Redirection vers la configuration des jours dans l'admin Django"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé à la configuration des jours."))
        return redirect('inventaire:config_jour_list')
    
    _log_inventaire_action(request, "Accès admin configuration jours")
    messages.info(request, _("Redirection vers la configuration des jours."))
    
    return redirect('/admin/inventaire/configurationjour/')


@login_required
def redirect_to_statistiques_admin(request):
    """Redirection vers les statistiques dans l'admin Django"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé aux statistiques."))
        return redirect('inventaire:inventaire_list')
    
    _log_inventaire_action(request, "Accès admin statistiques")
    messages.info(request, _("Redirection vers les statistiques."))
    
    return redirect('/admin/inventaire/statistiquesperiodiques/')


@login_required
def redirect_to_edit_inventaire_admin(request, inventaire_id):
    """Redirection vers l'édition d'un inventaire spécifique dans l'admin Django"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé à l'édition des inventaires."))
        return redirect('inventaire:inventaire_list')
    
    try:
        inventaire = InventaireJournalier.objects.get(id=inventaire_id)
        _log_inventaire_action(
            request, 
            "Édition inventaire admin", 
            f"Inventaire {inventaire.poste.nom} du {inventaire.date}"
        )
        messages.info(request, _(
            f"Redirection vers l'édition de l'inventaire {inventaire.poste.nom} du {inventaire.date}."
        ))
        
        return redirect(f'/admin/inventaire/inventairejournalier/{inventaire_id}/change/')
        
    except InventaireJournalier.DoesNotExist:
        messages.error(request, _("Inventaire non trouvé."))
        return redirect('/admin/inventaire/inventairejournalier/')


@login_required
def redirect_to_edit_recette_admin(request, recette_id):
    """Redirection vers l'édition d'une recette spécifique dans l'admin Django"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé à l'édition des recettes."))
        return redirect('inventaire:inventaire_list')
    
    try:
        recette = RecetteJournaliere.objects.get(id=recette_id)
        _log_inventaire_action(
            request, 
            "Édition recette admin", 
            f"Recette {recette.poste.nom} du {recette.date}"
        )
        messages.info(request, _(
            f"Redirection vers l'édition de la recette {recette.poste.nom} du {recette.date}."
        ))
        
        return redirect(f'/admin/inventaire/recettejournaliere/{recette_id}/change/')
        
    except RecetteJournaliere.DoesNotExist:
        messages.error(request, _("Recette non trouvée."))
        return redirect('/admin/inventaire/recettejournaliere/')


@login_required
def redirect_to_add_inventaire_admin(request):
    """Redirection vers l'ajout d'inventaire dans l'admin Django"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé à la création d'inventaires."))
        return redirect('inventaire:inventaire_list')
    
    _log_inventaire_action(request, "Ajout inventaire admin")
    messages.info(request, _("Redirection vers l'ajout d'inventaire."))
    
    return redirect('/admin/inventaire/inventairejournalier/add/')


@login_required
def redirect_to_add_recette_admin(request):
    """Redirection vers l'ajout de recette dans l'admin Django"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé à la création de recettes."))
        return redirect('inventaire:inventaire_list')
    
    _log_inventaire_action(request, "Ajout recette admin")
    messages.info(request, _("Redirection vers l'ajout de recette."))
    
    return redirect('/admin/inventaire/recettejournaliere/add/')


@login_required
def redirect_to_add_config_jour_admin(request):
    """Redirection vers l'ajout de configuration de jour dans l'admin Django"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé à la configuration des jours."))
        return redirect('inventaire:config_jour_list')
    
    _log_inventaire_action(request, "Ajout configuration jour admin")
    messages.info(request, _("Redirection vers l'ajout de configuration de jour."))
    
    return redirect('/admin/inventaire/configurationjour/add/')


# ===================================================================
# API POUR INTÉGRATION AVEC ADMIN DJANGO ET DASHBOARD
# ===================================================================

@login_required
@require_http_methods(["GET"])
def inventaire_stats_api(request):
    """API pour les statistiques des inventaires"""
    
    if not _check_admin_permission(request.user):
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    try:
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        stats = {
            'inventaires_today': InventaireJournalier.objects.filter(date=today).count(),
            'inventaires_week': InventaireJournalier.objects.filter(date__gte=week_ago).count(),
            'inventaires_month': InventaireJournalier.objects.filter(date__gte=month_ago).count(),
            'total_inventaires': InventaireJournalier.objects.count(),
        }
        
        return JsonResponse(stats)
        
    except Exception as e:
        logger.error(f"Erreur API inventaire stats: {str(e)}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)



@login_required
@require_http_methods(["GET"])
def recette_stats_api(request):
    """API pour statistiques des recettes"""
    user = request.user
    
    if not _check_admin_permission(user):
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    try:
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Sum, Avg, Count
        
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # Statistiques de base
        stats = {
            'recettes_today': RecetteJournaliere.objects.filter(date=today).count(),
            'recettes_week': RecetteJournaliere.objects.filter(date__gte=week_ago).count(),
            'recettes_month': RecetteJournaliere.objects.filter(date__gte=month_ago).count(),
            'total_recettes': RecetteJournaliere.objects.count(),
        }
        
        # Montants et taux
        week_data = RecetteJournaliere.objects.filter(date__gte=week_ago).aggregate(
            montant_total=Sum('montant_declare'),
            montant_potentiel=Sum('recette_potentielle'),
            taux_moyen=Avg('taux_deperdition'),
            nombre_jours_impertinents=Count('id', filter=Q(taux_deperdition__gte=0))
        )
        
        stats.update({
            'montant_total_week': float(week_data['montant_total'] or 0),
            'montant_potentiel_week': float(week_data['montant_potentiel'] or 0),
            'taux_moyen_week': float(week_data['taux_moyen'] or 0),
            'jours_impertinents_week': week_data['nombre_jours_impertinents'],
        })
        
        # Calcul écart
        if stats['montant_potentiel_week'] > 0:
            stats['ecart_week'] = stats['montant_total_week'] - stats['montant_potentiel_week']
            stats['pourcentage_ecart_week'] = (stats['ecart_week'] / stats['montant_potentiel_week']) * 100
        else:
            stats['ecart_week'] = 0
            stats['pourcentage_ecart_week'] = 0
        
        # Distribution par couleur d'alerte
        alerts_distribution = {
            'success': RecetteJournaliere.objects.filter(
                date__gte=week_ago, taux_deperdition__gt=-10
            ).count(),
            'warning': RecetteJournaliere.objects.filter(
                date__gte=week_ago, taux_deperdition__gte=-30, taux_deperdition__lte=-10
            ).count(),
            'danger': RecetteJournaliere.objects.filter(
                date__gte=week_ago, taux_deperdition__lt=-30
            ).count(),
        }
        
        stats['alerts_distribution'] = alerts_distribution
        
        return JsonResponse(stats)
        
    except Exception as e:
        logger.error(f"Erreur API recette stats: {str(e)}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


@login_required
@require_http_methods(["GET"])
def check_day_status_api(request):
    """API pour vérifier le statut d'un jour"""
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'Date requise'}, status=400)
    
    try:
        from datetime import datetime
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        try:
            config = ConfigurationJour.objects.get(date=target_date)
            status = {
                'date': target_date.isoformat(),
                'statut': config.statut,
                'statut_display': config.get_statut_display(),
                'configured': True,
                'cree_par': config.cree_par.nom_complet if config.cree_par else None,
                'commentaire': config.commentaire,
                'date_creation': config.date_creation.isoformat() if config.date_creation else None,
            }
        except ConfigurationJour.DoesNotExist:
            status = {
                'date': target_date.isoformat(),
                'statut': 'ferme',  # Par défaut fermé si pas configuré
                #'statut_display': 'Fermé - saisie verrouillée',
                'configured': False,
                'cree_par': None,
                'commentaire': 'Jour non configuré - fermé par défaut',
                'date_creation': None,
            }
        
        # Ajouter des informations sur les inventaires/recettes du jour
        inventaires_count = InventaireJournalier.objects.filter(date=target_date).count()
        recettes_count = RecetteJournaliere.objects.filter(date=target_date).count()
        
        # Détails par poste si demandé
        include_details = request.GET.get('include_details', '').lower() == 'true'
        details = {}
        
        if include_details:
            inventaires = InventaireJournalier.objects.filter(
                date=target_date
            ).select_related('poste', 'agent_saisie')
            
            details['inventaires'] = [
                {
                    'id': inv.id,
                    'poste_nom': inv.poste.nom,
                    'poste_code': inv.poste.code,
                    'total_vehicules': inv.total_vehicules,
                 #   'verrouille': inv.verrouille,
                   # 'valide': inv.valide,
                    'agent_saisie': inv.agent_saisie.nom_complet if inv.agent_saisie else None,
                }
                for inv in inventaires
            ]
            
            recettes = RecetteJournaliere.objects.filter(
                date=target_date
            ).select_related('poste', 'chef_poste')
            
            details['recettes'] = [
                {
                    'id': rec.id,
                    'poste_nom': rec.poste.nom,
                    'poste_code': rec.poste.code,
                    'montant_declare': float(rec.montant_declare),
                    'taux_deperdition': float(rec.taux_deperdition) if rec.taux_deperdition else None,
                    'couleur_alerte': rec.get_couleur_alerte(),
                    'chef_poste': rec.chef_poste.nom_complet if rec.chef_poste else None,
                }
                for rec in recettes
            ]
        
        status.update({
            'inventaires_count': inventaires_count,
            'recettes_count': recettes_count,
            'has_data': inventaires_count > 0 or recettes_count > 0,
            'details': details if include_details else None,
        })
        
        return JsonResponse(status)
        
    except ValueError:
        return JsonResponse({'error': 'Format de date invalide'}, status=400)
    except Exception as e:
        logger.error(f"Erreur API check day status: {str(e)}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


# ===================================================================
# API ACTIONS RAPIDES POUR LE DASHBOARD
# ===================================================================

@login_required
def quick_action_api(request):
    """API pour les actions rapides"""
    
    if not _check_admin_permission(request.user):
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    try:
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        action = data.get('action')
        
        if action == 'mark_impertinent':
            date_str = data.get('date')
            commentaire = data.get('commentaire', '')
            
            if not date_str:
                return JsonResponse({'error': 'Date requise'}, status=400)
            
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            config, created = ConfigurationJour.objects.get_or_create(
                date=target_date,
                defaults={
                    'statut': StatutJour.IMPERTINENT,
                    'cree_par': request.user,
                    'commentaire': commentaire or f'Marqué impertinent par {request.user.nom_complet}'
                }
            )
            
            if not created:
                config.statut = StatutJour.IMPERTINENT
                config.commentaire = commentaire
                config.save()
            
            log_user_action(
                request.user,
                "Marquage jour impertinent",
                f"Date: {target_date}, Commentaire: {commentaire}",
                request
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Jour {target_date.strftime("%d/%m/%Y")} marqué comme impertinent'
            })
        
        else:
            return JsonResponse({'error': 'Action non reconnue'}, status=400)
            
    except Exception as e:
        logger.error(f"Erreur API quick action: {str(e)}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


# ===================================================================
# VUES POUR RAPPORTS ET EXPORT
# ===================================================================

class RapportInventaireView(AdminRequiredMixin, View):
    """Vue pour générer des rapports d'inventaire"""
    
    def get(self, request):
        """Afficher la page de génération de rapports"""
        context = {
            'title': 'Génération de Rapports',
            'postes': Poste.objects.filter(is_active=True, type='peage').order_by('nom'),
            'current_year': timezone.now().year,
            'years': range(timezone.now().year - 5, timezone.now().year + 1),
        }
        return render(request, 'inventaire/rapport_generation.html', context)
    
    def post(self, request):
        """Générer et télécharger un rapport"""
        try:
            # Récupération des paramètres
            type_rapport = request.POST.get('type_rapport')
            format_export = request.POST.get('format_export', 'excel')
            date_debut = request.POST.get('date_debut')
            date_fin = request.POST.get('date_fin')
            poste_ids = request.POST.getlist('postes')
            
            # Validation des dates
            if not date_debut or not date_fin:
                messages.error(request, _("Les dates de début et fin sont obligatoires."))
                return redirect('inventaire:rapport_generation')
            
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
            
            if date_debut > date_fin:
                messages.error(request, _("La date de début ne peut pas être postérieure à la date de fin."))
                return redirect('inventaire:rapport_generation')
            
            # Construction de la requête base
            queryset = InventaireJournalier.objects.filter(
                date__gte=date_debut,
                date__lte=date_fin
            ).select_related('poste', 'agent_saisie')
            
            if poste_ids:
                queryset = queryset.filter(poste_id__in=poste_ids)
            
            # Génération selon le type de rapport
            if type_rapport == 'inventaires_detailles':
                return self._generer_rapport_inventaires_detailles(queryset, format_export)
            elif type_rapport == 'synthese_postes':
                return self._generer_rapport_synthese_postes(queryset, format_export)
            elif type_rapport == 'evolution_trafic':
                return self._generer_rapport_evolution_trafic(queryset, format_export)
            else:
                messages.error(request, _("Type de rapport non reconnu."))
                return redirect('inventaire:rapport_generation')
        
        except Exception as e:
            logger.error(f"Erreur génération rapport: {str(e)}")
            messages.error(request, _("Erreur lors de la génération du rapport."))
            return redirect('inventaire:rapport_generation')
    
    # def _generer_rapport_inventaires_detailles(self, queryset, format_export):
    #     """Générer un rapport détaillé des inventaires"""
    #     import io
    #     from django.http import HttpResponse
        
    #     if format_export == 'excel':
    #         try:
    #             import openpyxl
    #             from openpyxl.styles import Font, Alignment, PatternFill
                
    #             # Créer le workbook
    #             wb = openpyxl.Workbook()
    #             ws = wb.active
    #             ws.title = "Inventaires Détaillés"
                
    #             # En-têtes
    #             headers = [
    #                 'Date', 'Poste', 'Code Poste', 'Agent Saisie', 
    #                 'Total Véhicules', 'Périodes Saisies', 'Verrouillé', 
    #                 'Validé', 'Date Création'
    #             ]
                
    #             for col, header in enumerate(headers, 1):
    #                 cell = ws.cell(row=1, column=col, value=header)
    #                 cell.font = Font(bold=True)
    #                 cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
    #             # Données
    #             for row, inventaire in enumerate(queryset.order_by('date', 'poste__nom'), 2):
    #                 ws.cell(row=row, column=1, value=inventaire.date.strftime('%d/%m/%Y'))
    #                 ws.cell(row=row, column=2, value=inventaire.poste.nom)
    #                 ws.cell(row=row, column=3, value=inventaire.poste.code)
    #                 ws.cell(row=row, column=4, value=inventaire.agent_saisie.nom_complet if inventaire.agent_saisie else '')
    #                 ws.cell(row=row, column=5, value=inventaire.total_vehicules)
    #                 ws.cell(row=row, column=6, value=inventaire.nombre_periodes_saisies)
    #                 ws.cell(row=row, column=7, value='Oui' if inventaire.verrouille else 'Non')
    #                 ws.cell(row=row, column=8, value='Oui' if inventaire.valide else 'Non')
    #                 ws.cell(row=row, column=9, value=inventaire.date_creation.strftime('%d/%m/%Y %H:%M'))
                
    #             # Ajuster la largeur des colonnes
    #             for column in ws.columns:
    #                 max_length = 0
    #                 column_letter = column[0].column_letter
    #                 for cell in column:
    #                     try:
    #                         if len(str(cell.value)) > max_length:
    #                             max_length = len(str(cell.value))
    #                     except:
    #                         pass
    #                 adjusted_width = min(max_length + 2, 50)
    #                 ws.column_dimensions[column_letter].width = adjusted_width
                
    #             # Générer la réponse
    #             output = io.BytesIO()
    #             wb.save(output)
    #             output.seek(0)
                
    #             response = HttpResponse(
    #                 output.getvalue(),
    #                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #             )
    #             response['Content-Disposition'] = 'attachment; filename="rapport_inventaires_detailles.xlsx"'
    #             return response
                
    #         except ImportError:
    #             # Fallback vers CSV si openpyxl n'est pas disponible
    #             format_export = 'csv'
        
    #     if format_export == 'csv':
    #         import csv
            
    #         response = HttpResponse(content_type='text/csv')
    #         response['Content-Disposition'] = 'attachment; filename="rapport_inventaires_detailles.csv"'
            
    #         writer = csv.writer(response)
    #         writer.writerow([
    #             'Date', 'Poste', 'Code Poste', 'Agent Saisie', 
    #             'Total Véhicules', 'Périodes Saisies', 'Verrouillé', 
    #             'Validé', 'Date Création'
    #         ])
            
    #         for inventaire in queryset.order_by('date', 'poste__nom'):
    #             writer.writerow([
    #                 inventaire.date.strftime('%d/%m/%Y'),
    #                 inventaire.poste.nom,
    #                 inventaire.poste.code,
    #                 inventaire.agent_saisie.nom_complet if inventaire.agent_saisie else '',
    #                 inventaire.total_vehicules,
    #                 inventaire.nombre_periodes_saisies,
    #                 'Oui' if inventaire.verrouille else 'Non',
    #                 'Oui' if inventaire.valide else 'Non',
    #                 inventaire.date_creation.strftime('%d/%m/%Y %H:%M')
    #             ])
            
    #         return response
    
    # def _generer_rapport_synthese_postes(self, queryset, format_export):
    #     """Générer un rapport de synthèse par poste"""
    #     from django.db.models import Sum, Avg, Count
        
    #     # Agrégation des données par poste
    #     synthese = queryset.values('poste__nom', 'poste__code').annotate(
    #         total_vehicules=Sum('total_vehicules'),
    #         nombre_inventaires=Count('id'),
    #         moyenne_vehicules=Avg('total_vehicules'),
    #         inventaires_verrouilles=Count('id', filter=Q(verrouille=True)),
    #         inventaires_valides=Count('id', filter=Q(valide=True))
    #     ).order_by('poste__nom')
        
    #     if format_export == 'excel':
    #         try:
    #             import openpyxl
    #             from openpyxl.styles import Font, Alignment, PatternFill
    #             import io
                
    #             wb = openpyxl.Workbook()
    #             ws = wb.active
    #             ws.title = "Synthèse par Poste"
                
    #             # En-têtes
    #             headers = [
    #                 'Poste', 'Code', 'Total Véhicules', 'Nb Inventaires',
    #                 'Moyenne Véhicules/Jour', 'Inventaires Verrouillés',
    #                 'Inventaires Validés', 'Taux Validation (%)'
    #             ]
                
    #             for col, header in enumerate(headers, 1):
    #                 cell = ws.cell(row=1, column=col, value=header)
    #                 cell.font = Font(bold=True)
    #                 cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
    #             # Données
    #             for row, data in enumerate(synthese, 2):
    #                 taux_validation = (data['inventaires_valides'] / data['nombre_inventaires'] * 100) if data['nombre_inventaires'] > 0 else 0
                    
    #                 ws.cell(row=row, column=1, value=data['poste__nom'])
    #                 ws.cell(row=row, column=2, value=data['poste__code'])
    #                 ws.cell(row=row, column=3, value=data['total_vehicules'] or 0)
    #                 ws.cell(row=row, column=4, value=data['nombre_inventaires'])
    #                 ws.cell(row=row, column=5, value=round(data['moyenne_vehicules'] or 0, 2))
    #                 ws.cell(row=row, column=6, value=data['inventaires_verrouilles'])
    #                 ws.cell(row=row, column=7, value=data['inventaires_valides'])
    #                 ws.cell(row=row, column=8, value=round(taux_validation, 2))
                
    #             # Ajuster les colonnes
    #             for column in ws.columns:
    #                 max_length = 0
    #                 column_letter = column[0].column_letter
    #                 for cell in column:
    #                     try:
    #                         if len(str(cell.value)) > max_length:
    #                             max_length = len(str(cell.value))
    #                     except:
    #                         pass
    #                 adjusted_width = min(max_length + 2, 50)
    #                 ws.column_dimensions[column_letter].width = adjusted_width
                
    #             output = io.BytesIO()
    #             wb.save(output)
    #             output.seek(0)
                
    #             response = HttpResponse(
    #                 output.getvalue(),
    #                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #             )
    #             response['Content-Disposition'] = 'attachment; filename="rapport_synthese_postes.xlsx"'
    #             return response
                
    #         except ImportError:
    #             format_export = 'csv'
        
    #     if format_export == 'csv':
    #         import csv
            
    #         response = HttpResponse(content_type='text/csv')
    #         response['Content-Disposition'] = 'attachment; filename="rapport_synthese_postes.csv"'
            
    #     writer = csv.writer(response)
    #         writer.writerow([
    #             'Poste', 'Code', 'Total Véhicules', 'Nb Inventaires',
    #             'Moyenne Véhicules/Jour', 'Inventaires Verrouillés',
    #             'Inventaires Validés', 'Taux Validation (%)'
    #         ])
            
    #         for data in synthese:
    #             taux_validation = (data['inventaires_valides'] / data['nombre_inventaires'] * 100) if data['nombre_inventaires'] > 0 else 0
    #             writer.writerow([
    #                 data['poste__nom'],
    #                 data['poste__code'],
    #                 data['total_vehicules'] or 0,
    #                 data['nombre_inventaires'],
    #                 round(data['moyenne_vehicules'] or 0, 2),
    #                 data['inventaires_verrouilles'],
    #                 data['inventaires_valides'],
    #                 round(taux_validation, 2)
    #             ])
            
    #         return response
    
    # def _generer_rapport_evolution_trafic(self, queryset, format_export):
    #     """Générer un rapport d'évolution du trafic"""
    #     from django.db.models import Sum
    #     from collections import defaultdict
        
    #     # Agrégation par date
    #     evolution = queryset.values('date').annotate(
    #         total_vehicules=Sum('total_vehicules'),
    #         nombre_postes=Count('poste', distinct=True)
    #     ).order_by('date')
        
    #     if format_export == 'excel':
    #         try:
    #             import openpyxl
    #             from openpyxl.styles import Font, PatternFill
    #             from openpyxl.chart import LineChart, Reference
    #             import io
                
    #             wb = openpyxl.Workbook()
    #             ws = wb.active
    #             ws.title = "Evolution du Trafic"
                
    #             # En-têtes
    #             headers = ['Date', 'Total Véhicules', 'Nombre Postes', 'Moyenne par Poste']
                
    #             for col, header in enumerate(headers, 1):
    #                 cell = ws.cell(row=1, column=col, value=header)
    #                 cell.font = Font(bold=True)
    #                 cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
    #             # Données
    #             for row, data in enumerate(evolution, 2):
    #                 moyenne_poste = (data['total_vehicules'] / data['nombre_postes']) if data['nombre_postes'] > 0 else 0
                    
    #                 ws.cell(row=row, column=1, value=data['date'].strftime('%d/%m/%Y'))
    #                 ws.cell(row=row, column=2, value=data['total_vehicules'] or 0)
    #                 ws.cell(row=row, column=3, value=data['nombre_postes'])
    #                 ws.cell(row=row, column=4, value=round(moyenne_poste, 2))
                
    #             # Ajouter un graphique si possible
    #             if len(evolution) > 1:
    #                 chart = LineChart()
    #                 chart.title = "Evolution du trafic"
    #                 chart.y_axis.title = 'Nombre de véhicules'
    #                 chart.x_axis.title = 'Date'
                    
    #                 data_ref = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(evolution) + 1)
    #                 chart.add_data(data_ref, titles_from_data=True)
                    
    #                 ws.add_chart(chart, "F5")
                
    #             # Ajuster les colonnes
    #             for column in ws.columns:
    #                 max_length = 0
    #                 column_letter = column[0].column_letter
    #                 for cell in column:
    #                     try:
    #                         if len(str(cell.value)) > max_length:
    #                             max_length = len(str(cell.value))
    #                     except:
    #                         pass
    #                 adjusted_width = min(max_length + 2, 50)
    #                 ws.column_dimensions[column_letter].width = adjusted_width
                
    #             output = io.BytesIO()
    #             wb.save(output)
    #             output.seek(0)
                
    #             response = HttpResponse(
    #                 output.getvalue(),
    #                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #             )
    #             response['Content-Disposition'] = 'attachment; filename="rapport_evolution_trafic.xlsx"'
    #             return response
                
    #         except ImportError:
    #             format_export = 'csv'
        
    #     if format_export == 'csv':
    #         import csv
            
    #         response = HttpResponse(content_type='text/csv')
    #         response['Content-Disposition'] = 'attachment; filename="rapport_evolution_trafic.csv"'
            
    #         writer = csv.writer(response)
    #         writer.writerow(['Date', 'Total Véhicules', 'Nombre Postes', 'Moyenne par Poste'])
            
    #         for data in evolution:
    #             moyenne_poste = (data['total_vehicules'] / data['nombre_postes']) if data['nombre_postes'] > 0 else 0
    #             writer.writerow([
    #                 data['date'].strftime('%d/%m/%Y'),
    #                 data['total_vehicules'] or 0,
    #                 data['nombre_postes'],
    #                 round(moyenne_poste, 2)
    #             ])
            
    #         return response


# ===================================================================
# VUES D'AIDE ET DOCUMENTATION INVENTAIRE
# ===================================================================

@login_required
def inventaire_admin_help_view(request):
    """Vue d'aide pour l'utilisation de l'admin Django - module inventaire"""
    user = request.user
    
    if not _check_admin_permission(user):
        messages.error(request, _("Accès non autorisé."))
        return redirect('inventaire:inventaire_list')
    
    context = {
        'title': 'Aide - Administration Inventaires',
        'user_permissions': {
            'can_manage_inventaires': True,
            'can_manage_recettes': True,
            'can_config_jours': True,
            'can_view_statistics': True,
        },
        'help_sections': [
            {
                'title': 'Gestion des Inventaires',
                'description': 'Créer, modifier et valider les inventaires journaliers',
                'url': '/admin/inventaire/inventairejournalier/',
                'features': [
                    'Saisie par période horaire (8h-18h)',
                    'Calculs automatiques des totaux',
                    'Verrouillage et validation par responsable',
                    'Export des données en CSV/Excel',
                    'Historique complet des modifications'
                ]
            },
            {
                'title': 'Gestion des Recettes',
                'description': 'Suivi des recettes déclarées et calcul des taux de déperdition',
                'url': '/admin/inventaire/recettejournaliere/',
                'features': [
                    'Saisie des montants déclarés par les chefs de poste',
                    'Calcul automatique des taux de déperdition',
                    'Alertes visuelles par code couleur',
                    'Détection automatique des journées impertinentes',
                    'Rapports consolidés par période'
                ]
            },
            {
                'title': 'Configuration des Jours',
                'description': 'Gérer les jours ouverts/fermés pour la saisie',
                'url': '/admin/inventaire/configurationjour/',
                'features': [
                    'Ouverture/fermeture des jours de saisie',
                    'Marquage des jours impertinents',
                    'Commentaires et historique des modifications',
                    'Contrôle granulaire des accès',
                    'Planification avancée des périodes'
                ]
            },
            {
                'title': 'Statistiques et Rapports',
                'description': 'Consultation des statistiques consolidées',
                'url': '/admin/inventaire/statistiquesperiodiques/',
                'features': [
                    'Statistiques hebdomadaires, mensuelles, trimestrielles',
                    'Graphiques de tendances et évolutions',
                    'Comparaisons inter-postes',
                    'Export de rapports détaillés',
                    'Tableaux de bord personnalisés'
                ]
            }
        ],
        'quick_actions': [
            {
                'title': 'Ouvrir Jour Actuel',
                'description': 'Ouvrir le jour actuel pour la saisie',
                'action': 'open_today',
                'icon': 'fas fa-calendar-plus',
                'class': 'btn-success'
            },
            {
                'title': 'Consulter Inventaires',
                'description': 'Voir tous les inventaires récents',
                'url': '/admin/inventaire/inventairejournalier/',
                'icon': 'fas fa-list',
                'class': 'btn-primary'
            },
            {
                'title': 'Gérer Recettes',
                'description': 'Consulter et modifier les recettes',
                'url': '/admin/inventaire/recettejournaliere/',
                'icon': 'fas fa-euro-sign',
                'class': 'btn-warning'
            }
        ]
    }
    
    return render(request, 'inventaire/admin_help.html', context)


# ===================================================================
# DASHBOARD ET WIDGETS INVENTAIRE
# ===================================================================

@login_required
def inventaire_dashboard_widget(request):
    """Widget dashboard pour les inventaires"""
    if not request.user.peut_gerer_inventaire:
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    try:
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        
        # Statistiques rapides
        stats = {
            'inventaires_today': InventaireJournalier.objects.filter(date=today).count(),
            'inventaires_week': InventaireJournalier.objects.filter(date__gte=week_ago).count(),
            # 'inventaires_en_attente': InventaireJournalier.objects.filter(
            #     verrouille=False, date__lte=today
            # ).count(),
            'postes_actifs_today': InventaireJournalier.objects.filter(
                date=today
            ).values('poste').distinct().count(),
        }
        
        # Derniers inventaires pour l'utilisateur
        if request.user.acces_tous_postes or _check_admin_permission(request.user):
            derniers_inventaires = InventaireJournalier.objects.select_related(
                'poste', 'agent_saisie'
            ).order_by('-date_creation')[:5]
        else:
            derniers_inventaires = InventaireJournalier.objects.filter(
                poste=request.user.poste_affectation
            ).select_related('poste', 'agent_saisie').order_by('-date_creation')[:5]
        
        inventaires_data = []
        for inv in derniers_inventaires:
            inventaires_data.append({
                'id': inv.id,
                'date': inv.date.strftime('%d/%m/%Y'),
                'poste': inv.poste.nom,
                'total_vehicules': inv.total_vehicules,
              #  'verrouille': inv.verrouille,
               # 'valide': inv.valide,
                'agent': inv.agent_saisie.nom_complet if inv.agent_saisie else 'Non défini',
                'url': f'/inventaire/{inv.id}/'
            })
        
        # Actions rapides disponibles
        actions_rapides = []
        
        if _check_admin_permission(request.user):
            actions_rapides.extend([
                {
                    'title': 'Ouvrir Aujourd\'hui',
                    'action': 'open_today',
                    'icon': 'fas fa-calendar-plus',
                    'class': 'btn-success btn-sm'
                },
                {
                    'title': 'Admin Inventaires',
                    'url': '/admin/inventaire/inventairejournalier/',
                    'icon': 'fas fa-cogs',
                    'class': 'btn-primary btn-sm'
                }
            ])
        
        if request.user.poste_affectation:
            actions_rapides.append({
                'title': 'Nouvelle Saisie',
                'url': f'/inventaire/saisie/{request.user.poste_affectation.id}/',
                'icon': 'fas fa-plus',
                'class': 'btn-info btn-sm'
            })
        
        response_data = {
            'stats': stats,
            'derniers_inventaires': inventaires_data,
            'actions_rapides': actions_rapides,
            'can_admin': _check_admin_permission(request.user),
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Erreur widget dashboard inventaire: {str(e)}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


# ===================================================================
# GESTION DES ERREURS DE REDIRECTION INVENTAIRE
# ===================================================================

@login_required
def inventaire_redirect_error_handler(request, error_type='permission'):
    """Gestionnaire d'erreurs pour les redirections inventaire"""
    user = request.user
    
    if error_type == 'permission':
        messages.error(request, _(
            "Vous n'avez pas les permissions nécessaires pour accéder à cette section inventaire. "
            "Contactez un administrateur si vous pensez que c'est une erreur."
        ))
        
        # Journaliser la tentative d'accès non autorisée
        try:
            JournalAudit.objects.create(
                utilisateur=user,
                action="ACCÈS REFUSÉ - Redirection admin inventaire",
                details=f"Tentative d'accès non autorisée depuis {request.META.get('HTTP_REFERER', 'URL inconnue')}",
                adresse_ip=request.META.get('REMOTE_ADDR'),
                url_acces=request.path,
                methode_http=request.method,
                succes=False
            )
        except Exception as e:
            logger.error(f"Erreur journalisation tentative accès inventaire: {str(e)}")
    
    elif error_type == 'not_found':
        messages.error(request, _("L'inventaire ou la recette demandée n'a pas été trouvée."))
    
    elif error_type == 'locked':
        messages.warning(request, _("Cette ressource est verrouillée et ne peut plus être modifiée."))
    
    elif error_type == 'closed_day':
        messages.warning(request, _("Ce jour est fermé pour la saisie."))
    
    else:
        messages.error(request, _("Une erreur est survenue lors de la redirection."))
    
    # Rediriger vers la liste des inventaires
    return redirect('inventaire:inventaire_list')


# ===================================================================
# VUES POUR LA SAUVEGARDE ET RESTAURATION
# ===================================================================

@login_required
@require_http_methods(["POST"])
def backup_inventaires_api(request):
    """API pour sauvegarder les données d'inventaire"""
    user = request.user
    
    if not _check_admin_permission(user):
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    try:
        import json
        from django.core import serializers
        from datetime import datetime
        
        # Récupération des paramètres
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        date_debut = data.get('date_debut')
        date_fin = data.get('date_fin')
        
        if not date_debut or not date_fin:
            return JsonResponse({'error': 'Dates requises'}, status=400)
        
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        
        # Collecte des données
        inventaires = InventaireJournalier.objects.filter(
            date__gte=date_debut, date__lte=date_fin
        ).prefetch_related('details_periodes')
        
        recettes = RecetteJournaliere.objects.filter(
            date__gte=date_debut, date__lte=date_fin
        )
        
        configs = ConfigurationJour.objects.filter(
            date__gte=date_debut, date__lte=date_fin
        )
        
        # Sérialisation
        backup_data = {
            'metadata': {
                'date_creation': timezone.now().isoformat(),
                'cree_par': user.username,
                'date_debut': date_debut.isoformat(),
                'date_fin': date_fin.isoformat(),
                'version': '1.0'
            },
            'inventaires': json.loads(serializers.serialize('json', inventaires)),
            'details_periodes': [],
            'recettes': json.loads(serializers.serialize('json', recettes)),
            'configurations': json.loads(serializers.serialize('json', configs))
        }
        
        # Ajouter les détails des périodes
        for inventaire in inventaires:
            details = DetailInventairePeriode.objects.filter(inventaire=inventaire)
            backup_data['details_periodes'].extend(
                json.loads(serializers.serialize('json', details))
            )
        
        # Journaliser l'action
        _log_inventaire_action(
            request,
            "Sauvegarde données inventaire",
            f"Période: {date_debut} à {date_fin}, {len(inventaires)} inventaires"
        )
        
        # Retourner les données
        response = HttpResponse(
            json.dumps(backup_data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="backup_inventaires_{date_debut}_{date_fin}.json"'
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur sauvegarde inventaires: {str(e)}")
        return JsonResponse({'error': 'Erreur lors de la sauvegarde'}, status=500)


# ===================================================================
# VUES DE MAINTENANCE ET DIAGNOSTIC
# ===================================================================

# @login_required
# def diagnostic_inventaires_view(request):
#     """Vue de diagnostic pour les inventaires"""
#     user = request.user
    
#     if not _check_admin_permission(user):
#         messages.error(request, _("Accès non autorisé au diagnostic."))
#         return redirect('inventaire:inventaire_list')
    
#     try:
#         # Statistiques de santé
#         today = timezone.now().date()
#         week_ago = today - timedelta(days=7)
#         month_ago = today - timedelta(days=30)
        
#         # Détection des anomalies
#         anomalies = []
        
#         # Inventaires sans détails
#         inventaires_vides = InventaireJournalier.objects.filter(
#             details_periodes__isnull=True
#         ).count()
#         if inventaires_vides > 0:
#             anomalies.append({
#                 'type': 'warning',
#                 'message': f'{inventaires_vides} inventaires sans détails de périodes',
#                 'action': 'Vérifier les inventaires incomplets'
#             })
        
#         # Inventaires non verrouillés anciens
#         inventaires_anciens = InventaireJournalier.objects.filter(
#             date__lt=week_ago, verrouille=False
#         ).count()
#         if inventaires_anciens > 0:
#             anomalies.append({
#                 'type': 'info',
#                 'message': f'{inventaires_anciens} inventaires de plus de 7 jours non verrouillés',
#                 'action': 'Vérifier les inventaires en attente'
#             })
        
#         # Recettes sans inventaire associé
#         recettes_orphelines = RecetteJournaliere.objects.exclude(
#             poste__in=InventaireJournalier.objects.filter(
#                 date=models.OuterRef('date')
#             ).values('poste')
#         ).count()
#         if recettes_orphelines > 0:
#             anomalies.append({
#                 'type': 'error',
#                 'message': f'{recettes_orphelines} recettes sans inventaire associé',
#                 'action': 'Créer les inventaires manquants'
#             })
        
#         # Jours sans configuration
#         jours_non_configures = 0
#         for i in range(7):
#             day = today - timedelta(days=i)
#             if not ConfigurationJour.objects.filter(date=day).exists():
#                 jours_non_configures += 1
        
#         if jours_non_configures > 0:
#             anomalies.append({
#                 'type': 'warning',
#                 'message': f'{jours_non_configures} jours récents sans configuration',
#                 'action': 'Configurer les jours manquants'
#             })
        
#         # Statistiques de performance
#         stats_performance = {
#             'inventaires_total': InventaireJournalier.objects.count(),
#             'inventaires_month': InventaireJournalier.objects.filter(date__gte=month_ago).count(),
#             'taux_verrouillage': 0,
#             'taux_validation': 0,
#             'postes_actifs': Poste.objects.filter(is_active=True).count(),
#             'derniere_saisie': None
#         }
        
#         # Calculs des taux
#         inventaires_month = InventaireJournalier.objects.filter(date__gte=month_ago)
#         if inventaires_month.exists():
#             total_month = inventaires_month.count()
#             verrouilles_month = inventaires_month.filter(verrouille=True).count()
#             valides_month = inventaires_month.filter(valide=True).count()
            
#             stats_performance['taux_verrouillage'] = round((verrouilles_month / total_month) * 100, 2)
#             stats_performance['taux_validation'] = round((valides_month / total_month) * 100, 2)
        
#         # Dernière saisie
#         derniere_saisie = InventaireJournalier.objects.order_by('-date_creation').first()
#         if derniere_saisie:
#             stats_performance['derniere_saisie'] = {
#                 'date': derniere_saisie.date_creation.strftime('%d/%m/%Y %H:%M'),
#                 'poste': derniere_saisie.poste.nom,
#                 'agent': derniere_saisie.agent_saisie.nom_complet if derniere_saisie.agent_saisie else 'Non défini'
#             }
        
#         context = {
#             'title': 'Diagnostic Inventaires',
#             'anomalies': anomalies,
#             'stats_performance': stats_performance,
#             'diagnostic_date': timezone.now().strftime('%d/%m/%Y %H:%M'),
#         }
        
#         return render(request, 'inventaire/diagnostic.html', context)
        
    except Exception as e:
        logger.error(f"Erreur diagnostic inventaires: {str(e)}")
        messages.error(request, _("Erreur lors du diagnostic."))
        return redirect('inventaire:inventaire_list')

import calendar
from datetime import date

def is_admin(user):
    """Vérifier si l'utilisateur est admin"""
    return user.is_authenticated and user.is_superuser

# @login_required
# @user_passes_test(is_admin)
# def gerer_jours_inventaire(request, inventaire_id):
#     """Vue pour gérer les jours d'activation d'un inventaire mensuel"""
    
#     # Pour l'instant, utiliser l'inventaire journalier
#     # (Plus tard, remplacer par InventaireMensuel quand le modèle sera créé)
    
#     # Récupérer le mois et l'année depuis l'inventaire
#     # Pour le test, on va utiliser le mois et année courants
#     today = date.today()
#     mois = today.month
#     annee = today.year
    
#     # Obtenir le calendrier du mois
#     cal = calendar.monthcalendar(annee, mois)
    
#     # Obtenir les jours actuellement ouverts
#     jours_actifs = []
#     nb_jours = calendar.monthrange(annee, mois)[1]
    
#     for jour in range(1, nb_jours + 1):
#         date_jour = date(annee, mois, jour)
#         config = ConfigurationJour.objects.filter(date=date_jour).first()
#         if config and config.statut == 'ouvert':
#             jours_actifs.append(jour)
    
#     if request.method == 'POST':
#         # Traiter l'activation/désactivation des jours
#         jours_a_activer = request.POST.getlist('jours_actifs')
        
#         for jour in range(1, nb_jours + 1):
#             date_jour = date(annee, mois, jour)
            
#             if str(jour) in jours_a_activer:
#                 # Activer le jour
#                 config, created = ConfigurationJour.objects.get_or_create(
#                     date=date_jour,
#                     defaults={
#                         'statut': 'ouvert',
#                         'cree_par': request.user,
#                         'commentaire': f'Activé pour inventaire du {mois}/{annee}'
#                     }
#                 )
#                 if not created and config.statut != 'ouvert':
#                     config.statut = 'ouvert'
#                     config.save()
#             else:
#                 # Désactiver le jour
#                 config, created = ConfigurationJour.objects.get_or_create(
#                     date=date_jour,
#                     defaults={
#                         'statut': 'ferme',
#                         'cree_par': request.user,
#                         'commentaire': f'Fermé pour inventaire du {mois}/{annee}'
#                     }
#                 )
#                 if not created and config.statut != 'ferme':
#                     config.statut = 'ferme'
#                     config.save()
        
#         messages.success(request, f"Les jours du mois {mois}/{annee} ont été mis à jour.")
#         return redirect('admin:inventaire_inventairejournalier_changelist')
    
#     # Créer un objet fictif pour le template
#     inventaire = {
#         'id': inventaire_id,
#         'titre': f'Inventaire {calendar.month_name[mois]} {annee}',
#         'mois': mois,
#         'annee': annee,
#         'description': 'Gestion des jours d\'activation pour la saisie',
#         'get_nombre_postes': lambda: 'Tous les postes'
#     }
    
#     context = {
#         'inventaire': inventaire,
#         'calendrier': cal,
#         'jours_actifs': jours_actifs,
#         'title': f'Gérer les jours - {calendar.month_name[mois]} {annee}',
#     }
    
#     return render(request, 'admin/inventaire/gerer_jours.html', context)


@login_required
@require_permission('peut_gerer_inventaire')
def changer_agent_poste(request, poste_id):
    """
    Vue pour changer l'agent d'un poste
    """
    poste = get_object_or_404(Poste, id=poste_id)
    
    if not request.user.peut_acceder_poste(poste):
        messages.error(request, "Vous n'avez pas accès à ce poste.")
        return redirect('liste_postes_inventaires')
    
    if request.method == 'POST':
        nouvel_agent_id = request.POST.get('agent_id')
        commentaire = request.POST.get('commentaire', '')
        
        if nouvel_agent_id:
            try:
                nouvel_agent = UtilisateurSUPPER.objects.get(
                    id=nouvel_agent_id,
                    habilitation='agent_inventaire',
                    is_active=True
                )
                
                # Changer l'affectation
                ancien_agent = poste.agents_affectes.first()
                poste.agents_affectes.clear()
                nouvel_agent.poste_affectation = poste
                nouvel_agent.save()
                
                # Journaliser le changement
                log_user_action(
                    request.user,
                    "Changement agent poste",
                    f"Poste: {poste.nom} | Ancien: {ancien_agent.nom_complet if ancien_agent else 'Aucun'} "
                    f"| Nouveau: {nouvel_agent.nom_complet} | Commentaire: {commentaire}",
                    request
                )
                
                messages.success(request, f"Agent {nouvel_agent.nom_complet} affecté au poste {poste.nom}")
                return redirect('detail_poste_inventaires', poste_id=poste.id)
                
            except UtilisateurSUPPER.DoesNotExist:
                messages.error(request, "Agent sélectionné invalide.")
    
    # Agents inventaire disponibles
    agents_disponibles = UtilisateurSUPPER.objects.filter(
        habilitation='agent_inventaire',
        is_active=True
    ).exclude(
        poste_affectation=poste
    )
    
    context = {
        'poste': poste,
        'agents_disponibles': agents_disponibles,
        'title': f'Changer agent - {poste.nom}'
    }
    
    return render(request, 'inventaire/changer_agent_poste.html', context)



@login_required
def redirect_to_dashboard(request):
    """Redirection vers le dashboard approprié selon le rôle"""
    if request.user.is_authenticated:
        return redirect('admin:index')  # Rediriger vers l'admin Django
    else:
        return redirect('accounts:login')


@login_required
def programmer_inventaire(request):
    """
    Vue pour programmer des inventaires avec sélection automatique/manuelle
    
    MISE À JOUR - Permissions granulaires:
    - Permission requise: peut_programmer_inventaire
    - Log détaillé de chaque action utilisateur
    """
    import logging
    logger = logging.getLogger('supper')
    
    user = request.user
    
    # =========================================
    # VÉRIFICATION PERMISSION GRANULAIRE
    # =========================================
    if not has_permission(user, 'peut_programmer_inventaire'):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Programmation inventaire",
            f"Permission manquante: peut_programmer_inventaire | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | "
            f"IP: {request.META.get('REMOTE_ADDR')}",
            request
        )
        messages.error(request, _("Vous n'avez pas la permission de programmer des inventaires."))
        return redirect('common:dashboard')
    
    # Log de l'accès autorisé
    log_user_action(
        user,
        "Accès page programmation inventaire",
        f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | Méthode: {request.method}",
        request
    )
    
    context = {
        'mois_disponibles': [
            (date.today() + timedelta(days=30*i)).strftime('%Y-%m')
            for i in range(0, 6)
        ]
    }
    
    # Récupérer le mois sélectionné
    mois_str = request.GET.get('mois') or request.POST.get('mois')
    logger.debug(f"[DEBUG] Mois sélectionné: {mois_str}")
    
    if mois_str:
        context['mois_selectionne'] = mois_str
        try:
            mois = datetime.strptime(mois_str, '%Y-%m').date()
            context['mois'] = mois
        except Exception as e:
            logger.error(f"[ERROR] Format de mois invalide: {e}")
            messages.error(request, "Format de mois invalide")
            return render(request, 'inventaire/programmer_inventaire.html', context)
    
    # Si on clique sur "Générer" pour un motif
    if request.method == 'POST' and 'generer' in request.POST:
        motif = request.POST.get('motif')
        logger.info(f"[INFO] Génération pour motif: {motif}, mois: {mois_str}")
        context['motif_selectionne'] = motif
        
        # Log de la génération
        log_user_action(
            user,
            "Génération postes pour programmation",
            f"Motif: {motif} | Mois: {mois_str}",
            request
        )
        
        try:
            # MOTIF 1: RISQUE DE BAISSE ANNUEL
            if motif == 'risque_baisse':
                from inventaire.services.evolution_service import EvolutionService
                
                # Utiliser EXACTEMENT la même méthode que dans ProgrammationInventaire
                postes_data = EvolutionService.identifier_postes_en_baisse(
                    type_analyse='annuel',
                    seuil_baisse=-5
                )
                
                # Enrichir avec vérification des programmations existantes
                for item in postes_data:
                    item['deja_programme'] = ProgrammationInventaire.objects.filter(
                        poste=item['poste'],
                        mois=mois,
                        motif=motif,
                        actif=True
                    ).exists()
                    
                    # IMPORTANT : Utiliser les mêmes clés que calculer_risque_baisse_annuel
                    item['pourcentage_baisse'] = abs(item['taux_evolution'])
                    item['recettes_estimees'] = item.get('recettes_estimees', 0)
                    item['recettes_n1'] = item.get('recettes_precedentes', 0)
                
                context['postes_risque_baisse'] = postes_data
                logger.debug(f"[DEBUG] Postes risque baisse trouvés: {len(postes_data)}")
                
            # MOTIF 2: GRAND STOCK
            elif motif == 'grand_stock':
                postes_data = ProgrammationInventaire.get_postes_avec_grand_stock()
                # Enrichir avec des informations supplémentaires
                for item in postes_data:
                    # Vérifier si déjà programmé
                    item['deja_programme'] = ProgrammationInventaire.objects.filter(
                        poste=item['poste'],
                        mois=mois,
                        motif=motif,
                        actif=True
                    ).exists()
                    
                    # Formater la date pour l'affichage
                    item['date_epuisement_formatee'] = item['date_epuisement'].strftime('%d/%m/%Y')
                    
                    # Calculer le nombre de mois restants
                    mois_restants = (item['date_epuisement'].year - date.today().year) * 12 + \
                                (item['date_epuisement'].month - date.today().month)
                    item['mois_restants'] = max(0, mois_restants)
                
                context['postes_grand_stock'] = postes_data
                logger.debug(f"[DEBUG] Postes grand stock trouvés: {len(postes_data)}")
                
            # MOTIF 3: TAUX DE DÉPERDITION AUTOMATIQUE
            elif motif == 'taux_deperdition':
                # Récupérer tous les postes actifs
                tous_postes = Poste.objects.filter(is_active=True, type='peage').order_by('nom')
                logger.debug(f"[DEBUG] Total postes actifs: {tous_postes.count()}")
                
                # Séparer les postes selon leur taux de déperdition
                postes_auto_selectionnes = []
                postes_non_selectionnes = []
                
                for poste in tous_postes:
                    # Chercher le dernier taux de déperdition
                    derniere_recette = RecetteJournaliere.objects.filter(
                        poste=poste,
                        taux_deperdition__isnull=False
                    ).order_by('-date').first()
                    
                    if derniere_recette:
                        poste_data = {
                            'poste': poste,
                            'taux_deperdition': derniere_recette.taux_deperdition,
                            'date_calcul': derniere_recette.date,
                            'alerte': derniere_recette.get_couleur_alerte()
                        }
                        
                        # Sélection automatique si taux < -30%
                        if derniere_recette.taux_deperdition < -30:
                            poste_data['selection_auto'] = True
                            postes_auto_selectionnes.append(poste_data)
                
                context['postes_taux_auto'] = postes_auto_selectionnes
                context['postes_taux_manuel'] = postes_non_selectionnes
                
                logger.info(f"[INFO] Taux déperdition: {len(postes_auto_selectionnes)} auto-sélectionnés, "
                           f"{len(postes_non_selectionnes)} non sélectionnés")
            
            # MOTIF 4: PRÉSENCE ADMINISTRATIVE (NOUVEAU)
            elif motif == 'presence_admin':
                # Récupérer TOUS les postes, non cochés par défaut
                tous_postes = Poste.objects.filter(is_active=True, type='peage').order_by('nom')
                context['postes_presence_admin'] = tous_postes
                logger.debug(f"[DEBUG] Présence administrative: {tous_postes.count()} postes disponibles")
                
                # Aucun poste n'est pré-sélectionné pour la présence administrative
                context['aucune_preselection'] = True
            
            # Vérifier les programmations existantes pour éviter les doublons
            if mois_str and motif:
                prog_existantes = ProgrammationInventaire.objects.filter(
                    mois=mois,
                    motif=motif,
                    actif=True
                ).values_list('poste_id', flat=True)
                context['programmations_existantes'] = list(prog_existantes)
                logger.debug(f"[DEBUG] Programmations existantes: {len(prog_existantes)}")
                
        except Exception as e:
            logger.error(f"[ERROR] Erreur lors de la génération: {str(e)}")
            log_user_action(
                user,
                "ERREUR - Génération programmation",
                f"Motif: {motif} | Erreur: {str(e)}",
                request
            )
            messages.error(request, f"Erreur lors de la génération: {str(e)}")
    
    # Si on soumet le formulaire final de programmation
    if request.method == 'POST' and 'programmer' in request.POST:
        motif = request.POST.get('motif')
        postes_ids = request.POST.getlist('postes')
        logger.info(f"[INFO] Programmation finale - Motif: {motif}, Postes: {len(postes_ids)}")
        
        if not postes_ids:
            messages.error(request, "Veuillez sélectionner au moins un poste")
        else:
            postes_programmes = []
            postes_deja_programmes = []
            
            for poste_id in postes_ids:
                try:
                    poste = Poste.objects.get(id=poste_id)
                    
                    # Vérifier si déjà programmé
                    if ProgrammationInventaire.objects.filter(
                        poste=poste,
                        mois=mois,
                        motif=motif,
                        actif=True
                    ).exists():
                        postes_deja_programmes.append(poste.nom)
                        continue
                    
                    # Créer la programmation
                    prog = ProgrammationInventaire.objects.create(
                        poste=poste,
                        mois=mois,
                        motif=motif,
                        cree_par=request.user,
                        actif=True
                    )
                    
                    # Données spécifiques selon le motif
                    if motif == MotifInventaire.GRAND_STOCK:
                        stock_key = f'stock_{poste_id}'
                        if stock_key in request.POST:
                            try:
                                prog.stock_restant = int(request.POST[stock_key])
                            except ValueError:
                                pass
                    
                    elif motif == MotifInventaire.TAUX_DEPERDITION:
                        taux_key = f'taux_{poste_id}'
                        if taux_key in request.POST:
                            try:
                                prog.taux_deperdition_precedent = float(request.POST[taux_key])
                            except ValueError:
                                pass
                    
                    elif motif == MotifInventaire.RISQUE_BAISSE:
                        # Les données sont calculées automatiquement
                        prog.calculer_risque_baisse_annuel()
                    
                    prog.save()
                    postes_programmes.append(poste.nom)
                    logger.info(f"[INFO] Programmation créée pour: {poste.nom}")
                    
                except Poste.DoesNotExist:
                    logger.error(f"[ERROR] Poste {poste_id} introuvable")
                except Exception as e:
                    logger.error(f"[ERROR] Erreur pour poste {poste_id}: {str(e)}")
                    messages.error(request, f"Erreur pour le poste: {str(e)}")
            
            # Log détaillé de la programmation
            log_user_action(
                user,
                "Programmation inventaires créée",
                f"Motif: {motif} | Mois: {mois_str} | "
                f"Postes programmés: {len(postes_programmes)} | "
                f"Déjà programmés: {len(postes_deja_programmes)} | "
                f"Liste: {', '.join(postes_programmes[:5])}{'...' if len(postes_programmes) > 5 else ''}",
                request
            )
            
            # Messages de confirmation
            if postes_programmes:
                messages.success(request, f"✅ {len(postes_programmes)} programmation(s) créée(s) avec succès")
            if postes_deja_programmes:
                messages.warning(request, f"⚠️ {len(postes_deja_programmes)} poste(s) déjà programmé(s) pour ce motif")
            
            if postes_programmes or postes_deja_programmes:
                return redirect('inventaire:liste_programmations')
    
    # Log final du contexte
    logger.debug(f"[DEBUG] Contexte final: motif={context.get('motif_selectionne')}, "
                f"postes_auto={len(context.get('postes_taux_auto', []))}, "
                f"postes_manuel={len(context.get('postes_taux_manuel', []))}, "
                f"postes_presence={context.get('postes_presence_admin').count() if 'postes_presence_admin' in context else 0}")
    
    return render(request, 'inventaire/programmer_inventaire.html', context)

@login_required
@require_http_methods(["GET"])
def api_get_postes_par_motif(request):
    """API pour récupérer les postes selon le motif sélectionné"""
    
    motif = request.GET.get('motif')
    mois = request.GET.get('mois')
    
    if not motif:
        return JsonResponse({'error': 'Motif requis'}, status=400)
    
    # Convertir le mois si fourni
    mois_date = None
    if mois:
        try:
            mois_date = datetime.strptime(mois, '%Y-%m').date()
        except:
            pass
    
    data = {
        'postes': [],
        'postes_sugeres': [],
        'total_postes': 0
    }
    
    if motif == MotifInventaire.RISQUE_BAISSE:
        # Récupérer les postes avec risque de baisse
        postes_risque = ProgrammationInventaire.get_postes_avec_risque_baisse()
        
        for item in postes_risque:
            # Vérifier si déjà programmé pour ce mois/motif
            deja_programme = False
            if mois_date:
                deja_programme = ProgrammationInventaire.objects.filter(
                    poste=item['poste'],
                    mois=mois_date,
                    motif=motif,
                    actif=True
                ).exists()
            
            data['postes_sugeres'].append({
                'id': item['poste'].id,
                'nom': item['poste'].nom,
                'code': item['poste'].code,
                'region': item['poste'].get_region_display(),
                'recettes_actuelles': float(item['recettes_actuelles']),
                'recettes_precedentes': float(item['recettes_precedentes']),
                'pourcentage_baisse': float(item['pourcentage_baisse']),
                'selectionne': True,  # Pré-sélectionné
                'deja_programme': deja_programme
            })
    
    elif motif == MotifInventaire.GRAND_STOCK:
        # Récupérer les postes avec grand stock
        postes_stock = ProgrammationInventaire.get_postes_avec_grand_stock()
        
        for item in postes_stock:
            deja_programme = False
            if mois_date:
                deja_programme = ProgrammationInventaire.objects.filter(
                    poste=item['poste'],
                    mois=mois_date,
                    motif=motif,
                    actif=True
                ).exists()
            
            data['postes_sugeres'].append({
                'id': item['poste'].id,
                'nom': item['poste'].nom,
                'code': item['poste'].code,
                'region': item['poste'].get_region_display(),
                'stock_restant': item['stock_restant'],
                'date_epuisement': item['date_epuisement'].strftime('%d/%m/%Y'),
                'jours_restants': item['jours_restants'],
                'selectionne': True,  # Pré-sélectionné
                'deja_programme': deja_programme
            })
    
    elif motif == MotifInventaire.TAUX_DEPERDITION:
        # Récupérer TOUS les postes avec leur taux si disponible
        postes_taux = ProgrammationInventaire.get_postes_avec_taux_deperdition()
        postes_avec_taux_ids = [item['poste'].id for item in postes_taux]
        
        # Postes avec taux (pré-sélectionnés)
        for item in postes_taux:
            deja_programme = False
            if mois_date:
                deja_programme = ProgrammationInventaire.objects.filter(
                    poste=item['poste'],
                    mois=mois_date,
                    motif=motif,
                    actif=True
                ).exists()
            
            data['postes_sugeres'].append({
                'id': item['poste'].id,
                'nom': item['poste'].nom,
                'code': item['poste'].code,
                'region': item['poste'].get_region_display(),
                'taux_deperdition': float(item['taux_deperdition']),
                'date_calcul': item['date_calcul'].strftime('%d/%m/%Y'),
                'alerte': item['alerte'],
                'selectionne': float(item['taux_deperdition']) < -10,  # Sélectionné si taux < -10%
                'deja_programme': deja_programme
            })
        
        # Ajouter les autres postes (non sélectionnés)
        autres_postes = Poste.objects.filter(
            is_active=True, type='peage'
        ).exclude(id__in=postes_avec_taux_ids)
        
        for poste in autres_postes:
            deja_programme = False
            if mois_date:
                deja_programme = ProgrammationInventaire.objects.filter(
                    poste=poste,
                    mois=mois_date,
                    motif=motif,
                    actif=True
                ).exists()
            
            data['postes'].append({
                'id': poste.id,
                'nom': poste.nom,
                'code': poste.code,
                'region': poste.get_region_display(),
                'taux_deperdition': None,
                'selectionne': False,
                'deja_programme': deja_programme
            })
    
    data['total_postes'] = len(data['postes']) + len(data['postes_sugeres'])
    
    return JsonResponse(data)


@login_required
def liste_programmations(request):
    """Vue pour afficher la liste des programmations groupées par poste"""
    
    # Récupérer toutes les programmations actives
    if request.user.is_admin:
        programmations = ProgrammationInventaire.objects.filter(actif=True)
    else:
        programmations = ProgrammationInventaire.objects.filter(
            actif=True,
            poste__in=request.user.get_postes_accessibles()
        )
    
    # Grouper par poste et mois
    programmations_groupees = {}
    
    for prog in programmations.select_related('poste', 'cree_par'):
        key = f"{prog.poste.id}_{prog.mois.strftime('%Y-%m')}"
        
        if key not in programmations_groupees:
            programmations_groupees[key] = {
                'poste': prog.poste,
                'mois': prog.mois,
                'motifs': [],
                'cree_par': prog.cree_par,
                'date_creation': prog.date_creation
            }
        
        # Ajouter les détails du motif
        motif_detail = {
            'type': prog.motif,
            'display': prog.get_motif_display()
        }
        
        if prog.motif == MotifInventaire.TAUX_DEPERDITION:
            motif_detail['taux'] = prog.taux_deperdition_precedent
        elif prog.motif == MotifInventaire.RISQUE_BAISSE:
            motif_detail['pourcentage_baisse'] = prog.pourcentage_baisse
        elif prog.motif == MotifInventaire.GRAND_STOCK:
            motif_detail['stock_restant'] = prog.stock_restant
            motif_detail['date_epuisement'] = prog.date_epuisement_prevu
        
        programmations_groupees[key]['motifs'].append(motif_detail)
    
    # Trier par mois décroissant et poste
    programmations_liste = sorted(
        programmations_groupees.values(),
        key=lambda x: (x['mois'], x['poste'].nom),
        reverse=True
    )
    
    context = {
        'programmations': programmations_liste,
        'user_role': request.user.get_habilitation_display()
    }
    
    return render(request, 'inventaire/liste_programmations.html', context)


@login_required
@require_permission('peut_gerer_inventaire')
@require_http_methods(["POST"])
def desactiver_programmation(request, poste_id, mois, motif):
    """Désactive une programmation spécifique"""
    
    try:
        mois_date = datetime.strptime(mois, '%Y-%m').date()
        
        prog = get_object_or_404(
            ProgrammationInventaire,
            poste_id=poste_id,
            mois=mois_date,
            motif=motif,
            actif=True
        )
        
        prog.actif = False
        prog.save()
        
        messages.success(
            request,
            f"Programmation désactivée pour {prog.poste.nom} - {prog.get_motif_display()}"
        )
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
@require_permission('peut_gerer_inventaire')
@require_http_methods(["DELETE"])
def supprimer_programmation(request, poste_id, mois, motif):
    """Supprime définitivement une programmation"""
    
    try:
        mois_date = datetime.strptime(mois, '%Y-%m').date()
        
        prog = get_object_or_404(
            ProgrammationInventaire,
            poste_id=poste_id,
            mois=mois_date,
            motif=motif
        )
        
        poste_nom = prog.poste.nom
        motif_display = prog.get_motif_display()
        
        prog.delete()
        
        messages.success(
            request,
            f"Programmation supprimée pour {poste_nom} - {motif_display}"
        )
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_permission('peut_gerer_inventaire')
def detail_programmation(request, poste_id, mois):
    """Vue détaillée d'une programmation avec inventaires journaliers"""
    
    poste = get_object_or_404(Poste, id=poste_id)
    
    # Parser le mois (format: 2025-09)
    try:
        mois_date = datetime.strptime(mois, '%Y-%m').date()
    except ValueError:
        messages.error(request, "Format de mois invalide")
        return redirect('inventaire:liste_programmations')
    
    # Récupérer les programmations pour ce poste/mois
    programmations = ProgrammationInventaire.objects.filter(
        poste=poste,
        mois=mois_date,
        actif=True
    ).select_related('cree_par')
    
    if not programmations.exists():
        messages.error(request, "Aucune programmation trouvée")
        return redirect('inventaire:liste_programmations')
    
    # Calculer les dates du mois
    annee = mois_date.year
    mois_num = mois_date.month
    debut_mois = date(annee, mois_num, 1)
    dernier_jour = calendar.monthrange(annee, mois_num)[1]
    fin_mois = date(annee, mois_num, dernier_jour)
    
    # Récupérer les inventaires du mois
    inventaires = InventaireJournalier.objects.filter(
        poste=poste,
        date__range=[debut_mois, fin_mois]
    ).select_related('agent_saisie').prefetch_related('details_periodes').order_by('date')
    
    # Récupérer les recettes avec calculs
    recettes = RecetteJournaliere.objects.filter(
        poste=poste,
        date__range=[debut_mois, fin_mois]
    ).select_related('chef_poste', 'inventaire_associe').order_by('date')
    
    # Créer un dictionnaire date -> données
    donnees_par_jour = {}
    for inv in inventaires:
        donnees_par_jour[inv.date] = {
            'inventaire': inv,
            'recette': None,
            'taux_deperdition': None,
            'couleur_alerte': 'secondary',
            'stock_restant': None
        }
    
    # Ajouter les recettes et calculs
    for rec in recettes:
        if rec.date not in donnees_par_jour:
            donnees_par_jour[rec.date] = {
                'inventaire': None,
                'recette': rec,
                'taux_deperdition': rec.taux_deperdition,
                'couleur_alerte': rec.get_couleur_alerte() if rec.taux_deperdition else 'secondary',
                'stock_restant': rec.stock_tickets_restant
            }
        else:
            donnees_par_jour[rec.date]['recette'] = rec
            donnees_par_jour[rec.date]['taux_deperdition'] = rec.taux_deperdition
            donnees_par_jour[rec.date]['couleur_alerte'] = rec.get_couleur_alerte() if rec.taux_deperdition else 'secondary'
            donnees_par_jour[rec.date]['stock_restant'] = rec.stock_tickets_restant
    
    # Statistiques du mois
    stats = {
        'total_inventaires': inventaires.count(),
        'total_recettes': recettes.count(),
        'jours_saisis': len(donnees_par_jour),
        'total_vehicules': sum(inv.total_vehicules for inv in inventaires),
        'total_recettes_declarees': sum(rec.montant_declare for rec in recettes),
        'total_recettes_potentielles': sum(rec.recette_potentielle or 0 for rec in recettes),
        'taux_moyen': None,
        'jours_risque': 0,
        'jours_impertinents': 0
    }
    
    # Calculer le taux moyen et compter les jours à risque
    taux_list = [rec.taux_deperdition for rec in recettes if rec.taux_deperdition is not None]
    if taux_list:
        stats['taux_moyen'] = sum(taux_list) / len(taux_list)
        stats['jours_risque'] = sum(1 for t in taux_list if t < -30)
        stats['jours_impertinents'] = sum(1 for t in taux_list if t > -5)
    
    # Créer le calendrier
    cal = calendar.monthcalendar(annee, mois_num)
    calendrier_data = []
   
    
    for semaine in cal:
        semaine_data = []
        for jour in semaine:
            if jour == 0:
                semaine_data.append(None)
            else:
                date_jour = date(annee, mois_num, jour)
                jour_data = donnees_par_jour.get(date_jour, {
                    'inventaire': None,
                    'recette': None,
                    'taux_deperdition': None,
                    'couleur_alerte': 'light',
                    'stock_restant': None
                })
                jour_data['date'] = date_jour
                jour_data['jour'] = jour
                semaine_data.append(jour_data)
        calendrier_data.append(semaine_data)
    
    context = {
        'poste': poste,
        'programmations': programmations,
        'mois_date': mois_date,
        'inventaires': inventaires,
        'recettes': recettes,
        'donnees_par_jour': donnees_par_jour,
        'stats': stats,
        'calendrier': calendrier_data,
        'title': f'Détail programmation - {poste.nom} - {mois_date.strftime("%B %Y")}'
    }
    
    return render(request, 'inventaire/detail_programmation.html', context)

@login_required
def api_month_data(request):
    """API pour récupérer les données d'un mois"""
    year = int(request.GET.get('year'))
    month = int(request.GET.get('month'))
    poste_id = request.GET.get('poste_id')
    
    poste = get_object_or_404(Poste, id=poste_id)
    
    # Vérifier si le mois est programmé
    mois_date = date(year, month, 1)
    programmation_exists = ProgrammationInventaire.objects.filter(
        poste=poste,
        mois=mois_date,
        actif=True
    ).exists()
    
    # Récupérer les inventaires existants
    inventaires = InventaireJournalier.objects.filter(
        poste=poste,
        date__year=year,
        date__month=month
    ).values('date', 'id')
    
    data = {}
    for inv in inventaires:
        date_str = inv['date'].strftime('%Y-%m-%d')
        data[date_str] = {
            'has_inventory': True,
            'inventory_id': inv['id'],
            'is_programmable': False
        }
    
    # Marquer les jours programmables
    if programmation_exists:
        from calendar import monthrange
        days_in_month = monthrange(year, month)[1]
        today = timezone.now().date()
        
        for day in range(1, days_in_month + 1):
            date_obj = date(year, month, day)
            date_str = date_obj.strftime('%Y-%m-%d')
            
            if date_str not in data and date_obj <= today:
                data[date_str] = {
                    'has_inventory': False,
                    'is_programmable': True
                }
    
    return JsonResponse(data)

@login_required
@require_permission('peut_gerer_inventaire')
def selection_date_inventaire(request, poste_id=None):
    """Vue intermédiaire pour sélectionner la date avant la saisie"""
    
    # Déterminer le poste
    if poste_id:
        poste = get_object_or_404(Poste, id=poste_id)
    else:
        poste = request.user.poste_affectation
        if not poste:
            messages.error(request, "Aucun poste d'affectation configuré.")
            return redirect('common:dashboard')
    
    # Vérifier l'accès au poste
    if hasattr(request.user, 'peut_acceder_poste'):
        if not request.user.peut_acceder_poste(poste):
            messages.error(request, "Accès non autorisé à ce poste.")
            return redirect('inventaire:inventaire_list')
    
    # Si une date est passée en GET, rediriger vers la saisie
    if request.GET.get('date'):
        date_str = request.GET.get('date')
        return redirect('inventaire:saisie_inventaire_avec_date', 
                       poste_id=poste.id, 
                       date_str=date_str)
    
    # Sinon, afficher le calendrier de sélection
    return render(request, 'inventaire/selection_date_inventaire.html', {
        'poste': poste,
        'poste_id': poste.id,
        'today': timezone.now().date(),
        'current_month': timezone.now().strftime('%Y-%m'),
    })

@login_required
def programmations_desactivees(request):
    """
    Liste des programmations désactivées.
    MISE À JOUR: Permission peut_voir_programmation_desactivee requise
    """
    user = request.user
    
    # Vérifier permission avec système granulaire
    if not has_permission(user, 'peut_voir_programmation_desactivee'):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Programmations désactivées",
            f"Permission manquante: peut_voir_programmation_desactivee",
            request
        )
        messages.error(request, "Vous n'avez pas la permission de voir les programmations désactivées.")
        return redirect('inventaire:liste_programmations')
    
    programmations = ProgrammationInventaire.objects.filter(
        actif=False
    ).select_related('poste', 'cree_par').order_by('-date_creation')
    
    # Log de l'accès
    log_user_action(
        user,
        "Consultation programmations désactivées",
        f"Total: {programmations.count()}",
        request
    )
    
    if request.method == 'POST':
        prog_id = request.POST.get('programmation_id')
        action = request.POST.get('action')
        
        # Vérifier permission de modification
        if not _check_programmation_permission(user):
            messages.error(request, "Vous n'avez pas la permission de modifier les programmations.")
            return redirect('inventaire:programmations_desactivees')
        
        try:
            prog = ProgrammationInventaire.objects.get(id=prog_id)
            
            if action == 'reactiver':
                # Vérifier qu'il n'y a pas de programmation active existante
                existe = ProgrammationInventaire.objects.filter(
                    poste=prog.poste,
                    mois=prog.mois,
                    motif=prog.motif,
                    actif=True
                ).exists()
                
                if existe:
                    messages.error(
                        request, 
                        f"Une programmation active existe déjà pour {prog.poste.nom} - "
                        f"{prog.mois.strftime('%B %Y')} - {prog.get_motif_display()}"
                    )
                else:
                    prog.actif = True
                    prog.save()
                    
                    log_user_action(
                        user,
                        "Réactivation programmation",
                        f"Poste: {prog.poste.nom} | Mois: {prog.mois} | Motif: {prog.get_motif_display()}",
                        request
                    )
                    messages.success(request, "Programmation réactivée avec succès")
                    
            elif action == 'supprimer':
                info_prog = f"{prog.poste.nom} - {prog.mois} - {prog.get_motif_display()}"
                prog.delete()
                
                log_user_action(
                    user,
                    "Suppression programmation",
                    f"Programmation supprimée: {info_prog}",
                    request
                )
                messages.success(request, "Programmation supprimée définitivement")
                
        except ProgrammationInventaire.DoesNotExist:
            messages.error(request, "Programmation introuvable")
    
    context = {
        'programmations': programmations,
        'title': 'Programmations désactivées',
        'can_edit': _check_programmation_permission(user),
    }
    
    return render(request, 'inventaire/programmations_desactivees.html', context)

@login_required
def check_inventaire_exists(request):
    """Vérifie si un inventaire existe pour un poste et une date donnés"""
    poste_id = request.GET.get('poste_id')
    date_str = request.GET.get('date')
    
    if not poste_id or not date_str:
        return JsonResponse({'exists': False})
    
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        exists = InventaireJournalier.objects.filter(
            poste_id=poste_id,
            date=date_obj
        ).exists()
        return JsonResponse({'exists': exists})
    except (ValueError, Exception):
        return JsonResponse({'exists': False})

@login_required
def jours_impertinents_view(request):
    """
    Vue pour afficher les jours impertinents avec détails.
    MISE À JOUR: Permission peut_voir_jours_impertinents requise
    """
    user = request.user
    
    # Vérifier permission avec système granulaire
    if not _check_jours_impertinents_permission(user):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Jours impertinents",
            f"Permission manquante: peut_voir_jours_impertinents | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
            request
        )
        messages.error(request, "Vous n'avez pas la permission de voir les jours impertinents.")
        return redirect('common:dashboard')
    
    # Récupération des paramètres
    periode = request.GET.get('periode', 'mois')
    poste_id = request.GET.get('poste', 'tous')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    # Déterminer les dates selon la période
    if date_debut_str and date_fin_str:
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
    else:
        today = date.today()
        if periode == 'jour':
            date_debut = date_fin = today
        elif periode == 'semaine':
            date_debut = today - timedelta(days=today.weekday())
            date_fin = date_debut + timedelta(days=6)
        elif periode == 'mois':
            date_debut = today.replace(day=1)
            if today.month == 12:
                date_fin = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                date_fin = date(today.year, today.month + 1, 1) - timedelta(days=1)
        else:  # annuel
            date_debut = date(today.year, 1, 1)
            date_fin = date(today.year, 12, 31)
    
    # Récupérer les recettes avec taux > -5%
    jours_enrichis = []
    
    recettes_impertinentes = RecetteJournaliere.objects.filter(
        date__range=[date_debut, date_fin],
        inventaire_associe__isnull=False,
        taux_deperdition__gt=-5
    ).select_related('poste', 'chef_poste', 'inventaire_associe__agent_saisie')
    
    if poste_id != 'tous':
        recettes_impertinentes = recettes_impertinentes.filter(poste_id=poste_id)
    
    for recette in recettes_impertinentes:
        jours_enrichis.append({
            'date': recette.date,
            'poste': recette.poste,
            'agent_inventaire': recette.inventaire_associe.agent_saisie if recette.inventaire_associe else None,
            'inventaire': recette.inventaire_associe,
            'recette': recette,
            'taux_deperdition': recette.taux_deperdition,
            'commentaire': f"TD > -5% : {recette.taux_deperdition:.2f}%"
        })
    
    # Statistiques
    stats = {
        'total_jours': len(jours_enrichis),
        'par_agent': {},
        'par_poste': {}
    }
    
    for jour in jours_enrichis:
        if jour['agent_inventaire']:
            agent_nom = jour['agent_inventaire'].nom_complet
            if agent_nom not in stats['par_agent']:
                stats['par_agent'][agent_nom] = 0
            stats['par_agent'][agent_nom] += 1
        
        poste_nom = jour['poste'].nom
        if poste_nom not in stats['par_poste']:
            stats['par_poste'][poste_nom] = 0
        stats['par_poste'][poste_nom] += 1
    
    # Log de la consultation
    log_user_action(
        user,
        "Consultation jours impertinents",
        f"Période: {periode} | Du {date_debut} au {date_fin} | "
        f"Poste: {poste_id} | Total: {len(jours_enrichis)}",
        request
    )
    
    context = {
        'jours_impertinents': jours_enrichis,
        'stats': stats,
        'periode': periode,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'postes': Poste.objects.filter(is_active=True, type='peage'),
        'poste_selectionne': poste_id,
        'can_view_stats': _check_stats_deperdition_permission(user),
    }
    
    return render(request, 'inventaire/jours_impertinents.html', context)

@login_required
def redirect_to_delete_recette_admin(request, recette_id):
    """Redirection vers la suppression dans l'admin Django"""
    user = request.user
    
    if not user.is_admin:
        messages.error(request, "Accès non autorisé à la suppression.")
        return redirect('inventaire:liste_recettes')
    
    try:
        recette = RecetteJournaliere.objects.get(id=recette_id)
        messages.info(request, f"Suppression de la recette {recette.poste.nom} du {recette.date}")
        return redirect(f'/admin/inventaire/recettejournaliere/{recette_id}/delete/')
    except RecetteJournaliere.DoesNotExist:
        messages.error(request, "Recette non trouvée.")
        return redirect('inventaire:liste_recettes')

@login_required
def gestion_objectifs_annuels(request):
    """
    Vue pour gérer les objectifs annuels.
    
    Permission consultation: peut_voir_objectifs_peage
    Permission modification (POST): is_admin_user (admin_principal, coord_psrr, serv_info)
    
    MISE À JOUR: Contrôle d'accès lecture/écriture séparé
    """
    user = request.user
    
    # Vérifier permission de consultation avec système granulaire
    if not has_permission(user, 'peut_voir_objectifs_peage'):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Objectifs annuels",
            f"Permission manquante: peut_voir_objectifs_peage | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
            request
        )
        messages.error(request, "Vous n'avez pas la permission de voir les objectifs annuels.")
        return redirect('common:dashboard')
    
    # Récupérer l'année (paramètre ou année en cours)
    annee = int(request.GET.get('annee', date.today().year))
    
    # Déterminer si l'utilisateur peut modifier (admin uniquement)
    can_edit = is_admin_user(user)
    
    # Log de l'accès
    log_user_action(
        user,
        "Consultation objectifs annuels",
        f"Année: {annee} | Mode: {'modification' if can_edit else 'lecture'}",
        request
    )
    
    from inventaire.services.objectifs_service import ObjectifsService
    from django.db import transaction
    from decimal import Decimal
    
    # Année sélectionnée
    annee_actuelle = date.today().year
    annees_disponibles = list(range(annee_actuelle - 5, annee_actuelle + 6))
    
    # Récupérer TOUS les postes actifs de type péage
    postes = Poste.objects.filter(is_active=True, type='peage').select_related('region').order_by('region', 'nom')
    
    if request.method == 'POST':
        # ============================================================
        # VÉRIFICATION CRITIQUE: Seuls les admins peuvent modifier
        # ============================================================
        if not can_edit:
            log_user_action(
                user,
                "TENTATIVE MODIFICATION OBJECTIFS NON AUTORISÉE",
                f"Tentative de modification des objectifs péage sans permission | "
                f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
                request
            )
            messages.error(request, "Vous n'avez pas la permission de modifier les objectifs.")
            return redirect(f"{request.path}?annee={annee}")
        
        # Traitement du formulaire (admin uniquement)
        with transaction.atomic():
            objectifs_crees = 0
            objectifs_modifies = 0
            details_modifications = []
            
            for poste in postes:
                montant_key = f'objectif_{poste.id}'
                if montant_key in request.POST:
                    montant_str = request.POST[montant_key].replace(' ', '').replace(',', '')
                    
                    try:
                        montant = Decimal(montant_str) if montant_str else Decimal('0')
                        
                        if montant > 0:
                            objectif, was_created = ObjectifAnnuel.objects.update_or_create(
                                poste=poste,
                                annee=annee,
                                defaults={
                                    'montant_objectif': montant,
                                    'cree_par': request.user
                                }
                            )
                            
                            if was_created:
                                objectifs_crees += 1
                            else:
                                objectifs_modifies += 1
                            
                            details_modifications.append(f"{poste.nom}: {montant:,.0f} FCFA")
                    
                    except (ValueError, TypeError):
                        continue
            
            # Log détaillé de la modification
            log_user_action(
                user,
                "MODIFICATION_OBJECTIFS_PEAGE",
                f"Modification objectifs péage {annee} | "
                f"{objectifs_crees} créés, {objectifs_modifies} modifiés | "
                f"Détails: {', '.join(details_modifications[:5])}{'...' if len(details_modifications) > 5 else ''}",
                request
            )
            
            messages.success(
                request, 
                f"✓ {objectifs_crees} objectifs créés, {objectifs_modifies} modifiés pour {annee}"
            )
            return redirect(f"{request.path}?annee={annee}")
    
    # Construire les données pour TOUS les postes
    objectifs_data = []
    
    for poste in postes:
        # Récupérer l'objectif existant pour cette année
        try:
            objectif = ObjectifAnnuel.objects.get(poste=poste, annee=annee)
            montant_objectif = objectif.montant_objectif
        except ObjectifAnnuel.DoesNotExist:
            # Pas d'objectif : essayer de copier depuis l'année précédente
            try:
                objectif_precedent = ObjectifAnnuel.objects.get(
                    poste=poste, 
                    annee=annee-1
                )
                montant_objectif = objectif_precedent.montant_objectif
            except ObjectifAnnuel.DoesNotExist:
                montant_objectif = Decimal('0')
        
        # Calculer le réalisé pour cette année
        realise = RecetteJournaliere.objects.filter(
            poste=poste,
            date__year=annee
        ).aggregate(Sum('montant_declare'))['montant_declare__sum'] or Decimal('0')
        
        # Calculer le taux
        taux = (realise / montant_objectif * 100) if montant_objectif > 0 else 0
        
        objectifs_data.append({
            'poste': poste,
            'montant_objectif': montant_objectif,
            'realise': realise,
            'reste': montant_objectif - realise,
            'taux': float(taux)
        })
    
    # UTILISER LE SERVICE pour les totaux globaux
    stats_globales = ObjectifsService.calculer_objectifs_annuels(
        annee=annee,
        inclure_postes_inactifs=False
    )
    
    context = {
        'annee': annee,
        'annees_disponibles': annees_disponibles,
        'objectifs_data': objectifs_data,
        'total_global': stats_globales['total_objectif'],
        'total_realise': stats_globales['total_realise'],
        'taux_global': stats_globales['taux_realisation'],
        'title': f'Objectifs Annuels {annee}',
        # Variable clé pour contrôler l'affichage des champs de modification
        'can_edit': can_edit,
    }
    
    return render(request, 'inventaire/gestion_objectifs_annuels.html', context)

@login_required
def dupliquer_objectifs_annee(request):
    """
    Duplique les objectifs d'une année vers une autre
    
    MISE À JOUR - Permissions granulaires:
    - Permission requise: utilisateur admin (is_admin_user)
    - Log détaillé de chaque action utilisateur
    """
    user = request.user
    
    # =========================================
    # VÉRIFICATION PERMISSION GRANULAIRE
    # =========================================
    if not is_admin_user(user):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Duplication objectifs",
            f"Permission admin requise | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | "
            f"IP: {request.META.get('REMOTE_ADDR')}",
            request
        )
        messages.error(request, _("Seuls les administrateurs peuvent dupliquer les objectifs."))
        return redirect('common:dashboard')
    
    if request.method == 'POST':
        annee_source = int(request.POST.get('annee_source'))
        annee_cible = int(request.POST.get('annee_cible'))
        
        # Validation des années
        if annee_cible <= annee_source:
            messages.error(request, "L'année cible doit être supérieure à l'année source.")
            log_user_action(
                user,
                "ERREUR - Duplication objectifs",
                f"Validation échouée: année source ({annee_source}) >= année cible ({annee_cible})",
                request
            )
            return redirect('inventaire:gestion_objectifs_annuels')
        
        objectifs_source = ObjectifAnnuel.objects.filter(annee=annee_source)
        count = 0
        
        for obj in objectifs_source:
            _, created = ObjectifAnnuel.objects.get_or_create(
                poste=obj.poste,
                annee=annee_cible,
                defaults={
                    'montant_objectif': obj.montant_objectif,
                    'cree_par': request.user
                }
            )
            if created:
                count += 1
        
        # Log détaillé de la duplication
        log_user_action(
            user,
            "Duplication objectifs annuels",
            f"Année source: {annee_source} | Année cible: {annee_cible} | "
            f"Objectifs source: {objectifs_source.count()} | "
            f"Objectifs dupliqués: {count}",
            request
        )
        
        messages.success(request, f"✓ {count} objectifs dupliqués de {annee_source} vers {annee_cible}")
    
    return redirect('inventaire:gestion_objectifs_annuels')


from inventaire.services.forecasting_service import ForecastingService

@login_required
def simulateur_commandes(request):
    """
    Simulateur de commandes avec prévisions statistiques.
    MISE À JOUR: Permission peut_simuler_commandes_peage requise
    """
    user = request.user
    
    # Vérifier permission avec système granulaire
    if not has_permission(user, 'peut_simuler_commandes_peage'):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Simulateur commandes",
            f"Permission manquante: peut_simuler_commandes_peage | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
            request
        )
        messages.error(request, "Vous n'avez pas la permission d'accéder au simulateur de commandes.")
        return redirect('common:dashboard')
    
    postes = Poste.objects.filter(is_active=True, type='peage').order_by('nom')
    resultats = None
    erreur = None
    
    if request.method == 'POST':
        poste_id = request.POST.get('poste_id')
        
        if poste_id:
            poste = get_object_or_404(Poste, id=poste_id)
            
            # Utiliser le service de prévisions
            from inventaire.services.forecasting_service import ForecastingService
            resultats_prevision = ForecastingService.calculer_commande_tickets_optimale(poste)
            
            if resultats_prevision['success']:
                resultats = resultats_prevision
                
                # Log de la simulation
                log_user_action(
                    user,
                    "Simulation commande tickets",
                    f"Poste: {poste.nom} | "
                    f"Scénario moyen: {resultats['scenarios']['moyen']['montant']:.0f} FCFA | "
                    f"Jours analysés: {resultats.get('jours_analyses', 'N/A')}",
                    request
                )
            else:
                erreur = resultats_prevision.get('error', 'Erreur inconnue')
                log_user_action(
                    user,
                    "ERREUR - Simulation commande",
                    f"Poste: {poste.nom} | Erreur: {erreur}",
                    request
                )
                messages.error(request, erreur)
    
    context = {
        'postes': postes,
        'resultats': resultats,
        'erreur': erreur,
        'title': 'Simulateur de Commandes de Tickets (Prévisions Avancées)'
    }
    
    return render(request, 'inventaire/simulateur_commandes.html', context)

@login_required
def api_graphique_evolution(request):
    """API pour graphique d'évolution (7 derniers jours)"""
    from datetime import date, timedelta
    from django.db.models import Avg, Sum
    
    today = date.today()
    jours = [(today - timedelta(days=i)) for i in range(6, -1, -1)]
    
    dates = []
    taux_deperdition = []
    recettes = []
    
    for jour in jours:
        dates.append(jour.strftime('%d/%m'))
        
        stats = RecetteJournaliere.objects.filter(date=jour).aggregate(
            taux_moyen=Avg('taux_deperdition'),
            recettes_total=Sum('montant_declare')
        )
        
        taux_deperdition.append(round(float(stats['taux_moyen'] or 0), 1))
        recettes.append(float(stats['recettes_total'] or 0))
    
    return JsonResponse({
        'dates': dates,
        'taux_deperdition': taux_deperdition,
        'recettes': recettes
    })


@login_required
def api_statistiques_postes_ordonnes(request):
    """API pour statistiques postes ordonnées"""
    from django.db.models import Avg, Sum, Count
    
    tri = request.GET.get('tri', 'taux')
    region_filter = request.GET.get('region', '')
    limite = int(request.GET.get('limite', 100))
    
    # Construction requête
    queryset = RecetteJournaliere.objects.filter(
        date__gte=timezone.now().date() - timedelta(days=30)
    ).values(
        'poste__id',
        'poste__nom',
        'poste__code',
        'poste__region__nom'
    ).annotate(
        taux_moyen=Avg('taux_deperdition'),
        recettes_total=Sum('montant_declare'),
        nb_jours=Count('date', distinct=True)
    )
    
    if region_filter:
        queryset = queryset.filter(poste__region__nom=region_filter)
    
    # Tri
    if tri == 'recettes':
        queryset = queryset.order_by('-recettes_total')
    else:
        queryset = queryset.order_by('taux_moyen')
    
    queryset = queryset[:limite]
    
    # Formater les données
    postes_ordonnes = []
    for idx, item in enumerate(queryset):
        taux = float(item['taux_moyen'] or 0)
        
        postes_ordonnes.append({
            'rang': idx + 1,
            'nom': item['poste__nom'],
            'code': item['poste__code'],
            'region': item['poste__region__nom'] or 'Non défini',
            'statistiques': {
                'taux_moyen': round(taux, 1),
                'recettes_total': float(item['recettes_total'] or 0),
                'nb_jours_actifs': item['nb_jours']
            },
            'performance': {
                'score': 100 if taux >= -5 else 80 if taux >= -10 else 60 if taux >= -20 else 40 if taux >= -30 else 20,
                'niveau': 'Excellent' if taux >= -10 else 'Bon' if taux >= -20 else 'Moyen' if taux >= -30 else 'Critique'
            }
        })
    
    return JsonResponse({
        'postes_ordonnes': postes_ordonnes,
        'statistiques_globales': {
            'nb_postes_total': len(postes_ordonnes),
            'taux_global': round(sum([p['statistiques']['taux_moyen'] for p in postes_ordonnes]) / len(postes_ordonnes) if postes_ordonnes else 0, 1)
        }
    })


@login_required
def api_inventaire_stats(request):
    """API pour statistiques inventaires"""
    from django.db.models import Count
    
    today = timezone.now().date()
    
    # Activités récentes
    activites = JournalAudit.objects.select_related('utilisateur').filter(
        timestamp__gte=timezone.now() - timedelta(days=7)
    ).order_by('-timestamp')[:10]
    
    activites_recentes = []
    for act in activites:
        activites_recentes.append({
            'user': act.utilisateur.nom_complet,
            'action': act.action,
            'time': act.timestamp.strftime('%d/%m %H:%M')
        })
    
    return JsonResponse({
        'activites_recentes': activites_recentes,
        'stats': {
            'inventaires_today': InventaireJournalier.objects.filter(date=today).count(),
            'recettes_today': RecetteJournaliere.objects.filter(date=today).count()
        }
    })



@login_required
def calculer_objectifs_automatique(request):
    """
    Vue pour calculer automatiquement les objectifs d'une année
    en appliquant un pourcentage sur l'année précédente
    
    MISE À JOUR - Permissions granulaires:
    - Permission requise: utilisateur admin (is_admin_user)
    - Suppression de @user_passes_test(lambda u: u.is_admin)
    - Log détaillé de chaque action utilisateur
    """
    from inventaire.services.objectifs_service import ObjectifsService
    
    user = request.user
    
    # =========================================
    # VÉRIFICATION PERMISSION GRANULAIRE
    # (Remplace @user_passes_test(lambda u: u.is_admin))
    # =========================================
    if not is_admin_user(user):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Calcul objectifs automatique",
            f"Permission admin requise | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | "
            f"IP: {request.META.get('REMOTE_ADDR')}",
            request
        )
        messages.error(request, _("Seuls les administrateurs peuvent calculer les objectifs automatiquement."))
        return redirect('common:dashboard')
    
    if request.method == 'POST':
        annee_source = int(request.POST.get('annee_source'))
        annee_cible = int(request.POST.get('annee_cible'))
        pourcentage = float(request.POST.get('pourcentage', 0))
        
        # Validation
        if annee_cible <= annee_source:
            messages.error(request, "L'année cible doit être supérieure à l'année source.")
            log_user_action(
                user,
                "ERREUR - Calcul objectifs automatique",
                f"Validation échouée: année cible ({annee_cible}) <= année source ({annee_source})",
                request
            )
            return redirect('inventaire:gestion_objectifs')
        
        if pourcentage < -100 or pourcentage > 500:
            messages.error(request, "Le pourcentage doit être entre -100% et +500%.")
            log_user_action(
                user,
                "ERREUR - Calcul objectifs automatique",
                f"Validation échouée: pourcentage hors limites ({pourcentage}%)",
                request
            )
            return redirect('inventaire:gestion_objectifs')
        
        # Appliquer le calcul
        resultats = ObjectifsService.appliquer_objectifs_calcules(
            annee_source, annee_cible, pourcentage, request.user
        )
        
        if resultats['success']:
            # Log détaillé du succès
            log_user_action(
                user,
                "Calcul objectifs automatique réussi",
                f"Année source: {annee_source} | Année cible: {annee_cible} | "
                f"Pourcentage: {pourcentage:+.1f}% | "
                f"Objectifs créés: {resultats['objectifs_crees']} | "
                f"Objectifs modifiés: {resultats['objectifs_modifies']} | "
                f"Total: {resultats['total_objectif_cible']:,.0f} FCFA",
                request
            )
            
            messages.success(
                request,
                f"✓ Objectifs {annee_cible} calculés avec succès : "
                f"{resultats['objectifs_crees']} créés, {resultats['objectifs_modifies']} modifiés. "
                f"Total : {resultats['total_objectif_cible']:,.0f} FCFA "
                f"({pourcentage:+.1f}% par rapport à {annee_source})"
            )
        else:
            # Log de l'échec
            log_user_action(
                user,
                "ERREUR - Calcul objectifs automatique",
                f"Année source: {annee_source} | Année cible: {annee_cible} | "
                f"Message: {resultats.get('message', 'Erreur inconnue')}",
                request
            )
            messages.error(request, resultats.get('message', 'Erreur lors du calcul'))
        
        return redirect(f"/inventaire/objectifs-annuels/?annee={annee_cible}")
    
    # GET : afficher le formulaire de calcul
    annee_actuelle = date.today().year
    annees = list(range(annee_actuelle - 5, annee_actuelle + 6))
    
    # Log de l'accès à la page
    log_user_action(
        user,
        "Accès page calcul objectifs automatique",
        f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
        request
    )
    
    context = {
        'annees': annees,
        'annee_defaut_source': annee_actuelle - 1,
        'annee_defaut_cible': annee_actuelle,
        'title': 'Calculer Objectifs Automatiquement'
    }
    
    return render(request, 'inventaire/calculer_objectifs.html', context)


@login_required
def saisie_quittancement(request):
    """
    Vue pour saisir un quittancement.
    MISE À JOUR: Permission peut_saisir_quittance_peage requise
    """
    user = request.user
    
    # Vérifier permission avec système granulaire
    if not _check_quittance_permission(user):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Saisie quittancement",
            f"Permission manquante: peut_saisir_quittance_peage | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
            request
        )
        messages.error(request, "Vous n'avez pas la permission de saisir des quittancements.")
        return redirect('common:dashboard')
    
    # Nettoyer la session au début si nouvelle saisie
    if request.method == 'GET' and 'etape' not in request.GET:
        for key in ['exercice_temp', 'mois_temp', 'type_declaration_temp', 
                   'form_data_temp', 'image_temp_path']:
            request.session.pop(key, None)
    
    etape = int(request.POST.get('etape', request.GET.get('etape', 1)))
    
    # Log de l'accès à l'étape
    log_user_action(
        user,
        f"Saisie quittancement - Étape {etape}",
        f"Méthode: {request.method}",
        request
    )
    
    # ============================================================
    # ÉTAPE 1 : Paramètres globaux
    # ============================================================
    if etape == 1:
        if request.method == 'POST':
            exercice = request.POST.get('exercice')
            mois = request.POST.get('mois')  # Format: YYYY-MM
            type_declaration = request.POST.get('type_declaration')
            
            # Validation
            errors = []
            if not exercice:
                errors.append("Veuillez sélectionner un exercice")
            if not mois:
                errors.append("Veuillez sélectionner un mois")
            if not type_declaration:
                errors.append("Veuillez sélectionner un type de déclaration")
            
            if errors:
                for error in errors:
                    messages.error(request, f"⚠️ {error}")
                return redirect('inventaire:saisie_quittancement')
            
            # Stocker en session
            request.session['exercice_temp'] = exercice
            request.session['mois_temp'] = mois
            request.session['type_declaration_temp'] = type_declaration
            
            # Rediriger vers étape 2
            return redirect(f"{reverse('inventaire:saisie_quittancement')}?etape=2")
        
       # Retourner le rendu de l'étape 1 par défaut
        from datetime import datetime
        annee_courante = datetime.now().year
        annees = list(range(annee_courante - 5, annee_courante + 2))
        
        context = {
            'etape': etape,
            'annees': annees,
            'annee_courante': annee_courante,
            'types_declaration': [
                ('journaliere', 'Journalière (par jour)'),
                ('decade', 'Par décade')
            ],
        }
        return render(request, 'inventaire/saisie_quittancement_simple.html', context)    
    # ============================================================
    # ÉTAPE 2 : Saisie du quittancement + IMAGE
    # ============================================================
    elif etape == 2:
        # Récupérer les paramètres de session
        exercice = request.session.get('exercice_temp')
        mois = request.session.get('mois_temp')
        type_declaration = request.session.get('type_declaration_temp')
        
        if not all([exercice, mois, type_declaration]):
            messages.error(request, "❌ Session expirée, veuillez recommencer")
            return redirect('inventaire:saisie_quittancement')
        
        if request.method == 'POST':
            # Récupérer les données du formulaire
            form_data = {
                'numero_quittance': request.POST.get('numero_quittance'),
                'date_quittancement': request.POST.get('date_quittancement'),
                'montant': request.POST.get('montant'),
                'observations': request.POST.get('observations', ''),
            }
            
            # Poste
            if request.user.is_admin:
                form_data['poste_id'] = request.POST.get('poste')
            else:
                form_data['poste_id'] = request.user.poste_affectation.id if request.user.poste_affectation else None
            
            # Dates selon le type
            if type_declaration == 'journaliere':
                form_data['date_recette'] = request.POST.get('date_recette')
            else:  # decade
                form_data['date_debut_decade'] = request.POST.get('date_debut_decade')
                form_data['date_fin_decade'] = request.POST.get('date_fin_decade')
            
            # Gérer l'image uploadée
            has_image = False
            image_name = None
            if 'image_quittance' in request.FILES:
                image_file = request.FILES['image_quittance']
                import os
                from django.core.files.storage import default_storage
                
                temp_name = f"temp_{request.user.id}_{timezone.now().timestamp()}_{image_file.name}"
                temp_path = os.path.join('temp_quittances', temp_name)
                saved_path = default_storage.save(temp_path, image_file)
                request.session['image_temp_path'] = saved_path
                has_image = True
                image_name = image_file.name
                
                logger.info(f"Image temporaire sauvegardée : {saved_path}")
            
            form_data['has_image'] = has_image
            form_data['image_name'] = image_name
            
            # =====================================================
            # VALIDATION COMPLÈTE
            # =====================================================
            errors = []
            
            # --- Validation numéro de quittance ---
            if not form_data['numero_quittance']:
                errors.append("Le numéro de quittance est obligatoire")
            else:
                numero = form_data['numero_quittance'].strip().upper()
                
                # Vérifier dans la table Quittancement (péage)
                if Quittancement.objects.filter(numero_quittance__iexact=numero).exists():
                    errors.append(f"Le numéro de quittance '{numero}' existe déjà dans les quittancements péage")
                
                # Vérifier dans la table QuittancementPesage (pesage)
                try:
                    from inventaire.models_pesage import QuittancementPesage
                    if QuittancementPesage.objects.filter(numero_quittance__iexact=numero).exists():
                        errors.append(f"Le numéro de quittance '{numero}' existe déjà dans les quittancements pesage")
                except ImportError:
                    pass
            
            # --- Validation autres champs ---
            if not form_data['date_quittancement']:
                errors.append("La date de quittancement est obligatoire")
            if not form_data['montant']:
                errors.append("Le montant est obligatoire")
            if not form_data['poste_id']:
                errors.append("Le poste est obligatoire")
            
            # --- Image obligatoire ---
            if not has_image:
                errors.append("L'image de la quittance est obligatoire")
            
            # --- Validation dates selon type ---
            if type_declaration == 'journaliere':
                if not form_data.get('date_recette'):
                    errors.append("La date de recette est obligatoire")
            else:
                if not form_data.get('date_debut_decade'):
                    errors.append("La date de début de décade est obligatoire")
                if not form_data.get('date_fin_decade'):
                    errors.append("La date de fin de décade est obligatoire")
                # Vérifier que date_debut <= date_fin
                elif form_data.get('date_debut_decade') and form_data.get('date_fin_decade'):
                    if form_data['date_debut_decade'] > form_data['date_fin_decade']:
                        errors.append("La date de début doit être antérieure à la date de fin")
            
            # =====================================================
            # ✅ NOUVEAU : VALIDATION DES DATES DÉJÀ QUITTANCÉES
            # =====================================================
            if form_data['poste_id'] and not errors:  # Seulement si pas d'autres erreurs
                try:
                    poste = Poste.objects.get(id=form_data['poste_id'])
                    
                    # Importer la fonction de validation
                    from inventaire.utils_quittancement import valider_dates_quittancement_peage
                    
                    if type_declaration == 'journaliere':
                        is_valid, error_msg, _ = valider_dates_quittancement_peage(
                            poste=poste,
                            exercice=int(exercice),
                            mois=mois,
                            type_declaration='journaliere',
                            date_recette=form_data.get('date_recette')
                        )
                    else:  # decade
                        is_valid, error_msg, _ = valider_dates_quittancement_peage(
                            poste=poste,
                            exercice=int(exercice),
                            mois=mois,
                            type_declaration='decade',
                            date_debut=form_data.get('date_debut_decade'),
                            date_fin=form_data.get('date_fin_decade')
                        )
                    
                    if not is_valid:
                        errors.append(error_msg)
                        
                except Poste.DoesNotExist:
                    errors.append("Poste invalide")
                except Exception as e:
                    logger.error(f"Erreur validation dates: {e}")
            
            # =====================================================
            # GESTION DES ERREURS
            # =====================================================
            if errors:
                for error in errors:
                    messages.error(request, f"❌ {error}")
                
                # Supprimer l'image temporaire en cas d'erreur
                if has_image and request.session.get('image_temp_path'):
                    from django.core.files.storage import default_storage
                    temp_path = request.session.get('image_temp_path')
                    if default_storage.exists(temp_path):
                        default_storage.delete(temp_path)
                    request.session.pop('image_temp_path', None)
                
                # Réafficher le formulaire avec les données
                if request.user.is_admin:
                    postes = Poste.objects.filter(is_active=True, type='peage')
                else:
                    postes = [request.user.poste_affectation] if request.user.poste_affectation else []
                
                context = {
                    'etape': 2,
                    'exercice': exercice,
                    'mois': mois,
                    'type_declaration': type_declaration,
                    'postes': postes,
                    'form_data': form_data,
                }
                return render(request, 'inventaire/saisie_quittancement_simple.html', context)
            
            # Stocker les données validées en session
            request.session['form_data_temp'] = form_data
            
            # Rediriger vers étape 3
            return redirect(f"{reverse('inventaire:saisie_quittancement')}?etape=3")
        
        # GET : Afficher le formulaire vide
        if request.user.is_admin:
            postes = Poste.objects.filter(is_active=True, type='peage')
        else:
            postes = [request.user.poste_affectation] if request.user.poste_affectation else []
        
        context = {
            'etape': 2,
            'exercice': exercice,
            'mois': mois,
            'type_declaration': type_declaration,
            'postes': postes,
        }
        return render(request, 'inventaire/saisie_quittancement_simple.html', context)

    
    # ============================================================
    # ÉTAPE 3 : Confirmation et enregistrement
    # ============================================================
    elif etape == 3:
        # Récupérer toutes les données de session
        exercice = request.session.get('exercice_temp')
        mois = request.session.get('mois_temp')
        type_declaration = request.session.get('type_declaration_temp')
        form_data = request.session.get('form_data_temp')
        image_temp_path = request.session.get('image_temp_path')
        
        if not all([exercice, mois, type_declaration, form_data]):
            messages.error(request, "❌ Session expirée")
            return redirect('inventaire:saisie_quittancement')
        
        if request.method == 'POST':
            action = request.POST.get('action')
            
            if action == 'confirmer':
                try:
                    # Récupérer le poste
                    poste = Poste.objects.get(id=form_data['poste_id'])
                    
                    # Créer le quittancement
                    quittancement = Quittancement(
                        numero_quittance=form_data['numero_quittance'],
                        poste=poste,
                        exercice=int(exercice),
                        mois=mois,
                        type_declaration=type_declaration,
                        date_quittancement=form_data['date_quittancement'],
                        montant=Decimal(form_data['montant']),
                        observations=form_data.get('observations', ''),
                        saisi_par=request.user,
                    )
                    
                    # Dates selon le type
                    if type_declaration == 'journaliere':
                        quittancement.date_recette = form_data['date_recette']
                    else:
                        quittancement.date_debut_decade = form_data['date_debut_decade']
                        quittancement.date_fin_decade = form_data['date_fin_decade']
                    
                    # ✅ CORRECTION : Récupérer l'image temporaire
                    if image_temp_path:
                        from django.core.files.storage import default_storage
                        import os
                        
                        # Lire le fichier temporaire
                        if default_storage.exists(image_temp_path):
                            temp_file = default_storage.open(image_temp_path, 'rb')
                            
                            # Créer le nom final
                            original_name = os.path.basename(image_temp_path).split('_', 3)[-1]
                            final_name = f"quittances/{quittancement.exercice}/{quittancement.mois}/{original_name}"
                            
                            # Sauvegarder dans le modèle
                            from django.core.files import File
                            quittancement.image_quittance.save(final_name, File(temp_file), save=False)
                            temp_file.close()
                            
                            # Supprimer le fichier temporaire
                            default_storage.delete(image_temp_path)
                            
                            logger.info(f"Image déplacée de {image_temp_path} vers {final_name}")
                    
                    # Sauvegarder (avec validation automatique)
                    quittancement.save()
                    
                    # Journaliser
                    log_user_action(
                        request.user,
                        "Création quittancement",
                        f"N°{quittancement.numero_quittance} - {poste.nom} - {quittancement.montant} FCFA"
                        + (f" | Image: Oui" if quittancement.image_quittance else " | Image: Non"),
                        request
                    )
                    
                    # Nettoyer la session
                    for key in ['exercice_temp', 'mois_temp', 'type_declaration_temp', 'form_data_temp', 'image_temp_path']:
                        request.session.pop(key, None)
                    
                    messages.success(
                        request, 
                        f"✅ Quittancement {quittancement.numero_quittance} enregistré avec succès"
                        + (" avec image" if quittancement.image_quittance else "")
                    )
                    return redirect('inventaire:liste_quittancements')
                
                except ValidationError as e:
                    messages.error(request, f"❌ Erreur de validation : {e}")
                except Exception as e:
                    messages.error(request, f"❌ Erreur : {str(e)}")
                    logger.error(f"Erreur création quittancement : {str(e)}", exc_info=True)
            
            elif action == 'retour':
                # Retour à l'étape 2
                return redirect(f"{reverse('inventaire:saisie_quittancement')}?etape=2")
            
            else:  # Annuler
                # Supprimer l'image temporaire si elle existe
                if image_temp_path:
                    from django.core.files.storage import default_storage
                    if default_storage.exists(image_temp_path):
                        default_storage.delete(image_temp_path)
                
                # Nettoyer la session
                for key in ['exercice_temp', 'mois_temp', 'type_declaration_temp', 'form_data_temp', 'image_temp_path']:
                    request.session.pop(key, None)
                messages.info(request, "Saisie annulée")
                return redirect('inventaire:liste_quittancements')
        
        # GET : Afficher la confirmation
        try:
            poste = Poste.objects.get(id=form_data['poste_id'])
        except:
            poste = None
        
        # Formatter la période
        if type_declaration == 'journaliere':
            periode = f"Jour : {form_data['date_recette']}"
        else:
            periode = f"Décade : du {form_data['date_debut_decade']} au {form_data['date_fin_decade']}"
        
        context = {
            'etape': 3,
            'exercice': exercice,
            'mois': mois,
            'type_declaration': type_declaration,
            'form_data': form_data,
            'poste': poste,
            'periode': periode,
            'has_image': form_data.get('has_image', False),
            'image_name': form_data.get('image_name'),
        }
        return render(request, 'inventaire/saisie_quittancement_simple.html', context)
    
@login_required
def liste_quittancements(request):
    """
    Liste des quittancements avec pagination et statistiques.
    
    MISE À JOUR - Corrections:
    1. Permission peut_voir_liste_quittances_peage requise
    2. Ajout de toutes les variables de contexte pour le template
    3. Pagination avec page_obj
    4. Statistiques (nombre_quittancements, total_montant)
    5. Log détaillé des actions utilisateur
    """
    user = request.user
    
    # =========================================
    # VÉRIFICATION PERMISSION GRANULAIRE
    # =========================================
    if not has_permission(user, 'peut_voir_liste_quittances_peage'):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Liste quittancements",
            f"Permission manquante: peut_voir_liste_quittances_peage | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | "
            f"IP: {request.META.get('REMOTE_ADDR')}",
            request
        )
        messages.error(request, "Vous n'avez pas la permission de voir les quittancements.")
        return redirect('common:dashboard')
    
    # =========================================
    # FILTRAGE SELON LES PERMISSIONS
    # =========================================
    
    # Déterminer l'accès aux postes
    if user_has_acces_tous_postes(user):
        quittancements = Quittancement.objects.all()
        postes_disponibles = Poste.objects.filter(is_active=True, type='peage').order_by('nom')
    elif user.poste_affectation:
        quittancements = Quittancement.objects.filter(poste=user.poste_affectation)
        postes_disponibles = None  # Pas de filtre poste si accès limité
    else:
        quittancements = Quittancement.objects.none()
        postes_disponibles = None
    
    # =========================================
    # APPLICATION DES FILTRES GET
    # =========================================
    
    poste_id = request.GET.get('poste')
    exercice = request.GET.get('exercice')
    type_declaration = request.GET.get('type_declaration')
    mois = request.GET.get('mois')  # Format: YYYY-MM
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    numero = request.GET.get('numero')
    montant_min = request.GET.get('montant_min')
    
    # Appliquer les filtres
    if poste_id:
        quittancements = quittancements.filter(poste_id=poste_id)
    
    if exercice:
        quittancements = quittancements.filter(exercice=exercice)
    
    if type_declaration:
        quittancements = quittancements.filter(type_declaration=type_declaration)
    
    if mois:
        # Format attendu: YYYY-MM
        quittancements = quittancements.filter(mois=mois)
    
    if date_debut:
        quittancements = quittancements.filter(date_quittancement__gte=date_debut)
    
    if date_fin:
        quittancements = quittancements.filter(date_quittancement__lte=date_fin)
    
    if numero:
        quittancements = quittancements.filter(numero_quittance__icontains=numero)
    
    if montant_min:
        try:
            montant_min_val = Decimal(montant_min)
            quittancements = quittancements.filter(montant__gte=montant_min_val)
        except (ValueError, InvalidOperation):
            pass
    
    # =========================================
    # OPTIMISATION ET TRI
    # =========================================
    
    quittancements = quittancements.select_related(
        'poste', 
        'saisi_par'
    ).order_by('-date_saisie', '-date_quittancement')
    
    # =========================================
    # CALCUL DES STATISTIQUES
    # =========================================
    
    # Nombre total de quittancements (avant pagination)
    nombre_quittancements = quittancements.count()
    
    # Montant total (avant pagination)
    total_montant = quittancements.aggregate(
        total=Sum('montant')
    )['total'] or Decimal('0')
    
    # =========================================
    # PAGINATION
    # =========================================
    
    paginator = Paginator(quittancements, 25)  # 25 éléments par page
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.get_page(page_number)
    except Exception:
        page_obj = paginator.get_page(1)
    
    # =========================================
    # PRÉPARATION DES DONNÉES POUR LES FILTRES
    # =========================================
    
    # Année courante et plage d'années pour le filtre
    annee_courante = timezone.now().year
    years_range = list(range(annee_courante - 5, annee_courante + 2))
    
    # =========================================
    # LOG DE LA CONSULTATION
    # =========================================
    
    filtres_actifs = []
    if poste_id:
        filtres_actifs.append(f"poste={poste_id}")
    if exercice:
        filtres_actifs.append(f"exercice={exercice}")
    if mois:
        filtres_actifs.append(f"mois={mois}")
    if type_declaration:
        filtres_actifs.append(f"type={type_declaration}")
    if numero:
        filtres_actifs.append(f"numero={numero}")
    
    log_user_action(
        user,
        "Consultation liste quittancements",
        f"Total résultats: {nombre_quittancements} | "
        f"Page: {page_number} | "
        f"Filtres: {', '.join(filtres_actifs) if filtres_actifs else 'Aucun'}",
        request
    )
    
    # =========================================
    # CONTEXTE POUR LE TEMPLATE
    # =========================================
    
    context = {
        # Variables de pagination (CRITIQUE - le template utilise page_obj)
        'page_obj': page_obj,
        
        # Statistiques affichées dans les cartes
        'nombre_quittancements': nombre_quittancements,
        'total_montant': total_montant,
        'exercice_courant': annee_courante,
        
        # Liste des postes pour le filtre (admin uniquement)
        'postes': postes_disponibles,
        
        # Plage d'années pour le filtre exercice
        'years_range': years_range,
        
        # Permissions pour les boutons d'action
        'can_edit': _check_quittance_permission(user),
        'can_comptabiliser': _check_comptabilisation_permission(user),
        
        # Accès admin pour affichage conditionnel
        'is_admin': user_has_acces_tous_postes(user),
    }
    
    return render(request, 'inventaire/liste_quittancements.html', context)



@login_required
def comptabilisation_quittancements(request):
    """
    Vue pour comptabiliser les quittancements.
    MISE À JOUR: Permission peut_comptabiliser_quittances_peage requise
    """
    user = request.user
    
    # Vérifier permission avec système granulaire
    if not _check_comptabilisation_permission(user):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Comptabilisation quittancements",
            f"Permission manquante: peut_comptabiliser_quittances_peage | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
            request
        )
        messages.error(request, "Vous n'avez pas la permission de comptabiliser les quittancements.")
        return redirect('common:dashboard')
    
    # Log de l'accès autorisé
    log_user_action(
        user,
        "Accès comptabilisation quittancements",
        f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
        request
    )
    import calendar
    from datetime import timedelta, date
    from decimal import Decimal

    # Récupération des paramètres
    annee_courante = timezone.now().year
    mois_courant = timezone.now().month

    periode = request.GET.get('periode', 'mois')
    poste_id = request.GET.get('poste')
    annee = int(request.GET.get('annee', annee_courante))
    mois = request.GET.get('mois')  # Format: YYYY-MM

    # Calcul des dates selon la période
    if periode == 'mois':
        if mois:
            try:
                annee_mois = mois.split('-')
                mois_int = int(annee_mois[1])
                annee = int(annee_mois[0])
            except:
                mois_int = mois_courant
        else:
            mois_int = mois_courant
        date_debut = date(annee, mois_int, 1)
        dernier_jour = calendar.monthrange(annee, mois_int)[1]
        date_fin = date(annee, mois_int, dernier_jour)
    elif periode == 'trimestre':
        trimestre = int(request.GET.get('trimestre', ((mois_courant - 1) // 3) + 1))
        mois_debut = (trimestre - 1) * 3 + 1
        mois_fin = trimestre * 3
        date_debut = date(annee, mois_debut, 1)
        dernier_jour = calendar.monthrange(annee, mois_fin)[1]
        date_fin = date(annee, mois_fin, dernier_jour)
    elif periode == 'semestre':
        semestre = int(request.GET.get('semestre', 1 if mois_courant <= 6 else 2))
        if semestre == 1:
            date_debut = date(annee, 1, 1)
            date_fin = date(annee, 6, 30)
        else:
            date_debut = date(annee, 7, 1)
            date_fin = date(annee, 12, 31)
    else:  # année
        date_debut = date(annee, 1, 1)
        date_fin = date(annee, 12, 31)

    # Filtrer les postes selon permissions
    if request.user.is_admin:
        if poste_id:
            postes = Poste.objects.filter(id=poste_id, is_active=True, type='peage')
        else:
            postes = Poste.objects.filter(is_active=True, type='peage')
    else:
        if request.user.poste_affectation:
            postes = Poste.objects.filter(id=request.user.poste_affectation.id)
        else:
            postes = Poste.objects.none()

    resultats = []
    total_declare_global = Decimal('0')
    total_quittance_global = Decimal('0')

    for poste in postes:
        # === LOGIQUE CORRIGÉE ===
        
        # 1. Récupérer toutes les recettes journalières de la période
        recettes_journalieres = RecetteJournaliere.objects.filter(
            poste=poste,
            date__range=[date_debut, date_fin]
        )
        
        # 2. Récupérer les quittancements journaliers
        quittancements_journaliers = Quittancement.objects.filter(
            poste=poste,
            type_declaration='journaliere',
            date_recette__range=[date_debut, date_fin]
        )
        
        # 3. Récupérer les quittancements de décade qui chevauchent la période
        quittancements_decades = Quittancement.objects.filter(
            poste=poste,
            type_declaration='decade'
        ).filter(
            models.Q(date_debut_decade__lte=date_fin) & 
            models.Q(date_fin_decade__gte=date_debut)
        )
        
        # 4. Calculer le total des recettes déclarées
        total_declare = recettes_journalieres.aggregate(
            Sum('montant_declare')
        )['montant_declare__sum'] or Decimal('0')
        
        # 5. Calculer le total des quittancements
        total_quittance = Decimal('0')
        ecart_details = []
        statut = 'conforme'
        statut_label = 'Conforme'
        statut_class = 'success'
        
        # 5a. Ajouter les quittancements journaliers
        for q_jour in quittancements_journaliers:
            total_quittance += q_jour.montant
            
            # Vérifier l'écart pour ce jour spécifique
            recette_jour = recettes_journalieres.filter(
                date=q_jour.date_recette
            ).aggregate(Sum('montant_declare'))['montant_declare__sum'] or Decimal('0')
            
            ecart_jour = q_jour.montant - recette_jour
            if abs(ecart_jour) >= 1:
                ecart_details.append({
                    'date': q_jour.date_recette,
                    'ecart': ecart_jour,
                    'type': 'journaliere'
                })
        
        # 5b. Traiter les quittancements de décade
        jours_incomplets = []
        for q_decade in quittancements_decades:
            # Déterminer la période effective de la décade dans notre période d'analyse
            debut_effectif = max(q_decade.date_debut_decade, date_debut)
            fin_effective = min(q_decade.date_fin_decade, date_fin)
            
            # Vérifier que tous les jours de la décade ont des recettes
            dates_decade = []
            current_date = q_decade.date_debut_decade
            while current_date <= q_decade.date_fin_decade:
                dates_decade.append(current_date)
                current_date += timedelta(days=1)
            
            # Récupérer les recettes de cette décade
            recettes_decade = recettes_journalieres.filter(
                date__in=dates_decade
            )
            
            # Vérifier si tous les jours ont des recettes
            dates_avec_recettes = set(recettes_decade.values_list('date', flat=True))
            dates_manquantes = [d for d in dates_decade if d not in dates_avec_recettes]
            
            if dates_manquantes:
                # Décade incomplète
                jours_incomplets.extend(dates_manquantes)
                statut = 'incomplet'
                statut_label = f'Données incomplètes ({len(dates_manquantes)} jour(s) manquant(s))'
                statut_class = 'warning'
            else:
                # Calculer la somme des recettes de la décade
                somme_recettes_decade = recettes_decade.aggregate(
                    Sum('montant_declare')
                )['montant_declare__sum'] or Decimal('0')
                
                # L'écart est : montant quittancé - somme des recettes
                ecart_decade = q_decade.montant - somme_recettes_decade
                
                # Si la décade est dans notre période, compter son montant
                if debut_effectif <= fin_effective:
                    total_quittance += q_decade.montant
                    
                    if abs(ecart_decade) >= 1:
                        ecart_details.append({
                            'debut': q_decade.date_debut_decade,
                            'fin': q_decade.date_fin_decade,
                            'ecart': ecart_decade,
                            'type': 'decade'
                        })
        
        # 6. Calculer l'écart global
        ecart = total_quittance - total_declare
        ecart_pourcentage = (ecart / total_declare * 100) if total_declare > 0 else Decimal('0')
        
        # 7. Déterminer le statut final
        if jours_incomplets:
            statut = 'incomplet'
            statut_label = f'Données incomplètes'
            statut_class = 'warning'
        elif abs(ecart) < 1:
            statut = 'conforme'
            statut_label = 'Conforme'
            statut_class = 'success'
        else:
            # Vérifier si justifié
            justification_existe = JustificationEcart.objects.filter(
                poste=poste,
                date_debut=date_debut,
                date_fin=date_fin
            ).exists()
            
            if justification_existe:
                statut = 'justifie'
                statut_label = 'Justifié'
                statut_class = 'info'
            else:
                statut = 'ecart'
                statut_label = 'Non justifié'
                statut_class = 'danger'
        
        # Ajouter au résultat
        resultats.append({
            'poste': poste,
            'poste_id': poste.id,
            'poste_nom': poste.nom,
            'total_quittance': total_quittance,
            'total_declare': total_declare,
            'ecart': ecart,
            'ecart_pourcentage': ecart_pourcentage,
            'ecart_details': ecart_details,
            'jours_incomplets': jours_incomplets,
            'justifie': statut == 'justifie',
            'statut': statut,
            'statut_label': statut_label,
            'statut_class': statut_class
        })
        
        total_declare_global += total_declare
        total_quittance_global += total_quittance

    # Calculer les statistiques globales
    ecart_global = total_quittance_global - total_declare_global
    
    statistiques = {
        'total_declare': total_declare_global,
        'total_quittance': total_quittance_global,
        'ecart_total': ecart_global,
        'ecart_pourcentage': (ecart_global / total_declare_global * 100) if total_declare_global > 0 else Decimal('0'),
        'nombre_postes': len(resultats),
        'nombre_conformes': len([r for r in resultats if r['statut'] == 'conforme']),
        'nombre_justifies': len([r for r in resultats if r['statut'] == 'justifie']),
        'nombre_ecarts': len([r for r in resultats if r['statut'] == 'ecart']),
        'nombre_incomplets': len([r for r in resultats if r['statut'] == 'incomplet']),
    }
    
    # Postes pour filtrage (admin seulement)
    postes_filtre = Poste.objects.filter(is_active=True, type='peage').order_by('nom') if request.user.is_admin else None
    
    # Générer la liste des années disponibles
    current_year = timezone.now().year
    annees_disponibles = list(range(current_year - 5, current_year + 1))
    
    # Journaliser l'action
    log_user_action(
        request.user,
        "Consultation comptabilisation quittancements",
        f"Période: {periode}, Date: {date_debut} au {date_fin}, Postes: {len(resultats)}",
        request
    )
    
    context = {
        'resultats': resultats,
        'statistiques': statistiques,
        'periode': periode,
        'annee': annee,
        'annees_disponibles': annees_disponibles,
        'mois': f"{annee}-{mois_int:02d}" if periode == 'mois' else None,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'postes_filtre': postes_filtre,
        'poste_selectionne': int(poste_id) if poste_id else None,
        'postes': postes_filtre,  
        'title': 'Comptabilisation des Quittancements',
        'can_justify': _check_comptabilisation_permission(user),

    }
    
    return render(request, 'inventaire/comptabilisation_quittancements.html', context)

@login_required
def justifier_ecart_periode(request, poste_id, date_debut, date_fin):
    """
    Vue pour justifier un écart entre quittancements et déclarations
    
    MISE À JOUR - Permissions granulaires:
    - Remplace: request.user.is_admin or request.user.is_chef_poste
    - Par: is_admin_user(user) or is_chef_poste(user) avec check_poste_access()
    - Log détaillé de chaque action utilisateur
    
    Args:
        request: Requête HTTP
        poste_id: ID du poste (int)
        date_debut: Date de début au format 'YYYY-MM-DD' (str)
        date_fin: Date de fin au format 'YYYY-MM-DD' (str)
    """
    user = request.user
    
    # Récupérer le poste
    poste = get_object_or_404(Poste, id=poste_id)
    
    # =========================================
    # VÉRIFICATION PERMISSIONS GRANULAIRES
    # (Remplace: request.user.is_admin or request.user.is_chef_poste)
    # =========================================
    
    # Vérifier si l'utilisateur a les droits de base (admin ou chef de poste)
    has_base_permission = is_admin_user(user) or is_chef_poste(user)
    
    if not has_base_permission:
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Justification écart",
            f"Habilitation insuffisante | "
            f"Habilitation actuelle: {getattr(user, 'habilitation', 'N/A')} | "
            f"Poste demandé: {poste.nom} | "
            f"IP: {request.META.get('REMOTE_ADDR')}",
            request
        )
        messages.error(
            request,
            "Seuls les administrateurs et chefs de poste peuvent justifier les écarts."
        )
        return redirect('inventaire:comptabilisation_quittancements')
    
    # Pour les chefs de poste non-admin, vérifier l'accès au poste spécifique
    if not is_admin_user(user):
        if not check_poste_access(user, poste):
            log_user_action(
                user,
                "ACCÈS REFUSÉ - Justification écart (poste non autorisé)",
                f"Chef de poste sans accès au poste demandé | "
                f"Poste demandé: {poste.nom} (ID: {poste_id}) | "
                f"Poste affectation: {getattr(user.poste_affectation, 'nom', 'Aucun')}",
                request
            )
            messages.error(
                request,
                "Vous n'avez pas accès à ce poste pour justifier les écarts."
            )
            return redirect('inventaire:comptabilisation_quittancements')
    
    # Convertir les dates
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except ValueError as e:
        messages.error(request, f"Format de date invalide : {str(e)}")
        log_user_action(
            user,
            "ERREUR - Justification écart (format date)",
            f"Date début: {date_debut} | Date fin: {date_fin} | Erreur: {str(e)}",
            request
        )
        return redirect('inventaire:comptabilisation_quittancements')
    
    # Calculer les totaux et l'écart
    total_quittance = Quittancement.objects.filter(
        poste=poste,
        date_quittancement__range=[date_debut_obj, date_fin_obj]
    ).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    
    total_declare = RecetteJournaliere.objects.filter(
        poste=poste,
        date__range=[date_debut_obj, date_fin_obj]
    ).aggregate(Sum('montant_declare'))['montant_declare__sum'] or Decimal('0')
    
    ecart = total_quittance - total_declare
    ecart_pourcentage = (ecart / total_declare * 100) if total_declare > 0 else 0
    
    # Vérifier si une justification existe déjà
    justification_existante = JustificationEcart.objects.filter(
        poste=poste,
        date_debut=date_debut_obj,
        date_fin=date_fin_obj
    ).select_related('justifie_par').first()
    
    # Log de l'accès à la page
    log_user_action(
        user,
        "Accès page justification écart",
        f"Poste: {poste.nom} | Période: {date_debut_obj} - {date_fin_obj} | "
        f"Écart calculé: {ecart} FCFA ({ecart_pourcentage:.1f}%) | "
        f"Justification existante: {'Oui' if justification_existante else 'Non'}",
        request
    )
    
    # Traitement du formulaire
    if request.method == 'POST':
        justification_texte = request.POST.get('justification', '').strip()
        
        # Validation
        if len(justification_texte) < 20:
            messages.error(
                request,
                "❌ La justification doit contenir au moins 20 caractères."
            )
            log_user_action(
                user,
                "ERREUR - Justification écart (texte trop court)",
                f"Poste: {poste.nom} | Longueur texte: {len(justification_texte)} caractères",
                request
            )
        else:
            try:
                if justification_existante:
                    # Mise à jour d'une justification existante
                    justification_existante.justification = justification_texte
                    justification_existante.justifie_par = request.user
                    justification_existante.date_justification = timezone.now()
                    justification_existante.montant_quittance = total_quittance
                    justification_existante.montant_declare = total_declare
                    justification_existante.ecart = ecart
                    justification_existante.save()
                    
                    log_user_action(
                        user,
                        "Mise à jour justification écart",
                        f"Poste: {poste.nom} | Période: {date_debut_obj} - {date_fin_obj} | "
                        f"Écart: {ecart} FCFA | "
                        f"Extrait: {justification_texte[:50]}...",
                        request
                    )
                    
                    messages.success(
                        request,
                        "✅ Justification mise à jour avec succès."
                    )
                else:
                    # Création d'une nouvelle justification
                    JustificationEcart.objects.create(
                        poste=poste,
                        date_debut=date_debut_obj,
                        date_fin=date_fin_obj,
                        montant_quittance=total_quittance,
                        montant_declare=total_declare,
                        ecart=ecart,
                        justification=justification_texte,
                        justifie_par=request.user
                    )
                    
                    log_user_action(
                        user,
                        "Création justification écart",
                        f"Poste: {poste.nom} | Période: {date_debut_obj} - {date_fin_obj} | "
                        f"Écart: {ecart} FCFA | "
                        f"Extrait: {justification_texte[:50]}...",
                        request
                    )
                    
                    messages.success(
                        request,
                        "✅ Justification enregistrée avec succès."
                    )
                
                return redirect('inventaire:comptabilisation_quittancements')
                
            except Exception as e:
                log_user_action(
                    user,
                    "ERREUR - Enregistrement justification écart",
                    f"Poste: {poste.nom} | Erreur: {str(e)}",
                    request
                )
                messages.error(
                    request,
                    f"❌ Erreur lors de l'enregistrement : {str(e)}"
                )
                logger.error(f"Erreur justification écart: {str(e)}")
    
    # Contexte pour le template (VARIABLES IDENTIQUES pour compatibilité template)
    context = {
        'poste': poste,
        'date_debut': date_debut_obj,
        'date_fin': date_fin_obj,
        'total_quittance': total_quittance,
        'total_declare': total_declare,
        'ecart': ecart,
        'ecart_pourcentage': ecart_pourcentage,
        'justification_existante': justification_existante,
    }
    
    return render(request, 'inventaire/justifier_ecart.html', context)





@login_required
def authentifier_document(request):
    """
    Vue pour authentifier un document.
    MISE À JOUR: Permission peut_authentifier_document requise
    """
    user = request.user
    
    # Vérifier permission avec système granulaire
    if not has_permission(user, 'peut_authentifier_document'):
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Authentification document",
            f"Permission manquante: peut_authentifier_document | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
            request
        )
        messages.error(request, "Vous n'avez pas la permission d'authentifier des documents.")
        return redirect('common:dashboard')
    
    resultat = None
    type_document = None
    
    if request.method == 'POST':
        numero = request.POST.get('numero', '').strip()
        type_recherche = request.POST.get('type_recherche', 'auto')
        
        if not numero:
            messages.error(request, "Veuillez saisir un numéro de document.")
        else:
            # Recherche dans les quittancements
            try:
                quittancement = Quittancement.objects.select_related(
                    'poste', 'saisi_par'
                ).get(numero_quittance=numero)
                
                resultat = {
                    'trouve': True,
                    'type': 'quittance',
                    'document': quittancement,
                    'details': {
                        'numero': quittancement.numero_quittance,
                        'poste': quittancement.poste.nom,
                        'montant': quittancement.montant,
                        'date_quittancement': quittancement.date_quittancement,
                        'periode': quittancement.get_periode_display(),
                        'saisi_par': quittancement.saisi_par.nom_complet if quittancement.saisi_par else 'Non défini',
                        'date_saisie': quittancement.date_saisie,
                    }
                }
                type_document = 'quittance'
                
            except Quittancement.DoesNotExist:
                # Chercher dans les bordereaux
                bordereaux = HistoriqueStock.objects.filter(
                    numero_bordereau=numero
                ).select_related('poste', 'effectue_par')
                
                if bordereaux.exists():
                    resultat = {
                        'trouve': True,
                        'type': 'bordereau',
                        'documents': bordereaux,
                    }
                    type_document = 'bordereau'
                else:
                    resultat = {
                        'trouve': False,
                        'message': f"Aucun document trouvé avec le numéro : {numero}"
                    }
                    messages.warning(request, f"Aucun document trouvé avec le numéro : {numero}")
            
            # Log de la recherche
            log_user_action(
                user,
                "Authentification document",
                f"Numéro recherché: {numero} | Type: {type_recherche} | "
                f"Résultat: {'Trouvé' if resultat and resultat.get('trouve') else 'Non trouvé'} | "
                f"Type document: {type_document or 'N/A'}",
                request
            )
    
    context = {
        'resultat': resultat,
        'type_document': type_document,
        'title': 'Authentification de Documents'
    }
    
    return render(request, 'inventaire/authentifier_document.html', context)

 



@login_required
def detail_quittancements_periode(request, poste_id, date_debut, date_fin):
    """
    Vue pour afficher les détails des quittancements d'un poste sur une période
    
    MISE À JOUR - Permissions granulaires:
    - Remplace: request.user.is_admin et request.user.poste_affectation.id != poste.id
    - Par: user_has_acces_tous_postes(user) et check_poste_access(user, poste)
    - Log détaillé de chaque action utilisateur
    
    LOGIQUE DE CALCUL :
    - Pour les quittancements journaliers : comparaison directe jour par jour
    - Pour les décades : vérification que tous les jours ont des recettes, 
      puis somme des recettes vs montant de la décade
    """
    from datetime import datetime, timedelta
    from decimal import Decimal
    from django.db.models import Sum
    
    user = request.user
    poste = get_object_or_404(Poste, id=poste_id)
    
    # =========================================
    # VÉRIFICATION PERMISSIONS GRANULAIRES
    # (Remplace: if not request.user.is_admin: if not request.user.poste_affectation...)
    # =========================================
    
    # Vérifier l'accès au poste avec le nouveau système
    if not user_has_acces_tous_postes(user):
        if not check_poste_access(user, poste):
            log_user_action(
                user,
                "ACCÈS REFUSÉ - Détails quittancements",
                f"Poste non autorisé | "
                f"Poste demandé: {poste.nom} (ID: {poste_id}) | "
                f"Poste affectation: {getattr(user.poste_affectation, 'nom', 'Aucun')} | "
                f"Habilitation: {getattr(user, 'habilitation', 'N/A')}",
                request
            )
            messages.error(request, "Vous n'avez pas accès aux données de ce poste.")
            return redirect('inventaire:comptabilisation_quittancements')
    
    # Conversion des dates
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except ValueError as e:
        log_user_action(
            user,
            "ERREUR - Détails quittancements (format date)",
            f"Date début: {date_debut} | Date fin: {date_fin} | Erreur: {str(e)}",
            request
        )
        messages.error(request, f"Format de date invalide : {str(e)}")
        return redirect('inventaire:comptabilisation_quittancements')
    
    # Récupérer les quittancements journaliers
    quittancements_journaliers = Quittancement.objects.filter(
        poste=poste,
        type_declaration='journaliere',
        date_recette__range=[date_debut_obj, date_fin_obj]
    ).select_related('poste', 'saisi_par').order_by('date_recette')
    
    # Récupérer les quittancements de décade qui chevauchent la période
    from django.db import models
    quittancements_decades = Quittancement.objects.filter(
        poste=poste,
        type_declaration='decade'
    ).filter(
        models.Q(date_debut_decade__lte=date_fin_obj) & 
        models.Q(date_fin_decade__gte=date_debut_obj)
    ).select_related('poste', 'saisi_par').order_by('date_debut_decade')
    
    # Récupérer toutes les recettes de la période
    recettes = RecetteJournaliere.objects.filter(
        poste=poste,
        date__range=[date_debut_obj, date_fin_obj]
    ).order_by('date')
    
    # === LOGIQUE DE CALCUL ===
    
    # Traiter les décades
    decades_details = []
    for q_decade in quittancements_decades:
        # Récupérer toutes les dates de la décade
        dates_decade = []
        current_date = q_decade.date_debut_decade
        while current_date <= q_decade.date_fin_decade:
            # Ne prendre que les dates dans notre période d'analyse
            if date_debut_obj <= current_date <= date_fin_obj:
                dates_decade.append(current_date)
            current_date += timedelta(days=1)
        
        # Récupérer les recettes de cette décade
        recettes_decade = recettes.filter(date__in=dates_decade)
        dates_avec_recettes = set(recettes_decade.values_list('date', flat=True))
        
        # Identifier les dates manquantes
        dates_manquantes = [d for d in dates_decade if d not in dates_avec_recettes]
        
        if dates_manquantes:
            # Décade incomplète - impossible de calculer l'écart
            decade_detail = {
                'quittance': q_decade,
                'dates': dates_decade,
                'dates_manquantes': dates_manquantes,
                'somme_recettes': Decimal('0'),
                'statut': 'incomplet',
                'ecart': None,
                'message': f"{len(dates_manquantes)} jour(s) sans recette"
            }
        else:
            # Calculer la somme des recettes journalières
            somme_recettes = recettes_decade.aggregate(
                Sum('montant_declare')
            )['montant_declare__sum'] or Decimal('0')
            
            # L'écart est la différence entre le montant quittancé et la somme des recettes
            ecart_decade = q_decade.montant - somme_recettes
            
            decade_detail = {
                'quittance': q_decade,
                'dates': dates_decade,
                'dates_manquantes': [],
                'somme_recettes': somme_recettes,
                'statut': 'ok' if abs(ecart_decade) < 1 else 'ecart',
                'ecart': ecart_decade
            }
        
        decades_details.append(decade_detail)
    
    # Créer la comparaison journalière détaillée
    comparaison_journaliere = []
    total_quittance = Decimal('0')
    total_declare = Decimal('0')
    
    # Parcourir toutes les dates de la période
    current_date = date_debut_obj
    while current_date <= date_fin_obj:
        # Récupérer la recette du jour
        recette_jour = recettes.filter(date=current_date).first()
        montant_declare = recette_jour.montant_declare if recette_jour else Decimal('0')
        
        # Récupérer les quittancements journaliers de ce jour
        quittances_jour = []
        montant_quittance_jour = Decimal('0')
        
        for q in quittancements_journaliers:
            if q.date_recette == current_date:
                quittances_jour.append(q)
                montant_quittance_jour += q.montant
        
        # Vérifier si ce jour fait partie d'une décade
        decade_couvrant = None
        for decade_detail in decades_details:
            if current_date in decade_detail['dates']:
                decade_couvrant = decade_detail
                break
        
        # Calculer l'écart du jour (seulement pour les quittancements journaliers)
        if decade_couvrant:
            # Pour les jours en décade, l'écart est calculé au niveau de la décade
            ecart_jour = None
            statut = 'decade'
            statut_label = 'Partie de décade'
            statut_class = 'info'
        else:
            # Pour les jours avec quittancement journalier
            ecart_jour = montant_quittance_jour - montant_declare
            
            if not recette_jour and montant_quittance_jour > 0:
                statut = 'manquant'
                statut_label = 'Recette manquante'
                statut_class = 'warning'
            elif abs(ecart_jour) < 1:
                statut = 'ok'
                statut_label = 'Conforme'
                statut_class = 'success'
            else:
                statut = 'ecart'
                statut_label = 'Écart détecté'
                statut_class = 'danger'
        
        comparaison_journaliere.append({
            'date': current_date,
            'recette': recette_jour,
            'montant_declare': montant_declare,
            'montant_quittance': montant_quittance_jour,
            'ecart': ecart_jour,
            'quittances_jour': quittances_jour,
            'decade_couvrant': decade_couvrant,
            'statut': statut,
            'statut_label': statut_label,
            'statut_class': statut_class,
        })
        
        # Ajouter aux totaux
        total_declare += montant_declare
        if not decade_couvrant:  # Ne pas compter deux fois pour les décades
            total_quittance += montant_quittance_jour
        
        current_date += timedelta(days=1)
    
    # Ajouter les montants des décades au total
    for decade_detail in decades_details:
        if decade_detail['statut'] != 'incomplet':
            # Ne compter que la partie de la décade dans notre période
            # Prorata si la décade dépasse notre période
            q_decade = decade_detail['quittance']
            if q_decade.date_debut_decade >= date_debut_obj and q_decade.date_fin_decade <= date_fin_obj:
                # Décade entièrement dans la période
                total_quittance += q_decade.montant
            else:
                # Calculer le prorata
                debut_effectif = max(q_decade.date_debut_decade, date_debut_obj)
                fin_effective = min(q_decade.date_fin_decade, date_fin_obj)
                jours_dans_periode = (fin_effective - debut_effectif).days + 1
                jours_total_decade = (q_decade.date_fin_decade - q_decade.date_debut_decade).days + 1
                montant_prorata = (q_decade.montant * jours_dans_periode) / jours_total_decade
                total_quittance += montant_prorata
    
    # Calcul de l'écart global
    ecart = total_quittance - total_declare
    
    # Vérifier les justifications existantes
    justifications = JustificationEcart.objects.filter(
        poste=poste,
        date_debut=date_debut_obj,
        date_fin=date_fin_obj
    ).select_related('justifie_par')
    
    # Log détaillé de la consultation
    log_user_action(
        user,
        "Consultation détails quittancements",
        f"Poste: {poste.nom} | Période: {date_debut_obj} - {date_fin_obj} | "
        f"Quittancements journaliers: {quittancements_journaliers.count()} | "
        f"Décades: {quittancements_decades.count()} | "
        f"Écart global: {ecart} FCFA",
        request
    )
    
    # Contexte pour le template (VARIABLES IDENTIQUES pour compatibilité template)
    context = {
        'poste': poste,
        'date_debut': date_debut_obj,
        'date_fin': date_fin_obj,
        'quittancements_journaliers': quittancements_journaliers,
        'quittancements_decades': quittancements_decades,
        'decades_details': decades_details,
        'recettes': recettes,
        'comparaison_journaliere': comparaison_journaliere,
        'total_quittance': total_quittance,
        'total_declare': total_declare,
        'ecart': ecart,
        'justifications': justifications,
        'quittancements': list(quittancements_journaliers) + list(quittancements_decades),
    }
    
    return render(request, 'inventaire/detail_quittancements.html', context)


@login_required
def export_quittancements(request):
    """Vue pour exporter les quittancements - FONCTION D'EXPORT"""
    
    format_export = request.GET.get('format', 'excel')
    
    # Récupérer les mêmes filtres que la comptabilisation
    periode = request.GET.get('periode', 'mois')
    poste_id = request.GET.get('poste')
    annee = int(request.GET.get('annee', timezone.now().year))
    mois = request.GET.get('mois')
    
    # Déterminer les dates
    if periode == 'mois' and mois:
        mois_int = int(mois)
        date_debut = date(annee, mois_int, 1)
        date_fin = date(annee, mois_int, calendar.monthrange(annee, mois_int)[1])
    elif periode == 'trimestre':
        trimestre = int(request.GET.get('trimestre', 1))
        date_debut = date(annee, (trimestre - 1) * 3 + 1, 1)
        date_fin = date(annee, trimestre * 3, calendar.monthrange(annee, trimestre * 3)[1])
    elif periode == 'semestre':
        semestre = int(request.GET.get('semestre', 1))
        date_debut = date(annee, 1 if semestre == 1 else 7, 1)
        date_fin = date(annee, 6 if semestre == 1 else 12, 
                       calendar.monthrange(annee, 6 if semestre == 1 else 12)[1])
    else:
        date_debut = date(annee, 1, 1)
        date_fin = date(annee, 12, 31)
    
    # Filtrer les quittancements
    quittancements = Quittancement.objects.filter(
        date_quittancement__range=[date_debut, date_fin]
    )
    
    if poste_id:
        quittancements = quittancements.filter(poste_id=poste_id)
    
    # Selon les permissions
    if not request.user.is_admin:
        if request.user.poste_affectation:
            quittancements = quittancements.filter(poste=request.user.poste_affectation)
    
    if format_export == 'excel':
        import openpyxl
        from django.http import HttpResponse
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=quittancements_{date_debut}_{date_fin}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Quittancements'
        
        # En-têtes
        headers = ['Numéro', 'Poste', 'Date', 'Montant', 'Type', 'Période', 'Saisi par', 'Date saisie']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Données
        for row, q in enumerate(quittancements, 2):
            ws.cell(row=row, column=1, value=q.numero_quittance)
            ws.cell(row=row, column=2, value=q.poste.nom)
            ws.cell(row=row, column=3, value=q.date_quittancement)
            ws.cell(row=row, column=4, value=float(q.montant))
            ws.cell(row=row, column=5, value=q.get_type_declaration_display())
            ws.cell(row=row, column=6, value=q.get_periode_display())
            ws.cell(row=row, column=7, value=q.saisi_par.nom_complet if q.saisi_par else '')
            ws.cell(row=row, column=8, value=q.date_saisie)
        
        wb.save(response)
        return response
    
    else:  # PDF
        messages.info(request, "Export PDF en cours de développement")
        return redirect('inventaire:comptabilisation_quittancements')


@login_required
def ajouter_image_quittancement(request, quittancement_id):
    """
    Vue pour ajouter/modifier l'image d'un quittancement existant
    
    MISE À JOUR - Permissions granulaires:
    - Remplace: request.user.is_admin
    - Par: is_admin_user(user)
    - Log détaillé de chaque action utilisateur
    
    UTILISATION :
    - Après avoir créé un quittancement sans image (étapes 1-2-3)
    - L'utilisateur peut cliquer sur "Ajouter une image" dans la liste
    - Upload simple et direct sans workflow complexe
    - Permet aussi de modifier/remplacer une image existante
    
    AVANTAGES :
    - Contourne le problème de perte d'image dans les redirections
    - Interface simple et intuitive
    - Modifiable à tout moment
    """
    user = request.user
    
    # Récupérer le quittancement
    quittancement = get_object_or_404(Quittancement, id=quittancement_id)
    
    # =========================================
    # VÉRIFICATION PERMISSIONS GRANULAIRES
    # (Remplace: request.user.is_admin or request.user == quittancement.saisi_par)
    # =========================================
    
    # Seuls l'admin ou la personne qui a créé le quittancement peuvent modifier
    can_modify = is_admin_user(user) or user == quittancement.saisi_par
    
    if not can_modify:
        log_user_action(
            user,
            "ACCÈS REFUSÉ - Modification image quittancement",
            f"N°{quittancement.numero_quittance} | "
            f"Saisi par: {quittancement.saisi_par.nom_complet if quittancement.saisi_par else 'N/A'} | "
            f"Habilitation: {getattr(user, 'habilitation', 'N/A')} | "
            f"IP: {request.META.get('REMOTE_ADDR')}",
            request
        )
        messages.error(
            request, 
            "❌ Vous n'avez pas les droits pour modifier ce quittancement"
        )
        return redirect('inventaire:liste_quittancements')
    
    if request.method == 'POST':
        # Vérifier qu'une image est uploadée
        if 'image_quittance' not in request.FILES:
            messages.error(request, "❌ Aucune image sélectionnée")
            log_user_action(
                user,
                "ERREUR - Ajout image quittancement (aucun fichier)",
                f"N°{quittancement.numero_quittance}",
                request
            )
            return redirect('inventaire:ajouter_image_quittancement', quittancement_id=quittancement_id)
        
        image_file = request.FILES['image_quittance']
        
        # Validation de la taille (5 MB max)
        max_size = 5 * 1024 * 1024  # 5 MB
        if image_file.size > max_size:
            log_user_action(
                user,
                "ERREUR - Ajout image quittancement (fichier trop volumineux)",
                f"N°{quittancement.numero_quittance} | Taille: {image_file.size / (1024*1024):.1f} MB",
                request
            )
            messages.error(
                request, 
                f"❌ Fichier trop volumineux ({image_file.size / (1024*1024):.1f} MB). Maximum : 5 MB"
            )
            return redirect('inventaire:ajouter_image_quittancement', quittancement_id=quittancement_id)
        
        # Validation du type de fichier
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'application/pdf']
        if image_file.content_type not in allowed_types:
            log_user_action(
                user,
                "ERREUR - Ajout image quittancement (type non autorisé)",
                f"N°{quittancement.numero_quittance} | Type: {image_file.content_type}",
                request
            )
            messages.error(
                request,
                f"❌ Type de fichier non autorisé. Formats acceptés : JPG, PNG, GIF, PDF"
            )
            return redirect('inventaire:ajouter_image_quittancement', quittancement_id=quittancement_id)
        
        try:
            # Supprimer l'ancienne image si elle existe
            old_path = None
            if quittancement.image_quittance:
                # Sauvegarder l'ancien chemin pour le log
                old_path = quittancement.image_quittance.name
                
                # Supprimer le fichier physique
                quittancement.image_quittance.delete(save=False)
                
                logger.info(f"Ancienne image supprimée : {old_path}")
            
            # Ajouter la nouvelle image
            quittancement.image_quittance = image_file
            quittancement.save()
            
            # Log détaillé de l'action
            log_user_action(
                user,
                "Ajout/Modification image quittancement",
                f"N°{quittancement.numero_quittance} | "
                f"Poste: {quittancement.poste.nom} | "
                f"Nouveau fichier: {image_file.name} ({image_file.size / 1024:.1f} KB) | "
                f"Ancien fichier: {old_path or 'Aucun'}",
                request
            )
            
            messages.success(
                request, 
                f"✅ Image ajoutée avec succès au quittancement {quittancement.numero_quittance}"
            )
            
            return redirect('inventaire:liste_quittancements')
            
        except Exception as e:
            log_user_action(
                user,
                "ERREUR - Enregistrement image quittancement",
                f"N°{quittancement.numero_quittance} | Erreur: {str(e)}",
                request
            )
            messages.error(request, f"❌ Erreur lors de l'enregistrement : {str(e)}")
            logger.error(f"Erreur ajout image quittancement {quittancement_id}: {str(e)}", exc_info=True)
            return redirect('inventaire:ajouter_image_quittancement', quittancement_id=quittancement_id)
    
    # GET : Afficher le formulaire d'upload
    # Log de l'accès à la page
    log_user_action(
        user,
        "Accès page ajout image quittancement",
        f"N°{quittancement.numero_quittance} | Poste: {quittancement.poste.nom} | "
        f"Image existante: {'Oui' if quittancement.image_quittance else 'Non'}",
        request
    )
    
    # Contexte pour le template (VARIABLES IDENTIQUES pour compatibilité template)
    context = {
        'quittancement': quittancement,
        'poste': quittancement.poste,
        'has_image': bool(quittancement.image_quittance),
    }
    
    return render(request, 'inventaire/ajouter_image_quittancement.html', context)
