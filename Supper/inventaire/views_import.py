# ===================================================================
# inventaire/views_import.py - Import Excel des recettes SUPPER
# VERSION MISE À JOUR - Intégration des permissions granulaires
#
# Permission requise: peut_importer_recettes_peage
# Habilitations concernées: admin_principal, coord_psrr, serv_info, serv_emission
# ===================================================================

import pandas as pd
from datetime import datetime
from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
import re
from difflib import SequenceMatcher

from accounts.models import Poste
from inventaire.models import RecetteJournaliere

# ===================================================================
# IMPORTS DES MODULES DE PERMISSIONS ET UTILITAIRES CENTRALISÉS
# ===================================================================

from common.utils import log_user_action

from common.permissions import (
    has_permission,
    is_admin_user,
    log_acces_refuse,
)

from common.decorators import (
    permission_required_granular,
    import_recettes_peage_required,
)

import logging

logger = logging.getLogger('supper')


# ===================================================================
# DÉCORATEUR POUR L'IMPORT DES RECETTES
# Permission: peut_importer_recettes_peage
# ===================================================================

def import_recettes_required(view_func):
    """
    Décorateur pour vérifier la permission d'importer des recettes.
    Permission: peut_importer_recettes_peage
    """
    from functools import wraps
    
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        
        if not has_permission(user, 'peut_importer_recettes_peage'):
            log_user_action(
                user,
                "Import recettes - permission refusée",
                f"Permission peut_importer_recettes_peage manquante - Vue: {view_func.__name__}",
                request
            )
            log_acces_refuse(user, view_func.__name__, "Permission import recettes manquante")
            messages.error(request, "Vous n'avez pas la permission d'importer des recettes.")
            return redirect('inventaire:liste_recettes')
        
        return view_func(request, *args, **kwargs)
    
    return wrapper


# ===================================================================
# VUE D'IMPORT DES RECETTES EXCEL
# ===================================================================

@import_recettes_required
def import_recettes_excel(request):
    """
    Vue pour importer des recettes depuis un fichier Excel au format matriciel
    (postes en lignes, dates en colonnes)
    
    Permission: peut_importer_recettes_peage
    """
    user = request.user
    
    if request.method == 'POST':
        fichier = request.FILES.get('fichier_excel')
        action_doublon = request.POST.get('action_doublon', 'sauter')
        
        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier Excel")
            return redirect('inventaire:import_recettes')
        
        if not fichier.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Le fichier doit être au format Excel (.xlsx ou .xls)")
            return redirect('inventaire:import_recettes')
        
        # Log du début de l'import
        log_user_action(
            user,
            "Import recettes Excel - début",
            f"Fichier: {fichier.name} | Taille: {fichier.size} bytes | Action doublon: {action_doublon}",
            request
        )
        
        try:
            # Lire le fichier Excel
            df = pd.read_excel(fichier)
            
            # Traiter l'import
            resultats = traiter_import_recettes_matriciel(
                df, action_doublon, user
            )
            
            # Afficher les résultats
            if resultats['success']:
                messages.success(
                    request,
                    f"Import terminé avec succès ! "
                    f"{resultats['nb_crees']} recettes créées, "
                    f"{resultats['nb_modifiees']} modifiées, "
                    f"{resultats['nb_sautes']} sautées, "
                    f"{resultats['nb_erreurs']} erreurs"
                )
                
                # Log détaillé de l'action
                log_user_action(
                    user,
                    "Import recettes Excel - succès",
                    f"Fichier: {fichier.name} | "
                    f"Créées: {resultats['nb_crees']} | "
                    f"Modifiées: {resultats['nb_modifiees']} | "
                    f"Sautées: {resultats['nb_sautes']} | "
                    f"Erreurs: {resultats['nb_erreurs']} | "
                    f"Postes non trouvés: {len(resultats.get('postes_non_trouves', []))}",
                    request
                )
                
                logger.info(
                    f"Import recettes par {user.username}: "
                    f"{resultats['nb_crees']} créées, "
                    f"{resultats['nb_modifiees']} modifiées"
                )
            else:
                messages.error(request, f"Erreur lors de l'import: {resultats['erreur']}")
                
                log_user_action(
                    user,
                    "Import recettes Excel - échec",
                    f"Fichier: {fichier.name} | Erreur: {resultats['erreur']}",
                    request
                )
                
                logger.error(f"Import recettes échoué pour {user.username}: {resultats['erreur']}")
            
            context = {
                'resultats': resultats,
                'is_admin': is_admin_user(user),
            }
            
            return render(request, 'inventaire/import_recettes_resultats.html', context)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier: {str(e)}")
            
            log_user_action(
                user,
                "Import recettes Excel - erreur lecture",
                f"Fichier: {fichier.name} | Erreur: {str(e)}",
                request
            )
            
            logger.error(f"Erreur lecture fichier Excel par {user.username}: {str(e)}", exc_info=True)
            
            return redirect('inventaire:import_recettes')
    
    # GET request - afficher le formulaire
    context = {
        'is_admin': is_admin_user(user),
        'title': 'Import des recettes',
    }
    return render(request, 'inventaire/import_recettes_form.html', context)


def traiter_import_recettes_matriciel(df, action_doublon, user):
    """
    Traite l'import des recettes depuis le format matriciel Excel.
    
    Args:
        df: DataFrame pandas contenant les données
        action_doublon: 'sauter' ou 'ecraser'
        user: L'utilisateur effectuant l'import
    
    Returns:
        dict: Résultats de l'import avec compteurs et détails
    """
    nb_crees = 0
    nb_modifiees = 0
    nb_sautes = 0
    nb_erreurs = 0
    erreurs_detail = []
    postes_non_trouves = {}
    dates_invalides = set()
    
    try:
        noms_colonnes = df.columns.tolist()
        colonne_postes = noms_colonnes[0]
        colonnes_dates = noms_colonnes[1:]
        
        # Créer un mapping des postes avec PLUSIEURS clés
        postes_db = {}
        for poste in Poste.objects.filter(is_active=True):
            # Clé principale : nom normalisé
            nom_normalise = normaliser_nom_poste(poste.nom)
            postes_db[nom_normalise] = poste
            
            # Clé alternative : sans parenthèses
            nom_sans_parentheses = re.sub(r'\([^)]*\)', '', nom_normalise).strip()
            nom_sans_parentheses = ' '.join(nom_sans_parentheses.split())
            if nom_sans_parentheses and nom_sans_parentheses != nom_normalise:
                postes_db[nom_sans_parentheses] = poste
            
            # Clé alternative : code normalisé
            if poste.code:
                code_normalise = normaliser_nom_poste(poste.code)
                postes_db[code_normalise] = poste
        
        logger.debug(f"Mapping postes créé avec {len(postes_db)} entrées")
        
        # Parser les dates
        mapping_dates = {}
        for col in colonnes_dates:
            try:
                if isinstance(col, datetime):
                    date_obj = col.date()
                elif isinstance(col, str):
                    date_obj = datetime.strptime(col.strip(), '%d/%m/%y').date()
                else:
                    continue
                mapping_dates[col] = date_obj
            except:
                dates_invalides.add(str(col))
                continue
        
        if not mapping_dates:
            return {
                'success': False,
                'erreur': "Aucune date valide trouvée dans les en-têtes",
                'nb_crees': 0,
                'nb_modifiees': 0,
                'nb_sautes': 0,
                'nb_erreurs': 0
            }
        
        logger.debug(f"Dates valides parsées: {len(mapping_dates)}")
        
        # Parcourir les lignes
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    nom_poste_excel = str(row[colonne_postes]).strip()
                    
                    if not nom_poste_excel or nom_poste_excel == 'nan':
                        continue
                    
                    # Utiliser la recherche flexible
                    poste = trouver_poste_flexible(nom_poste_excel, postes_db)
                    
                    if not poste:
                        # Stocker le nom original ET normalisé pour debug
                        nom_norm = normaliser_nom_poste(nom_poste_excel)
                        postes_non_trouves[nom_poste_excel] = nom_norm
                        nb_sautes += len(mapping_dates)
                        continue
                    
                    # Parcourir les dates
                    for col_date, date_obj in mapping_dates.items():
                        try:
                            montant_raw = row[col_date]
                            
                            if pd.isna(montant_raw):
                                continue
                            
                            # Nettoyer et convertir le montant
                            montant_str = str(montant_raw).replace(',', '.').replace(' ', '')
                            montant = Decimal(montant_str)
                            
                            if montant <= 0:
                                continue
                            
                            # Vérifier si une recette existe déjà
                            recette_existante = RecetteJournaliere.objects.filter(
                                poste=poste,
                                date=date_obj
                            ).first()
                            
                            if recette_existante:
                                if action_doublon == 'ecraser':
                                    recette_existante.montant_declare = montant
                                    recette_existante.chef_poste = user
                                    recette_existante.observations = f"Importé Excel {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                                    recette_existante.save()
                                    nb_modifiees += 1
                                else:
                                    nb_sautes += 1
                            else:
                                RecetteJournaliere.objects.create(
                                    poste=poste,
                                    date=date_obj,
                                    montant_declare=montant,
                                    chef_poste=user,
                                    observations=f"Importé Excel {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                                )
                                nb_crees += 1
                                
                        except Exception as e:
                            erreurs_detail.append(f"L{index + 2}, {col_date}: {str(e)}")
                            nb_erreurs += 1
                            
                except Exception as e:
                    erreurs_detail.append(f"Ligne {index + 2}: {str(e)}")
                    nb_erreurs += 1
        
        # Préparer la liste des postes non trouvés avec infos debug
        postes_non_trouves_liste = [
            f"{original} → normalisé: {normalise}" 
            for original, normalise in list(postes_non_trouves.items())[:30]
        ]
        
        return {
            'success': True,
            'nb_crees': nb_crees,
            'nb_modifiees': nb_modifiees,
            'nb_sautes': nb_sautes,
            'nb_erreurs': nb_erreurs,
            'erreurs_detail': erreurs_detail[:50],
            'postes_non_trouves': postes_non_trouves_liste,
            'dates_invalides': list(dates_invalides)[:20]
        }
        
    except Exception as e:
        logger.error(f"Erreur traitement import: {str(e)}", exc_info=True)
        return {
            'success': False,
            'erreur': str(e),
            'nb_crees': 0,
            'nb_modifiees': 0,
            'nb_sautes': 0,
            'nb_erreurs': 0
        }


def normaliser_nom_poste(nom):
    """
    Normalise un nom de poste pour la comparaison robuste.
    
    Transformations appliquées:
    - Conversion en majuscules
    - Suppression des accents
    - Normalisation des espaces
    - Remplacement tirets/underscores par espaces
    - Normalisation O/0
    """
    import unicodedata
    import re
    
    # Convertir en string et strip
    nom = str(nom).strip().upper()
    
    # Supprimer les accents
    nom = ''.join(
        c for c in unicodedata.normalize('NFD', nom)
        if unicodedata.category(c) != 'Mn'
    )
    
    # Remplacer tous les espaces multiples par un seul
    nom = re.sub(r'\s+', ' ', nom)
    
    # Remplacer tirets et underscores par espaces
    nom = nom.replace('-', ' ').replace('_', ' ')
    
    # Supprimer tous les caractères spéciaux SAUF parenthèses et espaces
    nom = re.sub(r'[^\w\s\(\)]', '', nom)
    
    # Re-normaliser les espaces après nettoyage
    nom = ' '.join(nom.split())
    
    # Normaliser le zéro et le O (cas MEKOTT0)
    nom = nom.replace('0', 'O')
    
    return nom


def trouver_poste_flexible(nom_excel, postes_db):
    """
    Recherche flexible d'un poste avec plusieurs stratégies.
    
    Stratégies de recherche (dans l'ordre):
    1. Match exact normalisé
    2. Match sans parenthèses
    3. Match par similarité (>90%)
    
    Args:
        nom_excel: Nom du poste depuis Excel
        postes_db: Dict des postes normalisés
    
    Returns:
        Poste trouvé ou None
    """
    # Stratégie 1 : Match exact normalisé
    nom_normalise = normaliser_nom_poste(nom_excel)
    if nom_normalise in postes_db:
        return postes_db[nom_normalise]
    
    # Stratégie 2 : Sans parenthèses
    nom_sans_parentheses = re.sub(r'\([^)]*\)', '', nom_normalise).strip()
    nom_sans_parentheses = ' '.join(nom_sans_parentheses.split())
    
    for cle, poste in postes_db.items():
        cle_sans_parentheses = re.sub(r'\([^)]*\)', '', cle).strip()
        cle_sans_parentheses = ' '.join(cle_sans_parentheses.split())
        
        if nom_sans_parentheses == cle_sans_parentheses:
            return poste
    
    # Stratégie 3 : Similarité avec difflib (>90%)
    meilleur_match = None
    meilleur_score = 0
    
    for cle, poste in postes_db.items():
        score = SequenceMatcher(None, nom_normalise, cle).ratio()
        if score > meilleur_score and score >= 0.90:
            meilleur_score = score
            meilleur_match = poste
    
    return meilleur_match


# ===================================================================
# TÉLÉCHARGEMENT DU MODÈLE EXCEL
# ===================================================================

@import_recettes_required
def telecharger_modele_excel(request):
    """
    Génère et télécharge un modèle Excel au format matriciel.
    
    Permission: peut_importer_recettes_peage
    """
    user = request.user
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from datetime import timedelta
    except ImportError:
        messages.error(request, "Module openpyxl non installé.")
        return redirect('inventaire:import_recettes')
    
    log_user_action(
        user,
        "Téléchargement modèle Excel recettes",
        "Génération du modèle Excel pour import des recettes",
        request
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recettes Avril 2025"
    
    # Style pour l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # A1: Titre
    ws['A1'] = "POSTE DE PEAGE"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    ws['A1'].border = thin_border
    
    # Générer les dates pour avril 2025 (colonnes B1 à AF1)
    date_debut = datetime(2025, 4, 1).date()
    for jour in range(30):  # Avril a 30 jours
        date_actuelle = date_debut + timedelta(days=jour)
        col_index = jour + 2  # Commence à la colonne B (index 2)
        
        cell = ws.cell(row=1, column=col_index)
        cell.value = date_actuelle.strftime('%d/%m/%y')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Récupérer quelques postes exemples (postes de péage uniquement)
    postes = Poste.objects.filter(is_active=True, type='peage').order_by('nom')[:10]
    
    # Remplir les noms de postes (colonne A, à partir de ligne 2)
    for row_num, poste in enumerate(postes, 2):
        cell = ws.cell(row=row_num, column=1)
        cell.value = poste.nom
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajouter quelques valeurs exemples
        for col in range(2, 10):  # Quelques jours seulement
            cell = ws.cell(row=row_num, column=col)
            cell.value = 50000 + (row_num * 1000)
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 35
    for col in range(2, 33):  # B à AF
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # Figer la première ligne et colonne
    ws.freeze_panes = 'B2'
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=modele_recettes_matriciel.xlsx'
    
    logger.info(f"Modèle Excel téléchargé par {user.username}")
    
    return response