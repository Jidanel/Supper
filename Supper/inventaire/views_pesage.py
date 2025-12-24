# ===================================================================
# inventaire/views_pesage.py - Vues pour le module Pesage SUPPER
# VERSION MISE À JOUR - Intégration des permissions granulaires
# 
# Gestion des permissions par PERMISSION GRANULAIRE (pas par rôle):
# - peut_saisir_amende: Saisie d'amendes
# - peut_saisir_pesee_jour: Saisie des pesées journalières
# - peut_valider_paiement_amende: Validation des paiements
# - peut_saisir_quittance_pesage: Saisie de quittancements
# - peut_lister_amendes: Consultation des amendes
# - peut_voir_historique_pesees: Historique des pesées
# - peut_voir_recettes_pesage: Recettes pesage
# - peut_voir_stats_pesage: Statistiques pesage
# - peut_comptabiliser_quittances_pesage: Comptabilisation
# - peut_voir_liste_quittancements_pesage: Liste quittancements
#
# Accès aux postes:
# - Habilitations avec acces_tous_postes: peuvent sélectionner n'importe quelle station
# - Habilitations pesage (chef_station, regisseur, chef_equipe): accès à leur station uniquement
# ===================================================================

import calendar
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate
from django.views.decorators.http import require_POST, require_GET
from django.core.exceptions import ValidationError
from functools import wraps
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import json
import logging
import os
import pytz

# ===================================================================
# IMPORTS DES MODULES DE PERMISSIONS ET UTILITAIRES CENTRALISÉS
# ===================================================================

from inventaire.utils_pesage import normalize_immatriculation
from common.utils import log_user_action

# Imports depuis le module de permissions centralisé
from common.permissions import (
    # Fonctions de classification utilisateur
    is_admin_user,
    is_service_central,
    is_cisop_pesage,
    is_operationnel_pesage,
    is_chef_station_pesage,
    is_regisseur_pesage,
    is_chef_equipe_pesage,
    
    # Fonctions d'accès aux postes
    user_has_acces_tous_postes,
    get_postes_accessibles,
    get_postes_pesage_accessibles,
    check_poste_access,
    get_poste_from_request,
    
    # Fonctions de vérification de permissions
    has_permission,
    has_any_permission,
    
    # Logging
    log_acces_refuse,
    
    # Constantes
    HABILITATIONS_OPERATIONNELS_PESAGE,
    PERMISSIONS_PESAGE,
)

# Imports depuis le module de décorateurs centralisé
from common.decorators import (
    permission_required_granular,
    saisie_amende_required,
    saisie_pesee_required,
    validation_paiement_amende_required,
    saisie_quittance_pesage_required,
    liste_amendes_required,
    historique_pesees_required,
    stats_pesage_required,
    comptabiliser_quittances_pesage_required,
    api_permission_required,
    operationnel_pesage_required,
)

from .models_pesage import *
from .forms_pesage import *
from accounts.models import Poste

logger = logging.getLogger('supper')


# ===================================================================
# CONSTANTES POUR LA LOGIQUE MÉTIER 9h-9h (CAMEROUN GMT+1)
# ===================================================================

CAMEROUN_TZ = pytz.timezone('Africa/Douala')
HEURE_DEBUT_JOURNEE = time(9, 0, 0)
HEURE_FIN_JOURNEE = time(8, 59, 59)


# ===================================================================
# FONCTIONS UTILITAIRES D'ACCÈS AUX POSTES (PESAGE)
# ===================================================================

def get_user_station_pesage(user):
    """
    Retourne la station de pesage de l'utilisateur.
    
    Returns:
        - None si l'utilisateur a accès à tous les postes (admin, services centraux, cisop_pesage)
        - Instance de Poste si utilisateur affecté à une station pesage
        - False si utilisateur non autorisé pour le pesage
    
    Utilise les permissions granulaires pour déterminer l'accès.
    """
    if not user or not user.is_authenticated:
        return False
    
    # Vérifier si l'utilisateur a accès à tous les postes (multi-postes)
    if user_has_acces_tous_postes(user):
        logger.debug(f"[PESAGE] Utilisateur {user.username} a accès à toutes les stations (multi-postes)")
        return None
    
    # Vérifier si l'utilisateur a au moins une permission pesage
    if not has_any_permission(user, PERMISSIONS_PESAGE):
        logger.warning(f"[PESAGE] Accès refusé: {user.username} n'a aucune permission pesage")
        return False
    
    # Vérifier le poste d'affectation
    poste = getattr(user, 'poste_affectation', None)
    if poste and poste.type == 'pesage' and poste.is_active:
        logger.debug(f"[PESAGE] Utilisateur {user.username} affecté à la station {poste.nom}")
        return poste
    
    logger.warning(f"[PESAGE] Utilisateur {user.username} sans station de pesage valide")
    return False


def get_stations_pesage_accessibles(user):
    """
    Retourne le queryset des stations de pesage accessibles pour l'utilisateur.
    Utilise la fonction centralisée get_postes_accessibles avec filtre type='pesage'.
    """
    return get_postes_accessibles(user, type_poste='pesage')


def get_station_context(request, user):
    """
    Récupère la station pour le contexte de la vue.
    
    Logic:
    1. Si l'utilisateur n'a PAS accès à tous les postes → utiliser son poste d'affectation
    2. Si l'utilisateur A accès à tous les postes → permettre la sélection via GET/POST/session
    
    Returns:
        tuple (station, redirect_response)
        - (station, None) si station trouvée
        - (None, redirect) si redirection nécessaire
        - (None, None) si admin sans station sélectionnée
    """
    # Cas 1: Utilisateur avec accès limité à son poste
    if not user_has_acces_tous_postes(user):
        station = get_user_station_pesage(user)
        
        if station is False:
            logger.warning(f"[PESAGE] {user.username} tente d'accéder sans permission pesage")
            log_user_action(
                user, 
                "Accès pesage refusé",
                f"Tentative d'accès au module pesage sans permission - Habilitation: {user.habilitation}",
                request
            )
            messages.error(request, _("Vous devez être affecté à une station de pesage pour accéder à cette fonctionnalité."))
            return None, redirect('common:dashboard')
        
        if station is None:
            logger.error(f"[PESAGE] Incohérence: {user.username} n'a pas accès tous postes mais station=None")
            messages.error(request, _("Erreur de configuration. Contactez un administrateur."))
            return None, redirect('common:dashboard')
        
        logger.debug(f"[PESAGE] Station {station.nom} assignée à {user.username}")
        return station, None
    
    # Cas 2: Utilisateur avec accès à tous les postes (admin, service central, cisop_pesage)
    station_id = (
        request.GET.get('station') or 
        request.POST.get('station') or 
        request.session.get('station_pesage_id')
    )
    
    if station_id:
        try:
            station = Poste.objects.get(pk=station_id, type='pesage', is_active=True)
            request.session['station_pesage_id'] = station.id
            logger.debug(f"[PESAGE] Admin {user.username} a sélectionné la station {station.nom}")
            return station, None
        except Poste.DoesNotExist:
            request.session.pop('station_pesage_id', None)
            logger.warning(f"[PESAGE] Station ID {station_id} invalide pour {user.username}")
    
    # Admin sans station sélectionnée
    return None, None


def valider_date_saisie(date_cible):
    """Valide qu'une date n'est pas dans le futur."""
    today = timezone.now().date()
    if date_cible > today:
        return False, _("Impossible de saisir des données pour une date future.")
    return True, None


# ===================================================================
# FONCTIONS UTILITAIRES TEMPORELLES (LOGIQUE 9h-9h)
# ===================================================================

def get_datetime_cameroun():
    """Retourne la date/heure actuelle au Cameroun (GMT+1)."""
    return timezone.now().astimezone(CAMEROUN_TZ)


def get_jour_travail_actuel():
    """Détermine le jour de travail actuel selon la logique 9h-9h."""
    maintenant = get_datetime_cameroun()
    if maintenant.time() < HEURE_DEBUT_JOURNEE:
        return (maintenant - timedelta(days=1)).date()
    else:
        return maintenant.date()


def calculer_periode_9h(periode_type, date_debut_custom=None, date_fin_custom=None):
    """Calcule les datetime de début et fin pour une période donnée avec logique 9h-9h."""
    jour_travail = get_jour_travail_actuel()
    
    if periode_type == 'personnalise' and date_debut_custom and date_fin_custom:
        if isinstance(date_debut_custom, str):
            try:
                date_debut = datetime.strptime(date_debut_custom, '%Y-%m-%d').date()
            except ValueError:
                date_debut = jour_travail.replace(day=1)
        else:
            date_debut = date_debut_custom
            
        if isinstance(date_fin_custom, str):
            try:
                date_fin = datetime.strptime(date_fin_custom, '%Y-%m-%d').date()
            except ValueError:
                date_fin = jour_travail
        else:
            date_fin = date_fin_custom
        
        datetime_debut = CAMEROUN_TZ.localize(datetime.combine(date_debut, HEURE_DEBUT_JOURNEE))
        datetime_fin = CAMEROUN_TZ.localize(datetime.combine(date_fin + timedelta(days=1), HEURE_FIN_JOURNEE))
        label = f"Du {date_debut.strftime('%d/%m/%Y')} 9h au {date_fin.strftime('%d/%m/%Y')} 9h"
        
        return {'datetime_debut': datetime_debut, 'datetime_fin': datetime_fin, 
                'date_debut': date_debut, 'date_fin': date_fin, 'label': label}
    
    if periode_type == 'jour':
        date_debut = jour_travail
        date_fin = jour_travail
        datetime_debut = CAMEROUN_TZ.localize(datetime.combine(jour_travail, HEURE_DEBUT_JOURNEE))
        datetime_fin = CAMEROUN_TZ.localize(datetime.combine(jour_travail + timedelta(days=1), HEURE_FIN_JOURNEE))
        label = f"Aujourd'hui ({jour_travail.strftime('%d/%m/%Y')}) de 9h à demain 9h"
        
    elif periode_type == 'semaine':
        date_debut = jour_travail - timedelta(days=6)
        date_fin = jour_travail
        datetime_debut = CAMEROUN_TZ.localize(datetime.combine(date_debut, HEURE_DEBUT_JOURNEE))
        datetime_fin = CAMEROUN_TZ.localize(datetime.combine(jour_travail + timedelta(days=1), HEURE_FIN_JOURNEE))
        label = f"7 derniers jours ({date_debut.strftime('%d/%m')} - {date_fin.strftime('%d/%m/%Y')})"
        
    elif periode_type == 'mois':
        date_debut = jour_travail.replace(day=1)
        date_fin = jour_travail
        datetime_debut = CAMEROUN_TZ.localize(datetime.combine(date_debut, HEURE_DEBUT_JOURNEE))
        datetime_fin = CAMEROUN_TZ.localize(datetime.combine(jour_travail + timedelta(days=1), HEURE_FIN_JOURNEE))
        label = f"Mois en cours ({date_debut.strftime('%d/%m')} - {date_fin.strftime('%d/%m/%Y')})"
        
    elif periode_type == 'trimestre':
        trimestre = (jour_travail.month - 1) // 3
        mois_debut = trimestre * 3 + 1
        date_debut = jour_travail.replace(month=mois_debut, day=1)
        date_fin = jour_travail
        datetime_debut = CAMEROUN_TZ.localize(datetime.combine(date_debut, HEURE_DEBUT_JOURNEE))
        datetime_fin = CAMEROUN_TZ.localize(datetime.combine(jour_travail + timedelta(days=1), HEURE_FIN_JOURNEE))
        label = f"Trimestre {trimestre + 1} ({date_debut.strftime('%d/%m')} - {date_fin.strftime('%d/%m/%Y')})"
        
    elif periode_type == 'annee':
        date_debut = jour_travail.replace(month=1, day=1)
        date_fin = jour_travail
        datetime_debut = CAMEROUN_TZ.localize(datetime.combine(date_debut, HEURE_DEBUT_JOURNEE))
        datetime_fin = CAMEROUN_TZ.localize(datetime.combine(jour_travail + timedelta(days=1), HEURE_FIN_JOURNEE))
        label = f"Année {jour_travail.year} ({date_debut.strftime('%d/%m')} - {date_fin.strftime('%d/%m/%Y')})"
        
    else:
        return calculer_periode_9h('mois')
    
    return {'datetime_debut': datetime_debut, 'datetime_fin': datetime_fin,
            'date_debut': date_debut, 'date_fin': date_fin, 'label': label}


def get_periode_dates(periode, date_ref=None):
    """Calcule les dates de début et fin pour une période donnée."""
    if date_ref is None:
        date_ref = timezone.now().date()
    
    if periode == 'jour':
        return date_ref, date_ref
    elif periode == 'semaine':
        date_debut = date_ref - timedelta(days=date_ref.weekday())
        date_fin = date_debut + timedelta(days=6)
        return date_debut, date_fin
    elif periode == 'mois':
        date_debut = date_ref.replace(day=1)
        if date_ref.month == 12:
            date_fin = date_ref.replace(day=31)
        else:
            date_fin = date_ref.replace(month=date_ref.month + 1, day=1) - timedelta(days=1)
        return date_debut, date_fin
    elif periode == 'trimestre':
        trimestre = (date_ref.month - 1) // 3
        date_debut = date_ref.replace(month=trimestre * 3 + 1, day=1)
        if trimestre == 3:
            date_fin = date_ref.replace(month=12, day=31)
        else:
            date_fin = date_ref.replace(month=(trimestre + 1) * 3 + 1, day=1) - timedelta(days=1)
        return date_debut, date_fin
    elif periode == 'annee':
        date_debut = date_ref.replace(month=1, day=1)
        date_fin = date_ref.replace(month=12, day=31)
        return date_debut, date_fin
    
    return get_periode_dates('mois', date_ref)


# ===================================================================
# DÉCORATEURS SPÉCIFIQUES PESAGE (AVEC GESTION DES POSTES)
# ===================================================================

def pesage_access_required(view_func):
    """
    Vérifie que l'utilisateur a accès au module pesage.
    Utilise la permission granulaire 'peut_lister_amendes' comme permission de base.
    """
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        
        # Vérifier si l'utilisateur a au moins une permission pesage
        if not has_any_permission(user, PERMISSIONS_PESAGE):
            log_user_action(
                user,
                "Accès module pesage refusé",
                f"Aucune permission pesage - Vue: {view_func.__name__}",
                request
            )
            log_acces_refuse(user, view_func.__name__, "Aucune permission pesage")
            messages.error(request, _("Accès non autorisé au module pesage."))
            return redirect('common:dashboard')
        
        # Vérifier l'accès à une station (sauf si accès tous postes)
        if not user_has_acces_tous_postes(user):
            station = get_user_station_pesage(user)
            if station is False:
                log_user_action(
                    user,
                    "Accès pesage - station manquante",
                    f"Permission pesage OK mais pas de station affectée",
                    request
                )
                messages.error(request, _("Vous devez être affecté à une station de pesage."))
                return redirect('common:dashboard')
        
        return view_func(request, *args, **kwargs)
    return wrapper


def saisie_pesage_required(view_func):
    """
    Vérifie que l'utilisateur peut saisir des données pesage (amendes ou pesées).
    Permissions: peut_saisir_amende OU peut_saisir_pesee_jour
    """
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        
        if not has_any_permission(user, ['peut_saisir_amende', 'peut_saisir_pesee_jour']):
            log_user_action(
                user,
                "Saisie pesage refusée",
                f"Permissions manquantes: peut_saisir_amende, peut_saisir_pesee_jour",
                request
            )
            log_acces_refuse(user, view_func.__name__, "Permission saisie pesage manquante")
            messages.error(request, _("Accès réservé aux chefs d'équipe et chefs de station pesage."))
            return redirect('common:dashboard')
        
        # Vérifier l'accès à une station
        if not user_has_acces_tous_postes(user):
            station = get_user_station_pesage(user)
            if station is False:
                messages.error(request, _("Vous devez être affecté à une station de pesage."))
                return redirect('common:dashboard')
        
        return view_func(request, *args, **kwargs)
    return wrapper


def validation_pesage_required(view_func):
    """
    Vérifie que l'utilisateur peut valider des paiements d'amendes.
    Permission: peut_valider_paiement_amende
    """
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        
        if not has_permission(user, 'peut_valider_paiement_amende'):
            log_user_action(
                user,
                "Validation paiement refusée",
                f"Permission manquante: peut_valider_paiement_amende",
                request
            )
            log_acces_refuse(user, view_func.__name__, "Permission validation paiement manquante")
            messages.error(request, _("Accès réservé aux régisseurs de station pesage."))
            return redirect('common:dashboard')
        
        # Vérifier l'accès à une station
        if not user_has_acces_tous_postes(user):
            station = get_user_station_pesage(user)
            if station is False:
                messages.error(request, _("Vous devez être affecté à une station de pesage."))
                return redirect('common:dashboard')
        
        return view_func(request, *args, **kwargs)
    return wrapper


def quittancement_pesage_required(view_func):
    """
    Vérifie que l'utilisateur peut gérer les quittancements pesage.
    Permission: peut_saisir_quittance_pesage
    """
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        
        if not has_permission(user, 'peut_saisir_quittance_pesage'):
            log_user_action(
                user,
                "Quittancement pesage refusé",
                f"Permission manquante: peut_saisir_quittance_pesage",
                request
            )
            log_acces_refuse(user, view_func.__name__, "Permission quittancement pesage manquante")
            messages.error(request, _("Accès réservé aux régisseurs pesage."))
            return redirect('common:dashboard')
        
        return view_func(request, *args, **kwargs)
    return wrapper


# ===================================================================
# SÉLECTION DE STATION (UTILISATEURS MULTI-POSTES UNIQUEMENT)
# ===================================================================

@login_required
def selectionner_station(request):
    """
    Page intermédiaire pour que les utilisateurs multi-postes sélectionnent une station.
    Accessible uniquement si l'utilisateur a accès à tous les postes.
    """
    user = request.user
    
    # Vérifier si l'utilisateur a réellement accès à tous les postes
    if not user_has_acces_tous_postes(user):
        log_user_action(
            user,
            "Sélection station - non autorisé",
            f"Utilisateur sans accès tous postes tente de sélectionner une station",
            request
        )
        messages.info(request, _("Vous êtes automatiquement affecté à votre station."))
        return redirect('common:dashboard')
    
    action = request.GET.get('action', request.POST.get('action', 'liste_amendes'))
    
    action_urls = {
        'saisir_amende': 'inventaire:saisir_amende',
        'saisir_pesees': 'inventaire:saisir_pesees',
        'saisie_quittancement_pesage': 'inventaire:saisie_quittancement_pesage',
        'liste_amendes': 'inventaire:liste_amendes',
        'amendes_a_valider': 'inventaire:liste_amendes_a_valider',
        'historique_pesees': 'inventaire:historique_pesees',
        'liste_quittancements_pesage': 'inventaire:liste_quittancements_pesage',
        'comptabilisation_quittancements_pesage': 'inventaire:comptabilisation_quittancements_pesage',
        'statistiques_pesage': 'inventaire:statistiques_pesage',
        'recettes_pesage': 'inventaire:recettes_pesage',
    }
    
    if request.method == 'POST':
        station_id = request.POST.get('station')
        if not station_id:
            messages.error(request, _("Veuillez sélectionner une station."))
        else:
            try:
                station = Poste.objects.get(pk=station_id, type='pesage', is_active=True)
                request.session['station_pesage_id'] = station.id
                
                log_user_action(
                    user,
                    "Station pesage sélectionnée",
                    f"Station: {station.nom} (ID: {station.id}) - Action: {action}",
                    request
                )
                
                url_name = action_urls.get(action, 'inventaire:liste_amendes')
                return redirect(f"{reverse(url_name)}?station={station.id}")
            except Poste.DoesNotExist:
                messages.error(request, _("Station invalide."))
    
    stations = Poste.objects.filter(type='pesage', is_active=True).order_by('region', 'nom')
    
    context = {
        'stations': stations,
        'action': action,
        'title': _('Sélectionner une station de pesage'),
        'is_admin': is_admin_user(user),
    }
    
    return render(request, 'pesage/selectionner_station.html', context)


# ===================================================================
# GESTION DES AMENDES - SAISIE
# Permission requise: peut_saisir_amende
# ===================================================================

@saisie_pesage_required
def saisir_amende(request):
    """
    Saisie d'une nouvelle amende.
    Permission: peut_saisir_amende
    """
    user = request.user
    today = timezone.now().date()
    
    # Vérifier la permission spécifique pour saisir une amende
    if not has_permission(user, 'peut_saisir_amende'):
        log_user_action(
            user,
            "Saisie amende - permission refusée",
            f"Permission peut_saisir_amende manquante",
            request
        )
        messages.error(request, _("Vous n'avez pas la permission de saisir des amendes."))
        return redirect('inventaire:liste_amendes')
    
    station, redirect_response = get_station_context(request, user)
    
    # Si utilisateur multi-postes sans station sélectionnée → rediriger vers sélection
    if user_has_acces_tous_postes(user) and station is None:
        return redirect(f"{reverse('inventaire:selectionner_station')}?action=saisir_amende")
    
    if redirect_response:
        return redirect_response
    
    if request.method == 'POST':
        form = AmendeEmiseForm(request.POST)
        
        # Validation de la date
        date_heure_emission_str = request.POST.get('date_heure_emission')
        if date_heure_emission_str:
            try:
                date_heure_emission = datetime.strptime(date_heure_emission_str, '%Y-%m-%d').date()
                is_valid, error_msg = valider_date_saisie(date_heure_emission)
                if not is_valid:
                    messages.error(request, error_msg)
                    return render(request, 'pesage/saisir_amende.html', {
                        'form': form, 
                        'station': station, 
                        'is_admin': is_admin_user(user),
                        'today': today, 
                        'title': _('Saisir une amende'),
                    })
            except ValueError:
                pass
        
        if form.is_valid():
            amende = form.save(commit=False)
            amende.station = station
            amende.saisi_par = user
            amende.statut = 'non_paye'
            amende.date_heure_emission = timezone.now()
            amende.save()
            
            messages.success(
                request,
                _("Amende %(numero)s créée avec succès. Montant: %(montant)s FCFA") % {
                    'numero': amende.numero_ticket,
                    'montant': f"{amende.montant_amende:,.0f}".replace(',', ' ')
                }
            )
            
            # Log détaillé de l'action
            log_user_action(
                user, 
                "Saisie amende pesage",
                f"N°Ticket: {amende.numero_ticket} | Montant: {amende.montant_amende} FCFA | "
                f"Immatriculation: {amende.immatriculation} | Station: {station.nom} | "
                f"Surcharge: {amende.est_surcharge} | Hors gabarit: {amende.est_hors_gabarit}",
                request
            )
            
            if 'saisir_autre' in request.POST:
                return redirect(f"{reverse('inventaire:saisir_amende')}?station={station.pk}")
            return redirect('inventaire:liste_amendes')
    else:
        form = AmendeEmiseForm()
    
    context = {
        'form': form, 
        'station': station, 
        'is_admin': is_admin_user(user),
        'today': today, 
        'title': _('Saisir une amende'),
        'peut_saisir_amende': has_permission(user, 'peut_saisir_amende'),
    }
    return render(request, 'pesage/saisir_amende.html', context)


# ===================================================================
# LISTE DES AMENDES
# Permission requise: peut_lister_amendes
# ===================================================================

@pesage_access_required
def liste_amendes(request):
    """
    Liste des amendes avec filtres de recherche.
    Permission: peut_lister_amendes
    """
    user = request.user
    stations_accessibles = get_stations_pesage_accessibles(user)
    
    station, redirect_response = get_station_context(request, user)
    if redirect_response:
        return redirect_response
    
    # Construction du queryset selon les droits
    if station is None and user_has_acces_tous_postes(user):
        # Utilisateur multi-postes sans station sélectionnée → toutes les amendes
        queryset = AmendeEmise.objects.all()
        station_filter = request.GET.get('station_filter')
        if station_filter:
            queryset = queryset.filter(station_id=station_filter)
            logger.debug(f"[PESAGE] {user.username} filtre les amendes par station ID {station_filter}")
    elif station:
        # Utilisateur avec station spécifique
        queryset = AmendeEmise.objects.filter(station=station)
    else:
        queryset = AmendeEmise.objects.none()
    
    # Filtres de recherche
    query = request.GET.get('q', '').strip()
    if query:
        queryset = queryset.filter(
            Q(numero_ticket__icontains=query) | 
            Q(immatriculation__icontains=query) |
            Q(transporteur__icontains=query) | 
            Q(provenance__icontains=query) |
            Q(destination__icontains=query)
        )
    
    statut_filter = request.GET.get('statut')
    if statut_filter:
        queryset = queryset.filter(statut=statut_filter)
    
    infraction_filter = request.GET.get('type_infraction')
    if infraction_filter == 'S':
        queryset = queryset.filter(est_surcharge=True, est_hors_gabarit=False)
    elif infraction_filter == 'HG':
        queryset = queryset.filter(est_hors_gabarit=True, est_surcharge=False)
    elif infraction_filter == 'S+HG':
        queryset = queryset.filter(est_surcharge=True, est_hors_gabarit=True)
    
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if date_debut:
        queryset = queryset.filter(date_heure_emission__date__gte=date_debut)
    if date_fin:
        queryset = queryset.filter(date_heure_emission__date__lte=date_fin)
    
    queryset = queryset.select_related('station', 'saisi_par').order_by('-date_heure_emission')
    
    # Statistiques
    stats = queryset.aggregate(
        total_montant=Sum('montant_amende'),
        total_paye=Sum('montant_amende', filter=Q(statut='paye')),
        total_non_paye=Sum('montant_amende', filter=Q(statut='non_paye')),
        count_total=Count('id'),
        count_paye=Count('id', filter=Q(statut='paye')),
        count_non_paye=Count('id', filter=Q(statut='non_paye')),
    )
    
    # Pagination
    paginator = Paginator(queryset, 25)
    page = request.GET.get('page')
    amendes = paginator.get_page(page)
    
    context = {
        'amendes': amendes, 
        'station': station, 
        'stations_accessibles': stations_accessibles,
        'stats': stats, 
        'is_admin': is_admin_user(user), 
        'query': query, 
        'statut_filter': statut_filter,
        'type_infraction': infraction_filter, 
        'date_debut': date_debut, 
        'date_fin': date_fin,
        'station_filter': request.GET.get('station_filter', ''), 
        'title': _('Liste des amendes'),
        'peut_valider_paiement': has_permission(user, 'peut_valider_paiement_amende'),
        'peut_saisir_amende': has_permission(user, 'peut_saisir_amende'),
    }
    return render(request, 'pesage/liste_amendes.html', context)


@pesage_access_required
def detail_amende(request, pk):
    """
    Détail d'une amende.
    AMÉLIORATION: Détecte si le véhicule a des amendes non payées 
    dans d'autres stations et affiche un avertissement.
    """
    user = request.user
    station, _unused = get_station_context(request, user)
    
    # Récupérer l'amende selon les droits
    if user_has_acces_tous_postes(user):
        amende = get_object_or_404(AmendeEmise, pk=pk)
    else:
        amende = get_object_or_404(AmendeEmise, pk=pk, station=station)
    
    # Récupérer les événements si disponibles
    events = []
    if hasattr(amende, 'events'):
        events = amende.events.select_related('effectue_par').order_by('-event_datetime')
    
    # =====================================================================
    # NOUVEAU: Détecter les amendes non payées du véhicule
    # =====================================================================
    from .utils_pesage import normalize_immatriculation
    
    immat_norm = normalize_immatriculation(amende.immatriculation)
    
    # Toutes les amendes non payées de ce véhicule (sauf celle-ci)
    amendes_non_payees = AmendeEmise.objects.filter(
        Q(immatriculation_normalise__iexact=immat_norm) |
        Q(immatriculation__iexact=amende.immatriculation) |
        Q(immatriculation__iexact=immat_norm),
        statut='non_paye'
    ).exclude(pk=amende.pk).select_related('station')
    
    # Séparer par station
    amendes_autres_stations = amendes_non_payees.exclude(station=amende.station)
    amendes_meme_station = amendes_non_payees.filter(station=amende.station)
    
    # Infos pour l'affichage
    infos_autres_amendes = None
    if amendes_non_payees.exists():
        infos_autres_amendes = {
            'total': amendes_non_payees.count(),
            'autres_stations': {
                'count': amendes_autres_stations.count(),
                'montant': amendes_autres_stations.aggregate(total=Sum('montant_amende'))['total'] or 0,
                'stations': list(amendes_autres_stations.values_list('station__nom', flat=True).distinct()),
            },
            'meme_station': {
                'count': amendes_meme_station.count(),
                'montant': amendes_meme_station.aggregate(total=Sum('montant_amende'))['total'] or 0,
            },
            'montant_total': amendes_non_payees.aggregate(total=Sum('montant_amende'))['total'] or 0,
        }
    
    # =====================================================================
    # FIN NOUVEAU CODE
    # =====================================================================
    
    log_user_action(
        user,
        "Consultation détail amende",
        f"Amende {amende.numero_ticket} - Station: {amende.station.nom}",
        request
    )
    
    context = {
        'amende': amende, 
        'events': events, 
        'station': station or amende.station,
        'is_admin': is_admin_user(user), 
        'title': f"Amende {amende.numero_ticket}",
        'peut_valider_paiement': has_permission(user, 'peut_valider_paiement_amende'),
        # Nouvelles variables
        'infos_autres_amendes': infos_autres_amendes,
        'has_autres_amendes': infos_autres_amendes is not None,
    }
    return render(request, 'pesage/detail_amende.html', context)

# ===================================================================
# VALIDATION DES PAIEMENTS
# Permission requise: peut_valider_paiement_amende
# ===================================================================

@validation_pesage_required
def liste_amendes_a_valider(request):
    """
    Liste des amendes en attente de validation.
    Permission: peut_valider_paiement_amende
    
    AMÉLIORATION: Détecte pour chaque amende si le véhicule a des amendes
    non payées dans d'autres stations et prépare les infos pour l'affichage.
    """
    user = request.user
    stations_accessibles = get_stations_pesage_accessibles(user)
    peut_valider = has_permission(user, 'peut_valider_paiement_amende')
    
    station, redirect_response = get_station_context(request, user)
    if redirect_response:
        return redirect_response
    
    # Construction du queryset
    if station is None and user_has_acces_tous_postes(user):
        queryset = AmendeEmise.objects.filter(statut='non_paye')
        station_filter = request.GET.get('station_filter')
        if station_filter:
            queryset = queryset.filter(station_id=station_filter)
    elif station:
        queryset = AmendeEmise.objects.filter(station=station, statut='non_paye')
    else:
        queryset = AmendeEmise.objects.none()
    
    queryset = queryset.select_related('station', 'saisi_par').order_by('-date_heure_emission')
    total_a_recouvrer = queryset.aggregate(total=Sum('montant_amende'))['total'] or Decimal('0')
    
    # =====================================================================
    # NOUVEAU: Détecter les véhicules avec amendes dans autres stations
    # =====================================================================
    
    # Récupérer toutes les immatriculations de la page actuelle
    paginator = Paginator(queryset, 25)
    page = request.GET.get('page')
    amendes_page = paginator.get_page(page)
    
    # Collecter les immatriculations uniques de cette page
    immatriculations_page = set()
    for amende in amendes_page:
        immat_norm = normalize_immatriculation(amende.immatriculation)
        immatriculations_page.add(immat_norm)
        immatriculations_page.add(amende.immatriculation)
    
    # Rechercher TOUTES les amendes non payées pour ces véhicules
    # (incluant celles dans d'autres stations)
    if immatriculations_page:
        amendes_autres_stations = AmendeEmise.objects.filter(
            Q(immatriculation__in=immatriculations_page) |
            Q(immatriculation_normalise__in=immatriculations_page),
            statut='non_paye'
        ).exclude(
            station=station  # Exclure la station actuelle
        ).values(
            'immatriculation', 
            'immatriculation_normalise',
            'station__nom',
            'station__id'
        ).annotate(
            nb_amendes=Count('pk'),
            montant_total=Sum('montant_amende')
        )
        
        # Construire un dictionnaire: immatriculation -> infos autres stations
        # Clé: immatriculation normalisée
        vehicules_avec_autres_amendes = {}
        
        for item in amendes_autres_stations:
            immat_norm = normalize_immatriculation(item['immatriculation'])
            
            if immat_norm not in vehicules_avec_autres_amendes:
                vehicules_avec_autres_amendes[immat_norm] = {
                    'stations': [],
                    'total_amendes': 0,
                    'montant_total': Decimal('0'),
                }
            
            # Ajouter cette station si pas déjà présente
            station_info = {
                'nom': item['station__nom'],
                'id': item['station__id'],
                'nb': item['nb_amendes'],
                'montant': item['montant_total'] or Decimal('0'),
            }
            
            # Éviter les doublons de station
            stations_existantes = [s['id'] for s in vehicules_avec_autres_amendes[immat_norm]['stations']]
            if item['station__id'] not in stations_existantes:
                vehicules_avec_autres_amendes[immat_norm]['stations'].append(station_info)
                vehicules_avec_autres_amendes[immat_norm]['total_amendes'] += item['nb_amendes']
                vehicules_avec_autres_amendes[immat_norm]['montant_total'] += (item['montant_total'] or Decimal('0'))
    else:
        vehicules_avec_autres_amendes = {}
    
    # Enrichir chaque amende avec les infos des autres stations
    amendes_enrichies = []
    for amende in amendes_page:
        immat_norm = normalize_immatriculation(amende.immatriculation)
        
        # Vérifier si ce véhicule a des amendes ailleurs
        infos_autres = vehicules_avec_autres_amendes.get(immat_norm)
        
        amendes_enrichies.append({
            'amende': amende,
            'has_autres_stations': infos_autres is not None,
            'autres_stations_info': infos_autres,
        })
    
    # =====================================================================
    # FIN NOUVEAU CODE
    # =====================================================================
    
    context = {
        'amendes': amendes_page,  # Pour la pagination
        'amendes_enrichies': amendes_enrichies,  # Avec les infos supplémentaires
        'vehicules_avec_autres_amendes': vehicules_avec_autres_amendes,  # Dictionnaire complet
        'station': station, 
        'stations_accessibles': stations_accessibles,
        'total_a_recouvrer': total_a_recouvrer, 
        'count_total': queryset.count(),
        'is_admin': is_admin_user(user), 
        'peut_valider': peut_valider,
        'station_filter': request.GET.get('station_filter', ''), 
        'title': _('Amendes à valider'),
    }
    return render(request, 'pesage/liste_amendes_a_valider.html', context)


# ===================================================================
# API pour récupérer les détails des amendes d'un véhicule (AJAX)
# ===================================================================

from django.http import JsonResponse
from django.views.decorators.http import require_GET

@validation_pesage_required
@require_GET
def api_amendes_vehicule_autres_stations(request, immatriculation):
    """
    API JSON qui retourne les amendes non payées d'un véhicule dans d'autres stations.
    Utilisé pour le modal de détails.
    """
    user = request.user
    station, _ = get_station_context(request, user)
    
    immat_norm = normalize_immatriculation(immatriculation)
    
    # Rechercher les amendes non payées dans AUTRES stations
    amendes = AmendeEmise.objects.filter(
        Q(immatriculation__iexact=immatriculation) |
        Q(immatriculation__iexact=immat_norm) |
        Q(immatriculation_normalise__iexact=immat_norm),
        statut='non_paye'
    ).exclude(
        station=station
    ).select_related('station').order_by('date_heure_emission')
    
    # Préparer les données JSON
    data = {
        'immatriculation': immatriculation,
        'count': amendes.count(),
        'montant_total': float(amendes.aggregate(total=Sum('montant_amende'))['total'] or 0),
        'amendes': []
    }
    
    for amende in amendes:
        data['amendes'].append({
            'id': amende.pk,
            'numero_ticket': amende.numero_ticket,
            'station': amende.station.nom,
            'station_id': amende.station.id,
            'date_emission': amende.date_heure_emission.strftime('%d/%m/%Y %H:%i'),
            'montant': float(amende.montant_amende),
            'transporteur': amende.transporteur or '-',
            'infraction': 'S+HG' if (amende.est_surcharge and amende.est_hors_gabarit) else ('S' if amende.est_surcharge else ('HG' if amende.est_hors_gabarit else '-')),
        })
    
    return JsonResponse(data)


@validation_pesage_required
@require_POST
def valider_paiement(request, pk):
    """
    Valide le paiement d'une amende.
    Permission: peut_valider_paiement_amende
    
    COMPORTEMENT:
    1. Vérifier si le véhicule a des amendes NON PAYÉES plus anciennes dans d'AUTRES stations
    2. Si oui → AVERTIR l'utilisateur mais VALIDER quand même le paiement
    3. Les amendes antérieures restent non payées et visibles dans la liste
    """
    user = request.user
    
    # Double vérification de la permission
    if not has_permission(user, 'peut_valider_paiement_amende'):
        log_user_action(
            user,
            "Validation paiement - permission refusée",
            f"Tentative de validation sans permission pour amende ID {pk}",
            request
        )
        messages.error(request, _("Vous n'avez pas la permission de valider les paiements."))
        return redirect('inventaire:liste_amendes_a_valider')
    
    station, _unused = get_station_context(request, user)
    
    # Récupérer l'amende selon les droits
    if user_has_acces_tous_postes(user):
        amende = get_object_or_404(AmendeEmise, pk=pk)
    else:
        amende = get_object_or_404(AmendeEmise, pk=pk, station=station)
    
    # Vérification 1: L'amende est-elle déjà payée ?
    if amende.statut == 'paye':
        messages.warning(request, _("Cette amende a déjà été validée."))
        return redirect('inventaire:liste_amendes_a_valider')
    
    # Vérification 2: Y a-t-il des amendes NON PAYÉES plus anciennes ailleurs ?
    # (Pour avertissement uniquement, pas de blocage)
    immat_norm = normalize_immatriculation(amende.immatriculation)
    
    amendes_anterieures_non_payees = AmendeEmise.objects.filter(
        Q(immatriculation_normalise__iexact=immat_norm) |
        Q(immatriculation__iexact=amende.immatriculation) |
        Q(immatriculation__iexact=immat_norm),
        statut='non_paye',
        date_heure_emission__lt=amende.date_heure_emission
    ).exclude(
        pk=amende.pk
    ).select_related('station').order_by('date_heure_emission')
    
    # Préparer l'avertissement si amendes antérieures
    avertissement_amendes = None
    if amendes_anterieures_non_payees.exists():
        # Séparer les amendes de la même station et des autres stations
        amendes_meme_station = amendes_anterieures_non_payees.filter(station=amende.station)
        amendes_autres_stations = amendes_anterieures_non_payees.exclude(station=amende.station)
        
        avertissement_amendes = {
            'total': amendes_anterieures_non_payees.count(),
            'meme_station': amendes_meme_station.count(),
            'autres_stations': amendes_autres_stations.count(),
            'stations': list(amendes_autres_stations.values_list('station__nom', flat=True).distinct()),
            'montant_total': amendes_anterieures_non_payees.aggregate(total=Sum('montant_amende'))['total'] or 0,
        }
    
    # VALIDER LE PAIEMENT (même s'il y a des amendes antérieures)
    try:
        amende.statut = 'paye'
        amende.date_paiement = timezone.now()
        amende.valide_par = user
        amende.save()
        
        # Message de succès
        messages.success(request,
            _("✅ Paiement validé pour l'amende %(numero)s - Montant: %(montant)s FCFA") % {
                'numero': amende.numero_ticket,
                'montant': f"{amende.montant_amende:,.0f}".replace(',', ' ')
            })
        
        # Avertissement sur les amendes antérieures (si présentes)
        if avertissement_amendes:
            if avertissement_amendes['autres_stations'] > 0:
                messages.warning(request,
                    _("⚠️ ATTENTION: Ce véhicule (%(immat)s) a encore %(nb)d amende(s) non payée(s) "
                      "dans d'autres stations: %(stations)s. Montant total impayé: %(montant)s FCFA") % {
                        'immat': amende.immatriculation,
                        'nb': avertissement_amendes['autres_stations'],
                        'stations': ', '.join(avertissement_amendes['stations']),
                        'montant': f"{avertissement_amendes['montant_total']:,.0f}".replace(',', ' ')
                    })
            
            if avertissement_amendes['meme_station'] > 0:
                messages.info(request,
                    _("ℹ️ Ce véhicule a également %(nb)d amende(s) antérieure(s) non payée(s) "
                      "dans cette station.") % {
                        'nb': avertissement_amendes['meme_station']
                    })
        
        # Journalisation
        details_log = (
            f"N°Ticket: {amende.numero_ticket} | Montant: {amende.montant_amende} FCFA | "
            f"Véhicule: {amende.immatriculation} | Station: {amende.station.nom} | "
            f"Date émission: {amende.date_heure_emission}"
        )
        
        if avertissement_amendes:
            details_log += (
                f" | AVERTISSEMENT: {avertissement_amendes['total']} amende(s) antérieure(s) "
                f"non payée(s) ({avertissement_amendes['autres_stations']} dans autres stations)"
            )
        
        log_user_action(
            user, 
            "Validation paiement amende",
            details_log,
            request
        )
        
        # Log spécifique si amendes antérieures
        if avertissement_amendes and avertissement_amendes['autres_stations'] > 0:
            logger.info(
                f"Validation avec avertissement: Amende {amende.numero_ticket} validée - "
                f"Véhicule {amende.immatriculation} a {avertissement_amendes['autres_stations']} "
                f"amendes antérieures dans: {avertissement_amendes['stations']}"
            )
        
    except Exception as e:
        logger.error(f"Erreur validation paiement: {e}", exc_info=True)
        log_user_action(
            user,
            "Erreur validation paiement",
            f"Amende {amende.numero_ticket} - Erreur: {str(e)}",
            request
        )
        messages.error(request, _("Erreur lors de la validation du paiement."))
    
    return redirect('inventaire:liste_amendes_a_valider')


@validation_pesage_required
@require_POST
def valider_paiements_masse(request):
    """
    Valide plusieurs paiements en une fois.
    Permission: peut_valider_paiement_amende
    
    COMPORTEMENT:
    - Valide toutes les amendes sélectionnées
    - Avertit si des véhicules ont des amendes antérieures non payées
    """
    user = request.user
    
    if not has_permission(user, 'peut_valider_paiement_amende'):
        log_user_action(
            user,
            "Validation masse - permission refusée",
            "Tentative de validation en masse sans permission",
            request
        )
        messages.error(request, _("Vous n'avez pas la permission de valider les paiements."))
        return redirect('inventaire:liste_amendes_a_valider')
    
    station, _unused = get_station_context(request, user)
    amende_ids = request.POST.getlist('amendes')
    
    if not amende_ids:
        messages.warning(request, _("Aucune amende sélectionnée."))
        return redirect('inventaire:liste_amendes_a_valider')
    
    count_success = 0
    montant_total = Decimal('0')
    amendes_validees = []
    vehicules_avec_anterieures = []  # Pour l'avertissement
    
    for amende_id in amende_ids:
        try:
            if user_has_acces_tous_postes(user):
                amende = AmendeEmise.objects.get(pk=amende_id, statut='non_paye')
            else:
                amende = AmendeEmise.objects.get(pk=amende_id, station=station, statut='non_paye')
            
            # Vérifier amendes antérieures (pour avertissement)
            immat_norm = normalize_immatriculation(amende.immatriculation)
            amendes_anterieures = AmendeEmise.objects.filter(
                Q(immatriculation_normalise__iexact=immat_norm) |
                Q(immatriculation__iexact=amende.immatriculation),
                statut='non_paye',
                date_heure_emission__lt=amende.date_heure_emission
            ).exclude(station=amende.station).exclude(pk=amende.pk)
            
            if amendes_anterieures.exists():
                vehicules_avec_anterieures.append({
                    'immat': amende.immatriculation,
                    'nb': amendes_anterieures.count(),
                    'stations': list(amendes_anterieures.values_list('station__nom', flat=True).distinct())
                })
            
            # Valider le paiement
            amende.statut = 'paye'
            amende.date_paiement = timezone.now()
            amende.valide_par = user
            amende.save()
            
            count_success += 1
            montant_total += amende.montant_amende
            amendes_validees.append(amende.numero_ticket)
            
        except AmendeEmise.DoesNotExist:
            continue
        except Exception as e:
            logger.warning(f"Erreur validation amende {amende_id}: {e}")
            continue
    
    if count_success > 0:
        messages.success(request,
            _("✅ %(count)d paiement(s) validé(s) - Total: %(montant)s FCFA") % {
                'count': count_success, 
                'montant': f"{montant_total:,.0f}".replace(',', ' ')
            })
        
        # Avertissement si des véhicules ont des amendes antérieures
        if vehicules_avec_anterieures:
            # Regrouper par véhicule
            vehicules_uniques = {}
            for v in vehicules_avec_anterieures:
                if v['immat'] not in vehicules_uniques:
                    vehicules_uniques[v['immat']] = v
            
            messages.warning(request,
                _("⚠️ %(nb_vehicules)d véhicule(s) ont des amendes antérieures non payées "
                  "dans d'autres stations. Ces amendes restent à payer.") % {
                    'nb_vehicules': len(vehicules_uniques)
                })
        
        log_user_action(
            user, 
            "Validation paiements en masse",
            f"{count_success} amendes validées | Total: {montant_total} FCFA | "
            f"N°Tickets: {', '.join(amendes_validees[:10])}{'...' if len(amendes_validees) > 10 else ''}"
            f"{' | ' + str(len(vehicules_avec_anterieures)) + ' véhicules avec amendes antérieures' if vehicules_avec_anterieures else ''}",
            request
        )
    else:
        messages.warning(request, _("Aucun paiement n'a pu être validé."))
    
    return redirect('inventaire:liste_amendes_a_valider')


# ===================================================================
# VUE OPTIONNELLE: CONSULTATION DES AMENDES ANTÉRIEURES
# Pour permettre au régisseur de voir les détails avant validation
# ===================================================================

@validation_pesage_required
@require_GET
def consulter_amendes_anterieures(request, pk):
    """
    Affiche les détails des amendes antérieures d'un véhicule avant validation.
    Page informative qui permet ensuite de valider.
    """
    user = request.user
    station, _unused = get_station_context(request, user)
    
    # Récupérer l'amende
    if user_has_acces_tous_postes(user):
        amende = get_object_or_404(AmendeEmise, pk=pk)
    else:
        amende = get_object_or_404(AmendeEmise, pk=pk, station=station)
    
    # Rechercher les amendes antérieures
    immat_norm = normalize_immatriculation(amende.immatriculation)
    
    amendes_anterieures = AmendeEmise.objects.filter(
        Q(immatriculation_normalise__iexact=immat_norm) |
        Q(immatriculation__iexact=amende.immatriculation) |
        Q(immatriculation__iexact=immat_norm),
        statut='non_paye',
        date_heure_emission__lt=amende.date_heure_emission
    ).exclude(pk=amende.pk).select_related('station').order_by('date_heure_emission')
    
    # Séparer par station
    amendes_autres_stations = amendes_anterieures.exclude(station=amende.station)
    amendes_meme_station = amendes_anterieures.filter(station=amende.station)
    
    # Historique complet du véhicule
    historique = AmendeEmise.objects.filter(
        Q(immatriculation_normalise__iexact=immat_norm) |
        Q(immatriculation__iexact=amende.immatriculation)
    ).select_related('station').order_by('-date_heure_emission')
    
    # Statistiques
    stats = historique.aggregate(
        total=Count('pk'),
        payees=Count('pk', filter=Q(statut='paye')),
        non_payees=Count('pk', filter=Q(statut='non_paye')),
        montant_total=Sum('montant_amende'),
        montant_paye=Sum('montant_amende', filter=Q(statut='paye')),
        montant_impaye=Sum('montant_amende', filter=Q(statut='non_paye')),
    )
    
    context = {
        'title': _('Détails avant validation'),
        'amende': amende,
        'station': station,
        'amendes_autres_stations': amendes_autres_stations,
        'amendes_meme_station': amendes_meme_station,
        'historique': historique[:20],  # Limiter à 20 dernières
        'stats': stats,
        'has_anterieures': amendes_anterieures.exists(),
        'montant_total_impaye_autres': amendes_autres_stations.aggregate(
            total=Sum('montant_amende'))['total'] or 0,
    }
    
    return render(request, 'pesage/consulter_amendes_anterieures.html', context)


# ===================================================================
# VUE DE VALIDATION DIRECTE DEPUIS LA PAGE DE CONSULTATION
# ===================================================================

@validation_pesage_required
@require_POST
def valider_paiement_avec_avertissement(request, pk):
    """
    Valide le paiement après que l'utilisateur a vu l'avertissement.
    Identique à valider_paiement mais avec confirmation explicite.
    """
    # Vérifier que l'utilisateur a coché la case de confirmation
    confirmation = request.POST.get('confirmer_validation', False)
    
    if not confirmation:
        messages.error(request, _("Veuillez confirmer que vous avez pris connaissance des amendes antérieures."))
        return redirect('inventaire:consulter_amendes_anterieures', pk=pk)
    
    # Réutiliser la logique de validation standard
    return valider_paiement(request, pk)



# ===================================================================
# PESÉES JOURNALIÈRES
# Permission requise: peut_saisir_pesee_jour
# ===================================================================

@saisie_pesage_required
def saisir_pesees(request):
    """
    Saisie du nombre de pesées journalières.
    Permission: peut_saisir_pesee_jour
    """
    user = request.user
    today = timezone.now().date()
    
    # Vérifier la permission spécifique
    if not has_permission(user, 'peut_saisir_pesee_jour'):
        log_user_action(
            user,
            "Saisie pesées - permission refusée",
            "Permission peut_saisir_pesee_jour manquante",
            request
        )
        messages.error(request, _("Vous n'avez pas la permission de saisir des pesées."))
        return redirect('inventaire:historique_pesees')
    
    station, redirect_response = get_station_context(request, user)
    
    if user_has_acces_tous_postes(user) and station is None:
        return redirect(f"{reverse('inventaire:selectionner_station')}?action=saisir_pesees")
    
    if redirect_response:
        return redirect_response
    
    # Gestion de la date
    date_str = request.GET.get('date') or request.POST.get('date')
    if date_str:
        try:
            date_selectionnee = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date_selectionnee = today
    else:
        date_selectionnee = today
    
    is_valid, error_msg = valider_date_saisie(date_selectionnee)
    if not is_valid:
        messages.error(request, error_msg)
        date_selectionnee = today
    
    # Vérifier si déjà saisi
    pesee_existante = PeseesJournalieres.objects.filter(
        station=station, 
        date=date_selectionnee
    ).first()
    deja_saisi = pesee_existante is not None
    
    if request.method == 'POST' and not deja_saisi:
        date_post = request.POST.get('date')
        if date_post:
            try:
                date_post_obj = datetime.strptime(date_post, '%Y-%m-%d').date()
                is_valid, error_msg = valider_date_saisie(date_post_obj)
                if not is_valid:
                    messages.error(request, error_msg)
                    return redirect('inventaire:saisir_pesees')
                date_selectionnee = date_post_obj
            except ValueError:
                pass
        
        form = PeseesJournalieresForm(request.POST)
        if form.is_valid():
            pesees = form.save(commit=False)
            pesees.station = station
            pesees.date = date_selectionnee
            pesees.saisi_par = user
            pesees.save()
            
            messages.success(request,
                _("Pesées du %(date)s enregistrées: %(nombre)d pesées") % {
                    'date': pesees.date.strftime('%d/%m/%Y'), 
                    'nombre': pesees.nombre_pesees
                })
            
            log_user_action(
                user, 
                "Saisie pesées journalières",
                f"Station: {station.nom} | Date: {pesees.date.strftime('%d/%m/%Y')} | "
                f"Nombre de pesées: {pesees.nombre_pesees}",
                request
            )
            
            return redirect('inventaire:historique_pesees')
    else:
        form = PeseesJournalieresForm()
    
    context = {
        'form': form, 
        'station': station, 
        'today': today, 
        'date_selectionnee': date_selectionnee,
        'deja_saisi': deja_saisi, 
        'pesee_existante': pesee_existante,
        'is_admin': is_admin_user(user), 
        'title': _('Saisir les pesées du jour'),
        'peut_saisir_pesee': has_permission(user, 'peut_saisir_pesee_jour'),
    }
    return render(request, 'pesage/saisir_pesees.html', context)


@pesage_access_required
def historique_pesees(request):
    """
    Historique des pesées journalières.
    Permission: peut_voir_historique_pesees
    """
    user = request.user
    
    # Vérifier la permission
    if not has_permission(user, 'peut_voir_historique_pesees'):
        log_user_action(
            user,
            "Historique pesées - permission refusée",
            "Permission peut_voir_historique_pesees manquante",
            request
        )
        messages.error(request, _("Vous n'avez pas la permission de voir l'historique des pesées."))
        return redirect('common:dashboard')
    
    stations_accessibles = get_stations_pesage_accessibles(user)
    
    station, redirect_response = get_station_context(request, user)
    if redirect_response:
        return redirect_response
    
    # Construction du queryset
    if station is None and user_has_acces_tous_postes(user):
        queryset = PeseesJournalieres.objects.all()
        station_filter = request.GET.get('station_filter')
        if station_filter:
            queryset = queryset.filter(station_id=station_filter)
    elif station:
        queryset = PeseesJournalieres.objects.filter(station=station)
    else:
        queryset = PeseesJournalieres.objects.none()
    
    queryset = queryset.select_related('station', 'saisi_par').order_by('-date')
    stats = queryset.aggregate(total_pesees=Sum('nombre_pesees'), count_jours=Count('id'))
    
    paginator = Paginator(queryset, 30)
    page = request.GET.get('page')
    pesees = paginator.get_page(page)
    
    context = {
        'pesees': pesees, 
        'station': station, 
        'stations_accessibles': stations_accessibles,
        'stats': stats, 
        'is_admin': is_admin_user(user),
        'station_filter': request.GET.get('station_filter', ''), 
        'title': _('Historique des pesées'),
        'peut_saisir_pesee': has_permission(user, 'peut_saisir_pesee_jour'),
    }
    return render(request, 'pesage/historique_pesees.html', context)


# ===================================================================
# QUITTANCEMENTS - SAISIE
# Permission requise: peut_saisir_quittance_pesage
# ===================================================================

@quittancement_pesage_required
def saisie_quittancement_pesage(request):
    """
    Vue pour saisir un quittancement pesage (version en 3 étapes)
    Permission: peut_saisir_quittance_pesage
    """
    user = request.user
    
    # Nettoyer la session au début si nouvelle saisie
    if request.method == 'GET' and 'etape' not in request.GET:
        for key in ['exercice_pesage_temp', 'mois_pesage_temp', 'type_declaration_pesage_temp', 
                   'form_data_pesage_temp', 'image_pesage_temp_path', 'station_pesage_id']:
            request.session.pop(key, None)

    # GESTION DE LA STATION
    if user_has_acces_tous_postes(user):
        station_id = (
            request.GET.get('station') or 
            request.POST.get('station') or 
            request.session.get('station_pesage_id')
        )
        
        if not station_id:
            return redirect(f"{reverse('inventaire:selectionner_station')}?action=saisie_quittancement_pesage")
        
        try:
            station = Poste.objects.get(pk=station_id, type='pesage', is_active=True)
            request.session['station_pesage_id'] = station.id
        except Poste.DoesNotExist:
            request.session.pop('station_pesage_id', None)
            messages.error(request, "❌ Station invalide, veuillez en sélectionner une autre")
            return redirect(f"{reverse('inventaire:selectionner_station')}?action=saisie_quittancement_pesage")
    else:
        poste = getattr(user, 'poste_affectation', None)
        if poste and poste.type == 'pesage':
            station = poste
        else:
            log_user_action(
                user,
                "Quittancement pesage - station manquante",
                "Utilisateur sans station de pesage tente de saisir un quittancement",
                request
            )
            messages.error(request, "❌ Vous devez être affecté à une station de pesage")
            return redirect('common:dashboard')
    
    etape = int(request.POST.get('etape', request.GET.get('etape', 1)))
    
    # ÉTAPE 1 : Paramètres globaux
    if etape == 1:
        if request.method == 'POST':
            exercice = request.POST.get('exercice')
            mois = request.POST.get('mois')
            type_declaration = request.POST.get('type_declaration')
            
            errors = []
            if not exercice: errors.append("Veuillez sélectionner un exercice")
            if not mois: errors.append("Veuillez sélectionner un mois")
            if not type_declaration: errors.append("Veuillez sélectionner un type de déclaration")
            
            if errors:
                for error in errors:
                    messages.error(request, f"⚠️ {error}")
                return redirect(f"{reverse('inventaire:saisie_quittancement_pesage')}?station={station.id}")
            
            request.session['exercice_pesage_temp'] = exercice
            request.session['mois_pesage_temp'] = mois
            request.session['type_declaration_pesage_temp'] = type_declaration
            
            log_user_action(
                user,
                "Quittancement pesage - étape 1",
                f"Station: {station.nom} | Exercice: {exercice} | Mois: {mois} | Type: {type_declaration}",
                request
            )
            
            return redirect(f"{reverse('inventaire:saisie_quittancement_pesage')}?etape=2&station={station.id}")
        
        annee_courante = datetime.now().year
        annees = list(range(annee_courante - 5, annee_courante + 2))
        
        context = {
            'etape': 1, 
            'station': station, 
            'annees': annees, 
            'annee_courante': annee_courante,
            'types_declaration': [('journaliere', 'Journalière (par jour)'), ('decade', 'Par décade')],
            'is_admin': is_admin_user(user),
        }
        return render(request, 'pesage/saisie_quittancement_pesage.html', context)
    
    # ÉTAPE 2 : Saisie du quittancement + IMAGE
    elif etape == 2:
        exercice = request.session.get('exercice_pesage_temp')
        mois = request.session.get('mois_pesage_temp')
        type_declaration = request.session.get('type_declaration_pesage_temp')
        
        if not all([exercice, mois, type_declaration]):
            messages.error(request, "❌ Session expirée, veuillez recommencer")
            return redirect(f"{reverse('inventaire:saisie_quittancement_pesage')}?station={station.id}")
       
        if request.method == 'POST':
            form_data = {
                'station_id': station.id,
                'numero_quittance': request.POST.get('numero_quittance'),
                'date_quittancement': request.POST.get('date_quittancement'),
                'montant': request.POST.get('montant'),
                'observations': request.POST.get('observations', ''),
            }
            
            if type_declaration == 'journaliere':
                form_data['date_recette'] = request.POST.get('date_recette')
            else:
                form_data['date_debut_decade'] = request.POST.get('date_debut_decade')
                form_data['date_fin_decade'] = request.POST.get('date_fin_decade')
            
            # Gérer l'image uploadée
            has_image = False
            image_name = None
            if 'image_quittance' in request.FILES:
                from django.core.files.storage import default_storage
                image_file = request.FILES['image_quittance']
                temp_name = f"temp_pesage_{user.id}_{timezone.now().timestamp()}_{image_file.name}"
                temp_path = os.path.join('temp_quittances_pesage', temp_name)
                saved_path = default_storage.save(temp_path, image_file)
                request.session['image_pesage_temp_path'] = saved_path
                has_image = True
                image_name = image_file.name
            
            form_data['has_image'] = has_image
            form_data['image_name'] = image_name
            
            # VALIDATION COMPLÈTE
            errors = []
            
            # Validation numéro de quittance
            if not form_data['numero_quittance']:
                errors.append("Le numéro de quittance est obligatoire")
            else:
                numero = form_data['numero_quittance'].strip().upper()
                
                if QuittancementPesage.objects.filter(numero_quittance__iexact=numero).exists():
                    errors.append(f"Le numéro de quittance '{numero}' existe déjà dans les quittancements pesage")
                
                try:
                    from inventaire.models import Quittancement
                    if Quittancement.objects.filter(numero_quittance__iexact=numero).exists():
                        errors.append(f"Le numéro de quittance '{numero}' existe déjà dans les quittancements péage")
                except ImportError:
                    pass
            
            if not form_data['date_quittancement']:
                errors.append("La date de quittancement est obligatoire")
            if not form_data['montant']:
                errors.append("Le montant est obligatoire")
            if not has_image:
                errors.append("L'image de la quittance est obligatoire")
            
            if type_declaration == 'journaliere':
                if not form_data.get('date_recette'):
                    errors.append("La date de recette est obligatoire")
            else:
                if not form_data.get('date_debut_decade'):
                    errors.append("La date de début de décade est obligatoire")
                if not form_data.get('date_fin_decade'):
                    errors.append("La date de fin de décade est obligatoire")
                elif form_data.get('date_debut_decade') and form_data.get('date_fin_decade'):
                    if form_data['date_debut_decade'] > form_data['date_fin_decade']:
                        errors.append("La date de début doit être antérieure à la date de fin")
            
            # Validation des dates déjà quittancées
            if not errors:
                try:
                    from inventaire.utils_quittancement import valider_dates_quittancement_pesage
                    
                    if type_declaration == 'journaliere':
                        is_valid, error_msg, _ = valider_dates_quittancement_pesage(
                            station=station,
                            exercice=int(exercice),
                            mois=mois,
                            type_declaration='journaliere',
                            date_recette=form_data.get('date_recette')
                        )
                    else:
                        is_valid, error_msg, _ = valider_dates_quittancement_pesage(
                            station=station,
                            exercice=int(exercice),
                            mois=mois,
                            type_declaration='decade',
                            date_debut=form_data.get('date_debut_decade'),
                            date_fin=form_data.get('date_fin_decade')
                        )
                    
                    if not is_valid:
                        errors.append(error_msg)
                        
                except Exception as e:
                    logger.error(f"Erreur validation dates pesage: {e}")
            
            if errors:
                for error in errors:
                    messages.error(request, f"❌ {error}")
                
                if has_image and request.session.get('image_pesage_temp_path'):
                    from django.core.files.storage import default_storage
                    temp_path = request.session.get('image_pesage_temp_path')
                    if default_storage.exists(temp_path):
                        default_storage.delete(temp_path)
                    request.session.pop('image_pesage_temp_path', None)
                
                context = {
                    'etape': 2, 
                    'station': station, 
                    'exercice': exercice, 
                    'mois': mois,
                    'type_declaration': type_declaration, 
                    'form_data': form_data, 
                    'is_admin': is_admin_user(user),
                }
                return render(request, 'pesage/saisie_quittancement_pesage.html', context)
            
            request.session['form_data_pesage_temp'] = form_data
            return redirect(f"{reverse('inventaire:saisie_quittancement_pesage')}?etape=3&station={station.id}")
        
        context = {
            'etape': 2, 
            'station': station, 
            'exercice': exercice, 
            'mois': mois,
            'type_declaration': type_declaration, 
            'is_admin': is_admin_user(user),
        }
        return render(request, 'pesage/saisie_quittancement_pesage.html', context)
    
    # ÉTAPE 3 : Confirmation et enregistrement
    elif etape == 3:
        exercice = request.session.get('exercice_pesage_temp')
        mois = request.session.get('mois_pesage_temp')
        type_declaration = request.session.get('type_declaration_pesage_temp')
        form_data = request.session.get('form_data_pesage_temp')
        image_temp_path = request.session.get('image_pesage_temp_path')
        
        if not all([exercice, mois, type_declaration, form_data]):
            messages.error(request, "❌ Session expirée")
            return redirect(f"{reverse('inventaire:saisie_quittancement_pesage')}?station={station.id}")
        
        if request.method == 'POST':
            action = request.POST.get('action')
            
            if action == 'confirmer':
                try:
                    quittancement = QuittancementPesage(
                        numero_quittance=form_data['numero_quittance'],
                        station=station,
                        exercice=int(exercice),
                        mois=mois,
                        type_declaration=type_declaration,
                        date_quittancement=form_data['date_quittancement'],
                        montant_quittance=Decimal(form_data['montant']),
                        observations=form_data.get('observations', ''),
                        saisi_par=user,
                    )
                    
                    if type_declaration == 'journaliere':
                        quittancement.date_recette = form_data['date_recette']
                    else:
                        quittancement.date_debut_decade = form_data['date_debut_decade']
                        quittancement.date_fin_decade = form_data['date_fin_decade']
                    
                    if image_temp_path:
                        from django.core.files.storage import default_storage
                        from django.core.files import File
                        if default_storage.exists(image_temp_path):
                            temp_file = default_storage.open(image_temp_path, 'rb')
                            original_name = os.path.basename(image_temp_path).split('_', 4)[-1]
                            final_name = f"quittances_pesage/{quittancement.exercice}/{quittancement.mois}/{original_name}"
                            quittancement.image_quittance.save(final_name, File(temp_file), save=False)
                            temp_file.close()
                            default_storage.delete(image_temp_path)
                    
                    quittancement.save()
                    
                    log_user_action(
                        user, 
                        "Création quittancement pesage",
                        f"N°{quittancement.numero_quittance} | Station: {station.nom} | "
                        f"Montant: {quittancement.montant_quittance} FCFA | "
                        f"Exercice: {exercice} | Mois: {mois} | Type: {type_declaration}",
                        request
                    )
                    
                    # Nettoyer la session
                    for key in ['exercice_pesage_temp', 'mois_pesage_temp', 'type_declaration_pesage_temp', 
                               'form_data_pesage_temp', 'image_pesage_temp_path']:
                        request.session.pop(key, None)
                    
                    messages.success(request, f"✅ Quittancement {quittancement.numero_quittance} enregistré avec succès")
                    return redirect('inventaire:liste_quittancements_pesage')
                
                except ValidationError as e:
                    messages.error(request, f"❌ Erreur de validation : {e}")
                except Exception as e:
                    messages.error(request, f"❌ Erreur : {str(e)}")
                    logger.error(f"Erreur création quittancement pesage : {str(e)}", exc_info=True)
            
            elif action == 'retour':
                return redirect(f"{reverse('inventaire:saisie_quittancement_pesage')}?etape=2&station={station.id}")
            
            else:
                # Annulation
                if image_temp_path:
                    from django.core.files.storage import default_storage
                    if default_storage.exists(image_temp_path):
                        default_storage.delete(image_temp_path)
                
                for key in ['exercice_pesage_temp', 'mois_pesage_temp', 'type_declaration_pesage_temp', 
                           'form_data_pesage_temp', 'image_pesage_temp_path']:
                    request.session.pop(key, None)
                
                messages.info(request, "Saisie annulée")
                return redirect('inventaire:liste_quittancements_pesage')
        
        # Affichage de la confirmation
        if type_declaration == 'journaliere':
            periode = f"Jour : {form_data['date_recette']}"
        else:
            periode = f"Décade : du {form_data['date_debut_decade']} au {form_data['date_fin_decade']}"
        
        context = {
            'etape': 3, 
            'station': station, 
            'exercice': exercice, 
            'mois': mois,
            'type_declaration': type_declaration, 
            'form_data': form_data, 
            'periode': periode,
            'has_image': form_data.get('has_image', False), 
            'image_name': form_data.get('image_name'),
            'is_admin': is_admin_user(user),
        }
        return render(request, 'pesage/saisie_quittancement_pesage.html', context)


# ===================================================================
# LISTE DES QUITTANCEMENTS
# Permission: peut_voir_liste_quittancements_pesage
# ===================================================================

@login_required
def liste_quittancements_pesage(request):
    """
    Liste des quittancements pesage avec filtres.
    Permission: peut_voir_liste_quittancements_pesage
    """
    user = request.user
    
    # Vérifier la permission
    if not has_permission(user, 'peut_voir_liste_quittancements_pesage'):
        log_user_action(
            user,
            "Liste quittancements - permission refusée",
            "Permission peut_voir_liste_quittancements_pesage manquante",
            request
        )
        messages.error(request, _("Accès non autorisé."))
        return redirect('inventaire:liste_amendes')
    
    # Construction du queryset selon les droits
    if user_has_acces_tous_postes(user):
        quittancements = QuittancementPesage.objects.all()
        stations = Poste.objects.filter(type='pesage', is_active=True).order_by('nom')
    else:
        poste = getattr(user, 'poste_affectation', None)
        if poste and poste.type == 'pesage':
            quittancements = QuittancementPesage.objects.filter(station=poste)
            stations = None
        else:
            quittancements = QuittancementPesage.objects.none()
            stations = None

    # Filtres
    station_id = request.GET.get('station')
    exercice = request.GET.get('exercice')
    type_declaration = request.GET.get('type_declaration')
    mois = request.GET.get('mois')

    if station_id: 
        quittancements = quittancements.filter(station_id=station_id)
    if exercice: 
        quittancements = quittancements.filter(exercice=exercice)
    if type_declaration: 
        quittancements = quittancements.filter(type_declaration=type_declaration)
    if mois: 
        quittancements = quittancements.filter(mois=mois)

    quittancements = quittancements.select_related('station', 'saisi_par').order_by('-date_quittancement', '-id')

    paginator = Paginator(quittancements, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_montant = quittancements.aggregate(Sum('montant_quittance'))['montant_quittance__sum'] or Decimal('0')

    current_year = datetime.now().year
    years_range = list(range(current_year, current_year - 5, -1))

    context = {
        'page_obj': page_obj, 
        'quittancements': page_obj, 
        'total_montant': total_montant,
        'nombre_quittancements': quittancements.count(), 
        'stations': stations,
        'exercice_courant': current_year, 
        'years_range': years_range,
        'types_declaration': [('journaliere', 'Journalière'), ('decade', 'Par décade')],
        'is_admin': is_admin_user(user), 
        'title': _('Liste des quittancements'),
        'peut_saisir_quittance': has_permission(user, 'peut_saisir_quittance_pesage'),
    }
    return render(request, 'pesage/liste_quittancements_pesage.html', context)


# ===================================================================
# COMPTABILISATION DES QUITTANCEMENTS
# Permission: peut_comptabiliser_quittances_pesage
# ===================================================================

@login_required
def comptabilisation_quittancements_pesage(request):
    """
    Comptabilisation des quittancements pesage.
    Permission: peut_comptabiliser_quittances_pesage
    """
    user = request.user
    
    # Vérifier la permission
    if not has_permission(user, 'peut_comptabiliser_quittances_pesage'):
        log_user_action(
            user,
            "Comptabilisation quittancements - permission refusée",
            "Permission peut_comptabiliser_quittances_pesage manquante",
            request
        )
        messages.error(request, _("Accès non autorisé."))
        return redirect('inventaire:liste_quittancements_pesage')
    
    annee_courante = timezone.now().year
    mois_courant = timezone.now().month

    periode = request.GET.get('periode', 'mois')
    station_id = request.GET.get('station')
    annee = int(request.GET.get('annee', annee_courante))
    mois = request.GET.get('mois')

    # Calcul des dates selon la période
    if periode == 'mois':
        if mois:
            try:
                annee_mois = mois.split('-')
                mois_int = int(annee_mois[1])
                annee = int(annee_mois[0])
            except:
                mois_int = mois_courant
        else:
            mois_int = mois_courant
        date_debut = date(annee, mois_int, 1)
        dernier_jour = calendar.monthrange(annee, mois_int)[1]
        date_fin = date(annee, mois_int, dernier_jour)
    elif periode == 'trimestre':
        trimestre = int(request.GET.get('trimestre', ((mois_courant - 1) // 3) + 1))
        mois_debut = (trimestre - 1) * 3 + 1
        mois_fin = trimestre * 3
        date_debut = date(annee, mois_debut, 1)
        dernier_jour = calendar.monthrange(annee, mois_fin)[1]
        date_fin = date(annee, mois_fin, dernier_jour)
    elif periode == 'semestre':
        semestre = int(request.GET.get('semestre', 1 if mois_courant <= 6 else 2))
        if semestre == 1:
            date_debut = date(annee, 1, 1)
            date_fin = date(annee, 6, 30)
        else:
            date_debut = date(annee, 7, 1)
            date_fin = date(annee, 12, 31)
    else:
        date_debut = date(annee, 1, 1)
        date_fin = date(annee, 12, 31)

    # Sélection des stations
    if station_id:
        stations = Poste.objects.filter(id=station_id, type='pesage', is_active=True)
    else:
        stations = Poste.objects.filter(type='pesage', is_active=True)
    
    stations_filtre = Poste.objects.filter(type='pesage', is_active=True).order_by('nom')

    # Calcul des résultats par station
    resultats = []
    total_attendu_global = Decimal('0')
    total_quittance_global = Decimal('0')

    for station in stations:
        amendes_payees = AmendeEmise.objects.filter(
            station=station, 
            statut='paye',
            date_paiement__date__gte=date_debut, 
            date_paiement__date__lte=date_fin
        )
        total_attendu = amendes_payees.aggregate(Sum('montant_amende'))['montant_amende__sum'] or Decimal('0')
        
        quittancements_journaliers = QuittancementPesage.objects.filter(
            station=station, 
            type_declaration='journaliere', 
            date_recette__range=[date_debut, date_fin]
        ).aggregate(Sum('montant_quittance'))['montant_quittance__sum'] or Decimal('0')
        
        quittancements_decades = QuittancementPesage.objects.filter(
            station=station, 
            type_declaration='decade'
        ).filter(
            Q(date_debut_decade__lte=date_fin) & Q(date_fin_decade__gte=date_debut)
        ).aggregate(Sum('montant_quittance'))['montant_quittance__sum'] or Decimal('0')
        
        total_quittance = quittancements_journaliers + quittancements_decades
        ecart = total_quittance - total_attendu
        ecart_pourcentage = (ecart / total_attendu * 100) if total_attendu > 0 else Decimal('0')
        
        if abs(ecart) < 1:
            statut, statut_label, statut_class = 'conforme', 'Conforme', 'success'
        else:
            justification_existe = JustificationEcartPesage.objects.filter(
                station=station, 
                date_debut=date_debut, 
                date_fin=date_fin
            ).exists()
            if justification_existe:
                statut, statut_label, statut_class = 'justifie', 'Justifié', 'info'
            else:
                statut, statut_label, statut_class = 'ecart', 'Non justifié', 'danger'
        
        resultats.append({
            'station': station, 
            'station_id': station.id, 
            'station_nom': station.nom,
            'total_quittance': total_quittance, 
            'total_attendu': total_attendu,
            'ecart': ecart, 
            'ecart_pourcentage': ecart_pourcentage,
            'statut': statut, 
            'statut_label': statut_label, 
            'statut_class': statut_class
        })
        total_attendu_global += total_attendu
        total_quittance_global += total_quittance

    ecart_global = total_quittance_global - total_attendu_global
    statistiques = {
        'total_attendu': total_attendu_global, 
        'total_quittance': total_quittance_global,
        'ecart_total': ecart_global, 
        'nombre_stations': len(resultats),
        'nombre_conformes': len([r for r in resultats if r['statut'] == 'conforme']),
        'nombre_justifies': len([r for r in resultats if r['statut'] == 'justifie']),
        'nombre_ecarts': len([r for r in resultats if r['statut'] == 'ecart']),
    }
    
    annees_disponibles = list(range(annee_courante - 5, annee_courante + 1))
    
    context = {
        'resultats': resultats, 
        'statistiques': statistiques, 
        'periode': periode,
        'annee': annee, 
        'annees_disponibles': annees_disponibles,
        'mois': f"{annee}-{mois_int:02d}" if periode == 'mois' else None,
        'date_debut': date_debut, 
        'date_fin': date_fin,
        'stations': stations_filtre, 
        'station_selectionne': int(station_id) if station_id else None,
        'is_admin': is_admin_user(user), 
        'title': _('Comptabilisation Quittancements'),
    }
    return render(request, 'pesage/comptabilisation_quittancements_pesage.html', context)


# ===================================================================
# JUSTIFICATION ÉCART PESAGE
#
# AVANT: @quittancement_pesage_required (regisseur + admin)
# APRÈS: Permission 'peut_comptabiliser_quittances_pesage'
# ===================================================================

@login_required
def justifier_ecart_pesage(request, station_id, date_debut, date_fin):
    """
    Justifier un écart entre quittancements et amendes payées.
    
    PERMISSION: peut_comptabiliser_quittances_pesage
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_comptabiliser_quittances_pesage'):
        logger.warning(
            f"Accès refusé justifier_ecart_pesage | User: {user.username} | "
            f"Permission peut_comptabiliser_quittances_pesage: {getattr(user, 'peut_comptabiliser_quittances_pesage', False)}"
        )
        messages.error(request, _("Vous n'avez pas la permission de justifier les écarts."))
        return redirect('common:dashboard')
    
    station = get_object_or_404(Poste, id=station_id, type='pesage')
    
    # Parser les dates
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except ValueError as e:
        messages.error(request, f"Format de date invalide : {str(e)}")
        return redirect('inventaire:comptabilisation_quittancements_pesage')
    
    # Calculer les totaux
    total_quittance = QuittancementPesage.objects.filter(
        station=station
    ).filter(
        Q(type_declaration='journaliere', date_recette__range=[date_debut_obj, date_fin_obj]) |
        Q(type_declaration='decade', date_debut_decade__lte=date_fin_obj, date_fin_decade__gte=date_debut_obj)
    ).aggregate(Sum('montant_quittance'))['montant_quittance__sum'] or Decimal('0')
    
    total_attendu = AmendeEmise.objects.filter(
        station=station,
        statut='paye',
        date_paiement__date__gte=date_debut_obj,
        date_paiement__date__lte=date_fin_obj
    ).aggregate(Sum('montant_amende'))['montant_amende__sum'] or Decimal('0')
    
    ecart = total_quittance - total_attendu
    ecart_pourcentage = (ecart / total_attendu * 100) if total_attendu > 0 else 0
    
    # Vérifier si une justification existe déjà
    justification_existante = JustificationEcartPesage.objects.filter(
        station=station,
        date_debut=date_debut_obj,
        date_fin=date_fin_obj
    ).select_related('justifie_par').first()
    
    if request.method == 'POST':
        justification_texte = request.POST.get('justification', '').strip()
        
        if len(justification_texte) < 20:
            messages.error(request, "❌ La justification doit contenir au moins 20 caractères.")
        else:
            try:
                if justification_existante:
                    # Mise à jour
                    justification_existante.justification = justification_texte
                    justification_existante.justifie_par = user
                    justification_existante.date_justification = timezone.now()
                    justification_existante.montant_quittance = total_quittance
                    justification_existante.montant_attendu = total_attendu
                    justification_existante.ecart = ecart
                    justification_existante.save()
                else:
                    # Création
                    JustificationEcartPesage.objects.create(
                        station=station,
                        date_debut=date_debut_obj,
                        date_fin=date_fin_obj,
                        montant_quittance=total_quittance,
                        montant_attendu=total_attendu,
                        ecart=ecart,
                        justification=justification_texte,
                        justifie_par=user
                    )
                
                log_user_action(
                    user,
                    "Justification écart quittancement pesage",
                    f"Station: {station.nom} | Période: {date_debut_obj} - {date_fin_obj} | "
                    f"Écart: {ecart} FCFA",
                    request
                )
                
                logger.info(
                    f"Justification écart pesage | User: {user.username} | "
                    f"Station: {station.nom} | Période: {date_debut_obj} - {date_fin_obj} | "
                    f"Écart: {ecart} FCFA"
                )
                
                messages.success(request, "✅ Justification enregistrée avec succès.")
                return redirect('inventaire:comptabilisation_quittancements_pesage')
            except Exception as e:
                logger.error(f"Erreur justification écart pesage: {e}", exc_info=True)
                messages.error(request, f"❌ Erreur lors de l'enregistrement : {str(e)}")
    
    context = {
        'station': station,
        'date_debut': date_debut_obj,
        'date_fin': date_fin_obj,
        'total_quittance': total_quittance,
        'total_attendu': total_attendu,
        'ecart': ecart,
        'ecart_pourcentage': ecart_pourcentage,
        'justification_existante': justification_existante,
        'is_admin': is_admin_user(user),
        'title': _('Justifier Écart'),
    }
    return render(request, 'pesage/justifier_ecart_pesage.html', context)


# ===================================================================
# DÉTAILS QUITTANCEMENTS PÉRIODE
#
# AVANT: regisseur, chef_station + admin
# APRÈS: Permission 'peut_voir_liste_quittancements_pesage'
# ===================================================================

@login_required
def detail_quittancements_periode_pesage(request, station_id, date_debut, date_fin):
    """
    Affiche les détails des quittancements et amendes payées pour une période.
    
    PERMISSION: peut_voir_liste_quittancements_pesage
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_voir_liste_quittancements_pesage'):
        logger.warning(
            f"Accès refusé detail_quittancements_periode_pesage | User: {user.username} | "
            f"Permission peut_voir_liste_quittancements_pesage: {getattr(user, 'peut_voir_liste_quittancements_pesage', False)}"
        )
        messages.error(request, _("Accès non autorisé."))
        return redirect('inventaire:liste_amendes')
    
    station = get_object_or_404(Poste, id=station_id, type='pesage')
    
    # Parser les dates
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Format de date invalide")
        return redirect('inventaire:comptabilisation_quittancements_pesage')
    
    # Récupérer les quittancements de la période
    quittancements = QuittancementPesage.objects.filter(station=station).filter(
        Q(type_declaration='journaliere', date_recette__range=[date_debut_obj, date_fin_obj]) |
        Q(type_declaration='decade', date_debut_decade__lte=date_fin_obj, date_fin_decade__gte=date_debut_obj)
    ).select_related('saisi_par').order_by('date_quittancement')
    
    # Récupérer les amendes payées de la période
    amendes_payees = AmendeEmise.objects.filter(
        station=station,
        statut='paye',
        date_paiement__date__gte=date_debut_obj,
        date_paiement__date__lte=date_fin_obj
    ).select_related('saisi_par', 'valide_par').order_by('date_paiement')
    
    # Totaux
    total_quittance = quittancements.aggregate(Sum('montant_quittance'))['montant_quittance__sum'] or Decimal('0')
    total_amendes = amendes_payees.aggregate(Sum('montant_amende'))['montant_amende__sum'] or Decimal('0')
    ecart = total_quittance - total_amendes
    
    # Vérifier si une justification existe
    justification = JustificationEcartPesage.objects.filter(
        station=station,
        date_debut=date_debut_obj,
        date_fin=date_fin_obj
    ).first()
    
    logger.info(
        f"Détail quittancements période | User: {user.username} | "
        f"Station: {station.nom} | Période: {date_debut_obj} - {date_fin_obj}"
    )
    
    context = {
        'station': station,
        'date_debut': date_debut_obj,
        'date_fin': date_fin_obj,
        'quittancements': quittancements,
        'amendes_payees': amendes_payees,
        'total_quittance': total_quittance,
        'total_amendes': total_amendes,
        'ecart': ecart,
        'justification': justification,
        'is_admin': is_admin_user(user),
        'title': _(f'Détail Quittancements - {station.nom}'),
    }
    return render(request, 'pesage/detail_quittancements_periode_pesage.html', context)


# ===================================================================
# EXPORT EXCEL QUITTANCEMENTS
#
# AVANT: @quittancement_pesage_required (regisseur + admin)
# APRÈS: Permission 'peut_comptabiliser_quittances_pesage'
# ===================================================================

@login_required
def export_quittancements_pesage(request):
    """
    Export Excel des quittancements pesage.
    
    PERMISSION: peut_comptabiliser_quittances_pesage
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_comptabiliser_quittances_pesage'):
        logger.warning(
            f"Accès refusé export_quittancements_pesage | User: {user.username} | "
            f"Permission peut_comptabiliser_quittances_pesage: {getattr(user, 'peut_comptabiliser_quittances_pesage', False)}"
        )
        messages.error(request, _("Vous n'avez pas la permission d'exporter les quittancements."))
        return redirect('inventaire:liste_quittancements_pesage')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        messages.error(request, _("Module openpyxl non installé."))
        return redirect('inventaire:liste_quittancements_pesage')
    
    # Filtres
    station_id = request.GET.get('station')
    exercice = request.GET.get('exercice')
    mois = request.GET.get('mois')
    
    # Construire le queryset
    quittancements = QuittancementPesage.objects.all()
    if station_id:
        quittancements = quittancements.filter(station_id=station_id)
    if exercice:
        quittancements = quittancements.filter(exercice=exercice)
    if mois:
        quittancements = quittancements.filter(mois=mois)
    
    quittancements = quittancements.select_related('station', 'saisi_par').order_by('-date_quittancement')
    
    # Créer le workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quittancements Pesage"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ['N° Quittance', 'Station', 'Exercice', 'Mois', 'Type', 'Date Quittancement',
               'Période', 'Montant Quittancé', 'Saisi par']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Données
    for row, q in enumerate(quittancements, 2):
        if q.type_declaration == 'journaliere':
            periode = q.date_recette.strftime('%d/%m/%Y') if q.date_recette else ''
        else:
            if q.date_debut_decade and q.date_fin_decade:
                periode = f"{q.date_debut_decade.strftime('%d/%m/%Y')} - {q.date_fin_decade.strftime('%d/%m/%Y')}"
            else:
                periode = ''
        
        ws.cell(row=row, column=1, value=q.numero_quittance).border = thin_border
        ws.cell(row=row, column=2, value=q.station.nom).border = thin_border
        ws.cell(row=row, column=3, value=q.exercice).border = thin_border
        ws.cell(row=row, column=4, value=q.mois).border = thin_border
        ws.cell(row=row, column=5, value=q.get_type_declaration_display()).border = thin_border
        ws.cell(row=row, column=6, value=q.date_quittancement.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(row=row, column=7, value=periode).border = thin_border
        ws.cell(row=row, column=8, value=float(q.montant_quittance)).border = thin_border
        ws.cell(row=row, column=9, value=q.saisi_par.nom_complet if q.saisi_par else '').border = thin_border
    
    # Ajuster les largeurs de colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # Préparer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=quittancements_pesage_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    log_user_action(
        user,
        "Export Excel quittancements pesage",
        f"Nombre: {quittancements.count()}",
        request
    )
    
    logger.info(
        f"Export quittancements pesage | User: {user.username} | "
        f"Nombre: {quittancements.count()}"
    )
    
    return response


# ===================================================================
# STATISTIQUES PESAGE
# Permission: peut_voir_stats_pesage
# ===================================================================

@login_required
def statistiques_pesage(request):
    """
    Statistiques détaillées du pesage.
    Permission: peut_voir_stats_pesage
    """
    user = request.user
    
    # Vérifier la permission
    if not has_permission(user, 'peut_voir_stats_pesage'):
        log_user_action(
            user,
            "Statistiques pesage - permission refusée",
            "Permission peut_voir_stats_pesage manquante",
            request
        )
        messages.error(request, _("Vous n'avez pas la permission de voir les statistiques."))
        return redirect('common:dashboard')
    
    stations_accessibles = get_stations_pesage_accessibles(user)
    today = timezone.now().date()
    
    station, redirect_response = get_station_context(request, user)
    if redirect_response:
        return redirect_response
    
    periode = request.GET.get('periode', 'mois')
    date_debut, date_fin = get_periode_dates(periode, today)
    
    stats = {
        'emissions': 0, 
        'hors_gabarit': 0, 
        'montant_emis': Decimal('0'),
        'montant_recouvre': Decimal('0'), 
        'reste_a_recouvrer': Decimal('0'),
        'nombre_pesees': 0, 
        'taux_recouvrement': 0,
    }
    
    # Construction du queryset
    if station:
        amendes = AmendeEmise.objects.filter(
            station=station, 
            date_heure_emission__date__gte=date_debut, 
            date_heure_emission__date__lte=date_fin
        )
    elif user_has_acces_tous_postes(user):
        amendes = AmendeEmise.objects.filter(
            date_heure_emission__date__gte=date_debut, 
            date_heure_emission__date__lte=date_fin
        )
    else:
        amendes = AmendeEmise.objects.none()
    
    if amendes.exists():
        agg = amendes.aggregate(
            count=Count('id'), 
            hg_count=Count('id', filter=Q(est_hors_gabarit=True)),
            montant_emis=Sum('montant_amende'), 
            montant_recouvre=Sum('montant_amende', filter=Q(statut='paye'))
        )
        stats['emissions'] = agg['count'] or 0
        stats['hors_gabarit'] = agg['hg_count'] or 0
        stats['montant_emis'] = agg['montant_emis'] or Decimal('0')
        stats['montant_recouvre'] = agg['montant_recouvre'] or Decimal('0')
        stats['reste_a_recouvrer'] = stats['montant_emis'] - stats['montant_recouvre']
        if stats['montant_emis'] > 0:
            stats['taux_recouvrement'] = round(
                (float(stats['montant_recouvre']) / float(stats['montant_emis'])) * 100, 2)
    
    # Pesées
    if station:
        pesees_qs = PeseesJournalieres.objects.filter(
            station=station, 
            date__gte=date_debut, 
            date__lte=date_fin
        )
    elif user_has_acces_tous_postes(user):
        pesees_qs = PeseesJournalieres.objects.filter(
            date__gte=date_debut, 
            date__lte=date_fin
        )
    else:
        pesees_qs = PeseesJournalieres.objects.none()
    
    stats['nombre_pesees'] = pesees_qs.aggregate(total=Sum('nombre_pesees'))['total'] or 0
    
    context = {
        'stats': stats, 
        'station': station, 
        'stations_accessibles': stations_accessibles,
        'periode': periode, 
        'date_debut': date_debut, 
        'date_fin': date_fin,
        'is_admin': is_admin_user(user), 
        'title': _('Statistiques Pesage'),
    }
    return render(request, 'pesage/statistiques_pesage.html', context)


# ===================================================================
# RECETTES PESAGE
# Permission: peut_voir_recettes_pesage
# ===================================================================

@login_required
def recettes_pesage(request):
    """
    Suivi des recettes pesage (amendes payées).
    Permission: peut_voir_recettes_pesage
    """
    user = request.user
    
    # Vérifier la permission
    if not has_permission(user, 'peut_voir_recettes_pesage'):
        log_user_action(
            user,
            "Recettes pesage - permission refusée",
            "Permission peut_voir_recettes_pesage manquante",
            request
        )
        messages.error(request, _("Accès non autorisé."))
        return redirect('inventaire:liste_amendes')
    
    stations_accessibles = get_stations_pesage_accessibles(user)
    station, redirect_response = get_station_context(request, user)
    if redirect_response:
        return redirect_response
    
    periode = request.GET.get('periode', 'mois')
    date_debut_param = request.GET.get('date_debut', '')
    date_fin_param = request.GET.get('date_fin', '')
    station_filter = request.GET.get('station_filter', '')
    
    periodes_predefinies = ['jour', 'semaine', 'mois', 'trimestre', 'annee']
    
    if periode in periodes_predefinies:
        periode_info = calculer_periode_9h(periode)
    elif periode == 'personnalise' and date_debut_param and date_fin_param:
        periode_info = calculer_periode_9h('personnalise', date_debut_param, date_fin_param)
    else:
        periode = 'mois'
        periode_info = calculer_periode_9h('mois')
    
    datetime_debut = periode_info['datetime_debut']
    datetime_fin = periode_info['datetime_fin']
    date_debut = periode_info['date_debut']
    date_fin = periode_info['date_fin']
    
    # Construction du queryset
    if station is None and user_has_acces_tous_postes(user):
        queryset = AmendeEmise.objects.filter(statut='paye')
        if station_filter: 
            queryset = queryset.filter(station_id=station_filter)
    elif station:
        queryset = AmendeEmise.objects.filter(station=station, statut='paye')
        station_filter = ''
    else:
        queryset = AmendeEmise.objects.none()
        station_filter = ''
    
    queryset = queryset.filter(
        date_paiement__gte=datetime_debut, 
        date_paiement__lt=datetime_fin
    )
    
    stats = queryset.aggregate(
        total_recouvre=Sum('montant_amende'), 
        count_paiements=Count('id')
    )
    
    recettes_par_jour = queryset.annotate(
        date_jour=TruncDate('date_paiement')
    ).values('date_jour').annotate(
        total=Sum('montant_amende'), 
        count_paiements=Count('id')
    ).order_by('-date_jour')
    
    paginator = Paginator(list(recettes_par_jour), 15)
    page = request.GET.get('page')
    recettes = paginator.get_page(page)
    
    context = {
        'recettes': recettes, 
        'station': station, 
        'stations_accessibles': stations_accessibles,
        'stats': stats, 
        'is_admin': is_admin_user(user), 
        'periode': periode,
        'date_debut': date_debut, 
        'date_fin': date_fin, 
        'station_filter': station_filter,
        'title': _('Recettes Pesage'),
    }
    return render(request, 'pesage/recettes_pesage.html', context)


# ===================================================================
# IMPRIMER RECETTE JOUR
#
# AVANT: @pesage_access_required (PESAGE_ROLES + admin)
# APRÈS: Permission 'peut_voir_recettes_pesage'
# ===================================================================

@login_required
def imprimer_recette_jour(request, date_str):
    """
    Page d'impression pour les recettes d'un jour.
    
    PERMISSION: peut_voir_recettes_pesage
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_voir_recettes_pesage'):
        logger.warning(
            f"Accès refusé imprimer_recette_jour | User: {user.username} | "
            f"Permission peut_voir_recettes_pesage: {getattr(user, 'peut_voir_recettes_pesage', False)}"
        )
        messages.error(request, _("Accès non autorisé."))
        return redirect('common:dashboard')
    
    # Récupérer le contexte de station
    station, redirect_response = get_station_context(request, user)
    if redirect_response:
        return redirect_response
    
    # Parser la date
    try:
        date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, _("Format de date invalide."))
        return redirect('inventaire:recettes_pesage')
    
    # Calculer datetime 9h-9h pour ce jour
    datetime_debut = CAMEROUN_TZ.localize(datetime.combine(date_cible, HEURE_DEBUT_JOURNEE))
    datetime_fin = CAMEROUN_TZ.localize(datetime.combine(date_cible + timedelta(days=1), HEURE_FIN_JOURNEE))
    
    # Récupérer les paiements du jour
    if station is None and user_has_acces_tous_postes(user):
        queryset = AmendeEmise.objects.filter(statut='paye')
        station_filter = request.GET.get('station')
        if station_filter:
            queryset = queryset.filter(station_id=station_filter)
            try:
                station_obj = Poste.objects.get(pk=station_filter)
            except Poste.DoesNotExist:
                station_obj = None
        else:
            station_obj = None
    elif station:
        queryset = AmendeEmise.objects.filter(station=station, statut='paye')
        station_obj = station
    else:
        queryset = AmendeEmise.objects.none()
        station_obj = None
    
    paiements = queryset.filter(
        date_paiement__gte=datetime_debut,
        date_paiement__lte=datetime_fin
    ).select_related('station', 'valide_par', 'saisi_par').order_by('date_paiement')
    
    total = paiements.aggregate(total=Sum('montant_amende'))['total'] or 0
    
    # Statistiques par type d'infraction
    stats_types = []
    
    surcharge_only = paiements.filter(est_surcharge=True, est_hors_gabarit=False)
    if surcharge_only.exists():
        stats_types.append({
            'type': 'Surcharge (S)',
            'count': surcharge_only.count(),
            'montant': surcharge_only.aggregate(m=Sum('montant_amende'))['m'] or 0
        })
    
    hg_only = paiements.filter(est_hors_gabarit=True, est_surcharge=False)
    if hg_only.exists():
        stats_types.append({
            'type': 'Hors Gabarit (HG)',
            'count': hg_only.count(),
            'montant': hg_only.aggregate(m=Sum('montant_amende'))['m'] or 0
        })
    
    both = paiements.filter(est_surcharge=True, est_hors_gabarit=True)
    if both.exists():
        stats_types.append({
            'type': 'Surcharge + Hors Gabarit (S+HG)',
            'count': both.count(),
            'montant': both.aggregate(m=Sum('montant_amende'))['m'] or 0
        })
    
    logger.info(
        f"Impression recette jour | User: {user.username} | "
        f"Date: {date_cible} | Station: {station_obj.nom if station_obj else 'Toutes'} | "
        f"Paiements: {paiements.count()} | Total: {total} FCFA"
    )
    
    context = {
        'paiements': paiements,
        'date': date_cible,
        'datetime_debut': datetime_debut,
        'datetime_fin': datetime_fin,
        'station': station_obj,
        'total': total,
        'count': paiements.count(),
        'stats_types': stats_types,
        'is_admin': is_admin_user(user),
        'title': f"Recettes du {date_cible.strftime('%d/%m/%Y')} (9h-9h)",
    }
    
    return render(request, 'pesage/imprimer_recette_jour.html', context)


# ===================================================================
# API ENDPOINTS (JSON)
# Utilisent les décorateurs API pour retourner des réponses JSON
# ===================================================================

@login_required
@require_GET
def api_check_pesees(request):
    """
    API pour vérifier si les pesées ont déjà été saisies.
    GET params: date (YYYY-MM-DD), station (ID)
    """
    date_str = request.GET.get('date')
    station_id = request.GET.get('station')
    
    if not date_str or not station_id:
        return JsonResponse({'error': 'Paramètres date et station requis'}, status=400)
    
    try:
        date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Format de date invalide'}, status=400)
    
    try:
        station = Poste.objects.get(pk=station_id, type='pesage')
    except Poste.DoesNotExist:
        return JsonResponse({'error': 'Station non trouvée'}, status=404)
    
    # Vérifier l'accès à la station
    if not check_poste_access(request.user, station):
        return JsonResponse({'error': 'Accès non autorisé à cette station'}, status=403)
    
    pesee = PeseesJournalieres.objects.filter(station=station, date=date_cible).first()
    
    if pesee:
        return JsonResponse({
            'deja_saisi': True,
            'nombre_pesees': pesee.nombre_pesees,
            'date': date_cible.strftime('%d/%m/%Y'),
            'saisi_par': pesee.saisi_par.nom_complet if pesee.saisi_par else None
        })
    
    return JsonResponse({
        'deja_saisi': False, 
        'date': date_cible.strftime('%d/%m/%Y')
    })


@login_required
@require_GET
def api_montant_attendu(request):
    """
    API pour calculer le montant attendu pour un quittancement.
    GET params: station, date_debut, date_fin (YYYY-MM-DD)
    """
    station_id = request.GET.get('station')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    if not all([station_id, date_debut_str, date_fin_str]):
        return JsonResponse({'error': 'Paramètres requis'}, status=400)
    
    try:
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Format date invalide'}, status=400)
    
    try:
        station = Poste.objects.get(pk=station_id, type='pesage')
    except Poste.DoesNotExist:
        return JsonResponse({'error': 'Station non trouvée'}, status=404)
    
    # Vérifier l'accès
    if not check_poste_access(request.user, station):
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    paiements = AmendeEmise.objects.filter(
        station=station,
        statut='paye',
        date_paiement__date__gte=date_debut,
        date_paiement__date__lte=date_fin
    )
    
    agg = paiements.aggregate(total=Sum('montant_amende'), count=Count('id'))
    
    return JsonResponse({
        'montant': float(agg['total'] or 0),
        'count': agg['count'] or 0,
        'station': station.nom
    })


@login_required
@require_GET
def api_paiements_jour(request):
    """
    API pour récupérer les paiements d'un jour.
    
    PERMISSION: peut_voir_recettes_pesage
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_voir_recettes_pesage'):
        logger.warning(
            f"Accès refusé api_paiements_jour | User: {user.username} | "
            f"Permission peut_voir_recettes_pesage: {getattr(user, 'peut_voir_recettes_pesage', False)}"
        )
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    station, _redirect = get_station_context(request, user)
    
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'Date requise'}, status=400)
    
    try:
        date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Format date invalide'}, status=400)
    
    if station is None and user_has_acces_tous_postes(user):
        queryset = AmendeEmise.objects.filter(statut='paye')
        station_filter = request.GET.get('station')
        if station_filter:
            queryset = queryset.filter(station_id=station_filter)
    elif station:
        queryset = AmendeEmise.objects.filter(station=station, statut='paye')
    else:
        queryset = AmendeEmise.objects.none()
    
    paiements = queryset.filter(date_paiement__date=date_cible).select_related('valide_par')
    
    data = {
        'date': date_cible.strftime('%d/%m/%Y'),
        'paiements': [{
            'id': p.pk,
            'numero_ticket': p.numero_ticket,
            'immatriculation': p.immatriculation,
            'transporteur': p.transporteur or '-',
            'montant': float(p.montant_amende),
            'heure': p.date_paiement.strftime('%H:%M') if p.date_paiement else '-',
        } for p in paiements],
        'total': float(paiements.aggregate(Sum('montant_amende'))['montant_amende__sum'] or 0),
        'count': paiements.count(),
    }
    
    return JsonResponse(data)


@login_required
@require_GET
def api_recherche_amende(request):
    """
    API de recherche rapide d'amende.
    
    PERMISSION: peut_lister_amendes
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_lister_amendes'):
        logger.warning(
            f"Accès refusé api_recherche_amende | User: {user.username} | "
            f"Permission peut_lister_amendes: {getattr(user, 'peut_lister_amendes', False)}"
        )
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    station, _redirect = get_station_context(request, user)
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    if station is None and user_has_acces_tous_postes(user):
        amendes = AmendeEmise.objects.all()
    elif station:
        amendes = AmendeEmise.objects.filter(station=station)
    else:
        amendes = AmendeEmise.objects.none()
    
    amendes = amendes.filter(
        Q(numero_ticket__icontains=query) |
        Q(immatriculation__icontains=query) |
        Q(transporteur__icontains=query)
    ).select_related('station')[:10]
    
    results = [{
        'id': a.pk,
        'numero_ticket': a.numero_ticket,
        'immatriculation': a.immatriculation,
        'montant': float(a.montant_amende),
        'statut': a.get_statut_display(),
        'station': a.station.nom,
    } for a in amendes]
    
    return JsonResponse({'results': results})


# ===================================================================
# API STATS JOUR
#
# AVANT: @pesage_access_required (PESAGE_ROLES + admin)
# APRÈS: Permission 'peut_voir_stats_pesage'
# ===================================================================

@login_required
@require_GET
def api_stats_jour(request):
    """
    API pour récupérer les statistiques d'un jour spécifique.
    
    PERMISSION: peut_voir_stats_pesage
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_voir_stats_pesage'):
        logger.warning(
            f"Accès refusé api_stats_jour | User: {user.username} | "
            f"Permission peut_voir_stats_pesage: {getattr(user, 'peut_voir_stats_pesage', False)}"
        )
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    # Récupérer le contexte de station
    station, _ = get_station_context(request, user)
    
    # Paramètres
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'Date requise'}, status=400)
    
    try:
        date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Format date invalide'}, status=400)
    
    # Construire le queryset selon l'accès
    if station is None and user_has_acces_tous_postes(user):
        amendes = AmendeEmise.objects.filter(date_heure_emission__date=date_cible)
        station_filter = request.GET.get('station')
        if station_filter:
            amendes = amendes.filter(station_id=station_filter)
    elif station:
        amendes = AmendeEmise.objects.filter(station=station, date_heure_emission__date=date_cible)
    else:
        amendes = AmendeEmise.objects.none()
    
    # Calculer les statistiques
    stats = amendes.aggregate(
        count=Count('id'),
        hg_count=Count('id', filter=Q(est_hors_gabarit=True)),
        surcharge_count=Count('id', filter=Q(est_surcharge=True)),
        montant_emis=Sum('montant_amende'),
        montant_recouvre=Sum('montant_amende', filter=Q(statut='paye')),
        count_paye=Count('id', filter=Q(statut='paye')),
        count_non_paye=Count('id', filter=Q(statut='non_paye')),
    )
    
    montant_emis = stats['montant_emis'] or Decimal('0')
    montant_recouvre = stats['montant_recouvre'] or Decimal('0')
    reste = montant_emis - montant_recouvre
    taux_recouvrement = 0
    if montant_emis > 0:
        taux_recouvrement = round((float(montant_recouvre) / float(montant_emis)) * 100, 2)
    
    logger.info(
        f"API stats jour | User: {user.username} | "
        f"Date: {date_cible} | Station: {station.nom if station else 'Toutes'} | "
        f"Emissions: {stats['count'] or 0}"
    )
    
    return JsonResponse({
        'date': date_cible.strftime('%d/%m/%Y'),
        'emissions': stats['count'] or 0,
        'hors_gabarit': stats['hg_count'] or 0,
        'surcharge': stats['surcharge_count'] or 0,
        'montant_emis': float(montant_emis),
        'montant_recouvre': float(montant_recouvre),
        'reste_a_recouvrer': float(reste),
        'taux_recouvrement': taux_recouvrement,
        'count_paye': stats['count_paye'] or 0,
        'count_non_paye': stats['count_non_paye'] or 0,
    })


# ===================================================================
# API STATS PÉRIODE
#
# AVANT: Authentifié uniquement
# APRÈS: Permission 'peut_voir_stats_pesage'
# ===================================================================

@login_required
@require_GET
def api_stats_periode(request):
    """
    API pour récupérer les statistiques d'une période.
    
    PERMISSION: peut_voir_stats_pesage
    """
    user = request.user
    
    # Vérifier la permission granulaire
    if not has_permission(user, 'peut_voir_stats_pesage'):
        logger.warning(
            f"Accès refusé api_stats_periode | User: {user.username} | "
            f"Permission peut_voir_stats_pesage: {getattr(user, 'peut_voir_stats_pesage', False)}"
        )
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    # Paramètres
    station_id = request.GET.get('station')
    periode = request.GET.get('periode', 'mois')
    date_str = request.GET.get('date')
    
    # Déterminer la date de référence
    if date_str:
        try:
            date_ref = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date_ref = timezone.now().date()
    else:
        date_ref = timezone.now().date()
    
    # Calculer les dates de la période
    date_debut, date_fin = get_periode_dates(periode, date_ref)
    
    # Construire le queryset
    amendes = AmendeEmise.objects.filter(
        date_heure_emission__date__gte=date_debut,
        date_heure_emission__date__lte=date_fin
    )
    
    station = None
    if station_id:
        try:
            station = Poste.objects.get(pk=station_id, type='pesage')
            amendes = amendes.filter(station=station)
        except Poste.DoesNotExist:
            return JsonResponse({'error': 'Station non trouvée'}, status=404)
    
    # Calculer les statistiques
    stats = amendes.aggregate(
        emissions=Count('id'),
        hors_gabarit=Count('id', filter=Q(est_hors_gabarit=True)),
        surcharge=Count('id', filter=Q(est_surcharge=True)),
        montant_emis=Sum('montant_amende'),
        montant_recouvre=Sum('montant_amende', filter=Q(statut='paye')),
        count_paye=Count('id', filter=Q(statut='paye')),
        count_non_paye=Count('id', filter=Q(statut='non_paye'))
    )
    
    # Pesées de la période
    pesees_qs = PeseesJournalieres.objects.filter(
        date__gte=date_debut, date__lte=date_fin
    )
    if station_id:
        pesees_qs = pesees_qs.filter(station_id=station_id)
    
    nombre_pesees = pesees_qs.aggregate(total=Sum('nombre_pesees'))['total'] or 0
    
    # Calculs finaux
    montant_emis = stats['montant_emis'] or Decimal('0')
    montant_recouvre = stats['montant_recouvre'] or Decimal('0')
    reste = montant_emis - montant_recouvre
    taux_recouvrement = 0
    if montant_emis > 0:
        taux_recouvrement = round((float(montant_recouvre) / float(montant_emis)) * 100, 2)
    
    logger.info(
        f"API stats période | User: {user.username} | "
        f"Période: {periode} ({date_debut} - {date_fin}) | "
        f"Station: {station.nom if station else 'Toutes'}"
    )
    
    return JsonResponse({
        'periode': periode,
        'date_debut': date_debut.strftime('%d/%m/%Y'),
        'date_fin': date_fin.strftime('%d/%m/%Y'),
        'emissions': stats['emissions'] or 0,
        'hors_gabarit': stats['hors_gabarit'] or 0,
        'surcharge': stats['surcharge'] or 0,
        'montant_emis': float(montant_emis),
        'montant_recouvre': float(montant_recouvre),
        'reste_a_recouvrer': float(reste),
        'taux_recouvrement': taux_recouvrement,
        'count_paye': stats['count_paye'] or 0,
        'count_non_paye': stats['count_non_paye'] or 0,
        'nombre_pesees': nombre_pesees,
        'station': station.nom if station else None,
    })